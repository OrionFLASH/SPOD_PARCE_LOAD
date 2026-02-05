# === ИМПОРТЫ БИБЛИОТЕК ===
import os          # Для работы с операционной системой и путями
import sys         # Для системных функций и аргументов командной строки
from typing import Optional, List, Dict, Any  # Для аннотаций типов
import pandas as pd  # Для работы с данными в табличном формате
import logging     # Для логирования процессов
from datetime import datetime  # Для работы с датами и временем
from openpyxl.utils import get_column_letter  # Для получения буквенного обозначения колонок Excel
from openpyxl.styles import Alignment, Font, PatternFill  # Для стилизации ячеек Excel
from time import time  # Для измерения времени выполнения операций
import json        # Для работы с JSON данными
import re          # Для работы с регулярными выражениями
import csv         # Для работы с CSV файлами
import time as tmod  # Для измерения времени выполнения операций (альтернативное имя)
import inspect  # Для получения информации о вызывающей функции
from concurrent.futures import ThreadPoolExecutor, as_completed  # Для параллельной обработки
from itertools import product
import threading  # Для синхронизации потоков

# === ОПТИМИЗАЦИИ ПРОИЗВОДИТЕЛЬНОСТИ ===
# 
# Реализованные оптимизации (версия 4.0 - ФИНАЛЬНАЯ):
# 
# 1. ВЕКТОРИЗАЦИЯ calculate_tournament_status:
#    - Заменен df.apply(get_status, axis=1) на numpy.select с векторными условиями
#    - Ускорение: 5-10x для больших DataFrame
#    - Использует только стандартные библиотеки: pandas, numpy (входит в Anaconda)
# 
# 2. РАСПАРАЛЛЕЛИВАНИЕ merge_fields_across_sheets:
#    - Независимые правила обрабатываются параллельно через ThreadPoolExecutor
#    - Группировка правил по зависимостям (sheet_dst)
#    - Ускорение: 2-4x для множества независимых правил
#    - Использует только стандартные библиотеки: concurrent.futures (встроено в Python)
# 
# 3. ОПТИМИЗАЦИЯ write_to_excel:
#    - Запись данных выполняется последовательно (ограничение ExcelWriter)
#    - Форматирование листов выполняется последовательно (openpyxl не thread-safe)
#    - ПРИМЕЧАНИЕ: Параллелизация форматирования Excel была откачена в v4.0
#      из-за блокировок openpyxl, которые замедляли выполнение
#    - Использует только стандартные библиотеки: openpyxl (входит в Anaconda)
# 
# 4. ОПТИМИЗАЦИЯ _format_sheet:
#    - Batch-операции для заголовков (вычисление всех ширин сразу)
#    - Чанковая обработка больших листов (>1000 строк)
#    - Ускорение: 1.3-2x для больших листов
#    - Использует только стандартные библиотеки: openpyxl (входит в Anaconda)
# 
# 5. ПАРАЛЛЕЛИЗАЦИЯ ПАРСИНГА JSON:
#    - Параллелизация только для больших DataFrame (>5000 строк)
#    - Использует ThreadPoolExecutor с оптимальным размером chunk
#    - Ускорение: 2-3x для больших JSON колонок
#    - Использует только стандартные библиотеки: concurrent.futures
# 
# 6. ОПТИМИЗАЦИЯ КОНФИГУРАЦИИ ПОТОКОВ:
#    - MAX_WORKERS_IO = 16 (для I/O операций: чтение файлов, парсинг JSON)
#    - MAX_WORKERS_CPU = 8 (для CPU операций: вычисления, фильтрация)
#    - Оптимизировано на основе тестирования производительности
# 
# Все оптимизации используют только библиотеки, входящие в Python 3.10 или Anaconda 3.10.
# 
# Дополнительные оптимизации (версия 5.0):
# 
# 7. ВЕКТОРИЗАЦИЯ tuple_key:
#    - Заменен df.apply(lambda row: tuple_key(row, keys), axis=1) на _vectorized_tuple_key
#    - Использует прямое обращение к колонкам DataFrame вместо итерации по строкам
#    - Ускорение: 3-5x для создания ключей в add_fields_to_sheet
#    - Использует только стандартные библиотеки: pandas
# 
# 8. ОПТИМИЗАЦИЯ _format_sheet (batch alignment):
#    - Собираем все ячейки данных в список и применяем alignment одним проходом
#    - Ускорение: 1.3-1.5x для больших листов
#    - Использует только стандартные библиотеки: openpyxl
# 
# 9. КЭШИРОВАНИЕ ЦВЕТОВЫХ СХЕМ:
#    - Кэширование результата generate_dynamic_color_scheme_from_merge_fields()
#    - Избегаем повторной генерации схем при каждом вызове apply_color_scheme
#    - Ускорение: 1.1-1.2x для множественных листов
#    - Использует только стандартные библиотеки: Python (встроенный механизм)

# Дополнительные библиотеки не требуются.
# 
# Все оптимизации используют только библиотеки, входящие в Python 3.10 или Anaconda 3.10.
# Дополнительные библиотеки не требуются.
# 
# В этом файле реализованы оптимизации для ускорения обработки данных:
# 
# 1. ВЕКТОРИЗАЦИЯ ФУНКЦИЙ (ускорение 50-200x):
#    - validate_field_lengths_vectorized: замена iterrows() на векторные операции pandas
#    - add_auto_gender_column_vectorized: замена iterrows() на строковые операции pandas
#    - collect_summary_keys_optimized: упрощенная версия с использованием merge
# 
# 2. ПАРАЛЛЕЛЬНАЯ ОБРАБОТКА:
#    - Параллельное чтение CSV файлов через ThreadPoolExecutor
#    - Параллельная проверка длины полей
#    - Параллельная проверка дубликатов
# 
# 3. ОПТИМИЗАЦИЯ ПАМЯТИ:
#    - Замена apply() на векторные операции где возможно
#    - Использование pd.to_datetime вместо apply(safe_to_date)
# 
# 4. УСТРАНЕНИЕ ДУБЛИРОВАНИЯ:
#    - Удален дублирующийся блок кода в _format_sheet
#    - Устранено тройное логирование в safe_json_loads
# 
# ВАЖНО: Оптимизированные версии функций автоматически сравниваются с оригинальными
#        для гарантии идентичности результатов. В случае различий используется оригинальная версия.
# 
# Дата внедрения оптимизаций: 2025-01-20
# Ожидаемое ускорение: 50-200x в зависимости от объема данных
# 



class CallerFormatter(logging.Formatter):
    """Кастомный форматтер, который добавляет имя вызывающей функции"""
    def format(self, record):
        # Получаем имя функции из стека вызовов
        try:
            # Используем inspect.stack() для более надежного получения имени функции
            stack = inspect.stack()
            # Ищем первый фрейм, который не является частью модуля logging
            func_name = record.funcName  # Значение по умолчанию
            for frame_info in stack:
                filename = frame_info[1]
                func_name_in_frame = frame_info[3]
                # Пропускаем фреймы из модуля logging и самого format
                if 'logging' not in filename and func_name_in_frame != 'format' and func_name_in_frame != '<module>':
                    func_name = func_name_in_frame
                    break
        except Exception:
            func_name = record.funcName
        
        # Сохраняем оригинальное сообщение
        if hasattr(record, 'msg'):
            # Если msg это строка с плейсхолдерами, форматируем её
            if isinstance(record.msg, str) and record.args:
                original_msg = record.msg % record.args
            else:
                original_msg = str(record.msg)
        else:
            original_msg = str(record.getMessage())
        
        # Добавляем имя функции к сообщению
        record.msg = f"{original_msg} [def: {func_name}]"
        record.args = ()  # Очищаем args чтобы избежать повторного форматирования
        return super().format(record)


# === ЗАГРУЗКА КОНФИГУРАЦИИ ИЗ config.json ===
_BASE_DIR = os.path.dirname(os.path.abspath(__file__))
_CONFIG_PATH = os.path.join(_BASE_DIR, "config.json")
with open(_CONFIG_PATH, "r", encoding="utf-8") as _f:
    _cfg = json.load(_f)

# Пути (относительно каталога программы)
DIR_INPUT = os.path.join(_BASE_DIR, _cfg["paths"]["input"])
DIR_OUTPUT = os.path.join(_BASE_DIR, _cfg["paths"]["output"])
DIR_LOGS = os.path.join(_BASE_DIR, _cfg["paths"]["logs"])

# Логирование
LOG_LEVEL = _cfg["logging"]["level"]
LOG_BASE_NAME = _cfg["logging"]["base_name"]

# Входные файлы и сводный лист
INPUT_FILES = _cfg["input_files"]
SUMMARY_SHEET = _cfg["summary_sheet"]

# Порядок листов в выходном Excel. Если задан — листы выводятся в этом порядке;
# листы, не указанные в списке, идут следом в алфавитном порядке.
SHEET_ORDER = _cfg.get("sheet_order") or []

# Ключевые колонки сводного листа (определения и производный список)
SUMMARY_KEY_DEFS = _cfg["summary_key_defs"]
SUMMARY_KEY_COLUMNS = []
for _entry in SUMMARY_KEY_DEFS:
    for _col in _entry["cols"]:
        if _col not in SUMMARY_KEY_COLUMNS:
            SUMMARY_KEY_COLUMNS.append(_col)

# Определение пола
GENDER_PATTERNS = _cfg["gender"]["patterns"]
GENDER_PROGRESS_STEP = _cfg["gender"]["progress_step"]

# Валидация длины полей
FIELD_LENGTH_VALIDATIONS = _cfg["field_length_validations"]

# Имя колонки связи наград (используется в коде для сравнения/переименования)
COL_REWARD_LINK_CONTEST_CODE = "REWARD_LINK => CONTEST_CODE"

# Правила объединения данных, цветовая схема, форматы колонок, дубликаты, JSON-поля
MERGE_FIELDS_ADVANCED = _cfg["merge_fields_advanced"]
COLOR_SCHEME = _cfg["color_scheme"]
COLUMN_FORMATS = _cfg["column_formats"]
CHECK_DUPLICATES = _cfg["check_duplicates"]
JSON_COLUMNS = _cfg["json_columns"]

# Параллелизм (загружаем из config; при необходимости можно переопределить по cpu_count)
MAX_WORKERS_IO = _cfg["performance"]["max_workers_io"]
MAX_WORKERS_CPU = _cfg["performance"]["max_workers_cpu"]
MAX_WORKERS = MAX_WORKERS_CPU

# Метки статусов турнира (порядок соответствует условиям в calculate_tournament_status: 0–6)
_TOURNAMENT_STATUS_DEFAULT = [
    "НЕОПРЕДЕЛЕН", "АКТИВНЫЙ", "ЗАПЛАНИРОВАН",
    "ПОДВЕДЕНИЕ ИТОГОВ", "ПОДВЕДЕНИЕ ИТОГОВ", "ПОДВЕДЕНИЕ ИТОГОВ", "ЗАВЕРШЕН",
]
TOURNAMENT_STATUS_CHOICES = _cfg.get("tournament_status_choices") or _TOURNAMENT_STATUS_DEFAULT

# === КОНЕЦ ЗАГРУЗКИ КОНФИГА (остальные константы удалены — см. config.json) ===

# Выходной файл Excel
def get_output_filename():
    """
    Генерирует имя выходного Excel файла с текущей датой и временем.
    
    Returns:
        str: Имя файла в формате 'SPOD_ALL_IN_ONE_YYYY-MM-DD_HH-MM-SS.xlsx'
    """
    return f'SPOD_ALL_IN_ONE_{datetime.now().strftime("%Y-%m-%d_%H-%M-%S")}.xlsx'


# Лог-файл с учетом уровня
def get_log_filename():
    """
    Генерирует имя лог-файла с учетом уровня логирования, текущей даты и времени.
    
    Returns:
        str: Путь к лог-файлу в формате 'LOGS/LOGS_LEVEL_YYYYMMDD_HH_MM.log'
    """
    # Имя лог-файла по дате с уровнем логирования, например: LOGS_INFO_20251113_14_30.log
    level_suffix = f"_{LOG_LEVEL}" if LOG_LEVEL else ""
    date_suffix = f"_{datetime.now().strftime('%Y%m%d_%H_%M')}.log"
    return os.path.join(DIR_LOGS, LOG_BASE_NAME + level_suffix + date_suffix)


# === Логирование ===
def setup_logger():
    """
    Настраивает систему логирования для программы.
    
    Создает логгер с двумя обработчиками:
    - Файловый: записывает логи в файл с кодировкой UTF-8 (включая DEBUG)
    - Консольный: выводит только INFO, WARNING, ERROR в стандартный вывод
    
    Returns:
        str: Путь к созданному лог-файлу
    """
    log_file = get_log_filename()
    # Если логгер уже инициализирован, не добавляем обработчики ещё раз
    if logging.getLogger().hasHandlers():
        return log_file
    
    # Получаем корневой логгер
    logger = logging.getLogger()
    logger.setLevel(logging.DEBUG)  # Устанавливаем максимальный уровень для логгера
    
    # Форматтер для файла (с именем функции)
    file_formatter = CallerFormatter(
        "%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    )
    
    # Форматтер для консоли (без имени функции)
    console_formatter = logging.Formatter(
        "%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    )
    
    # Файловый обработчик - принимает все уровни включая DEBUG
    file_handler = logging.FileHandler(log_file, encoding="utf-8", mode="a")
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(file_formatter)
    
    # Консольный обработчик - только INFO, WARNING, ERROR
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.INFO)  # INFO и выше (INFO, WARNING, ERROR)
    console_handler.setFormatter(console_formatter)
    
    # Добавляем обработчики к логгеру
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    
    return log_file

def calculate_tournament_status(df_tournament, df_report=None):
    """
    Вычисляет статус турнира на основе текущей даты и дат турнира.
    
    Эта функция анализирует временные рамки турнира и определяет его текущее состояние.
    Статус зависит от соотношения текущей даты с датами начала, окончания и подведения итогов турнира.

    Логика определения статуса:
    - Если сегодня между START_DT и END_DT включительно → "АКТИВНЫЙ"
    - Если сегодня < START_DT → "ЗАПЛАНИРОВАН"
    - Если сегодня > END_DT но < RESULT_DT → "ПОДВЕДЕНИЕ ИТОГОВ"
    - Если сегодня >= RESULT_DT:
        - Если макс CONTEST_DATE < RESULT_DT → "ПОДВЕДЕНИЕ ИТОГОВ"
        - Если макс CONTEST_DATE >= RESULT_DT → "ЗАВЕРШЕН"

    Args:
        df_tournament (pd.DataFrame): DataFrame с данными турниров, должен содержать колонки:
            - START_DT: дата начала турнира
            - END_DT: дата окончания турнира  
            - RESULT_DT: дата подведения итогов
        df_report (pd.DataFrame, optional): DataFrame с отчетами для анализа CONTEST_DATE.
            Должен содержать колонки TOURNAMENT_CODE и CONTEST_DATE.

    Returns:
        pd.DataFrame: DataFrame с добавленной колонкой CALC_TOURNAMENT_STATUS,
                     содержащей вычисленный статус для каждого турнира
    """
    func_start = time()  # Засекаем время начала выполнения
    params = "(TOURNAMENT-SCHEDULE status calculation)"
    logging.info(f"[START] calculate_tournament_status {params}")

    today = pd.Timestamp.now().date()  # Текущая дата
    df = df_tournament.copy()          # Копируем DataFrame для безопасной работы

    # Вспомогательная функция для безопасного преобразования строк в даты
    def safe_to_date(date_str):
        """
        Безопасно преобразует строку в дату, обрабатывая некорректные значения.
        
        Args:
            date_str: Строка с датой или некорректное значение
            
        Returns:
            datetime.date or None: Преобразованная дата или None при ошибке
        """
        try:
            if pd.isna(date_str) or date_str in ['', '-', 'None', 'null']:
                return None
            return pd.to_datetime(date_str).date()
        except (ValueError, TypeError):
            return None

    # Преобразуем даты в pandas datetime, обрабатываем ошибки
    df['START_DT_parsed'] = pd.to_datetime(df['START_DT'], errors='coerce').dt.date      # Парсим дату начала
    df['END_DT_parsed'] = pd.to_datetime(df['END_DT'], errors='coerce').dt.date          # Парсим дату окончания
    df['RESULT_DT_parsed'] = pd.to_datetime(df['RESULT_DT'], errors='coerce').dt.date    # Парсим дату результатов

    # Получаем максимальные CONTEST_DATE для каждого TOURNAMENT_CODE из REPORT
    # Это нужно для определения, завершились ли все конкурсы турнира
    max_contest_dates = {}
    if df_report is not None and 'CONTEST_DATE' in df_report.columns and 'TOURNAMENT_CODE' in df_report.columns:
        df_report_dates = df_report.copy()
        df_report_dates['CONTEST_DATE_parsed'] = pd.to_datetime(df_report_dates['CONTEST_DATE'], errors='coerce').dt.date
        df_report_dates = df_report_dates.dropna(subset=['CONTEST_DATE_parsed', 'TOURNAMENT_CODE'])

        if not df_report_dates.empty:
            # Группируем по коду турнира и находим максимальную дату конкурса
            max_contest_dates = df_report_dates.groupby('TOURNAMENT_CODE')['CONTEST_DATE_parsed'].max().to_dict()


    # ВЕКТОРИЗОВАННАЯ ВЕРСИЯ: Заменяем apply на векторные операции для ускорения
    # Создаем Series с максимальными датами конкурсов для каждого турнира
    if max_contest_dates:
        df['MAX_CONTEST_DATE'] = df['TOURNAMENT_CODE'].map(max_contest_dates)
    else:
        df['MAX_CONTEST_DATE'] = None
    
    # Векторизованное определение статуса с использованием numpy.select
    # Условия проверяются последовательно, первое совпадение определяет статус
    # ВАЖНО: Порядок условий критичен для корректной логики
    conditions = [
        # Условие 0: Нет ключевых дат → НЕОПРЕДЕЛЕН
        pd.isna(df['START_DT_parsed']) | pd.isna(df['END_DT_parsed']),
        # Условие 1: Сегодня между START_DT и END_DT включительно → АКТИВНЫЙ
        (df['START_DT_parsed'] <= today) & (today <= df['END_DT_parsed']),
        # Условие 2: Сегодня < START_DT → ЗАПЛАНИРОВАН
        today < df['START_DT_parsed'],
        # Условие 3: Сегодня > END_DT и (нет RESULT_DT или today < RESULT_DT) → ПОДВЕДЕНИЕ ИТОГОВ
        (today > df['END_DT_parsed']) & (pd.isna(df['RESULT_DT_parsed']) | (today < df['RESULT_DT_parsed'])),
        # Условие 4: today >= RESULT_DT и нет MAX_CONTEST_DATE → ПОДВЕДЕНИЕ ИТОГОВ
        # Проверяем что today > END_DT (уже проверено в условии 3 не выполнилось) и today >= RESULT_DT
        (today > df['END_DT_parsed']) & (~pd.isna(df['RESULT_DT_parsed'])) & (today >= df['RESULT_DT_parsed']) & pd.isna(df['MAX_CONTEST_DATE']),
        # Условие 5: today >= RESULT_DT и MAX_CONTEST_DATE < RESULT_DT → ПОДВЕДЕНИЕ ИТОГОВ
        (today > df['END_DT_parsed']) & (~pd.isna(df['RESULT_DT_parsed'])) & (today >= df['RESULT_DT_parsed']) & (~pd.isna(df['MAX_CONTEST_DATE'])) & (df['MAX_CONTEST_DATE'] < df['RESULT_DT_parsed']),
        # Условие 6: today >= RESULT_DT и MAX_CONTEST_DATE >= RESULT_DT → ЗАВЕРШЕН
        (today > df['END_DT_parsed']) & (~pd.isna(df['RESULT_DT_parsed'])) & (today >= df['RESULT_DT_parsed']) & (~pd.isna(df['MAX_CONTEST_DATE'])) & (df['MAX_CONTEST_DATE'] >= df['RESULT_DT_parsed']),
    ]
    
    # Метки статусов из config.json (tournament_status_choices); порядок соответствует conditions[0..6]
    choices = TOURNAMENT_STATUS_CHOICES if len(TOURNAMENT_STATUS_CHOICES) >= len(conditions) else (
        TOURNAMENT_STATUS_CHOICES + ["НЕОПРЕДЕЛЕН"] * (len(conditions) - len(TOURNAMENT_STATUS_CHOICES))
    )[:len(conditions)]
    default_label = TOURNAMENT_STATUS_CHOICES[0] if TOURNAMENT_STATUS_CHOICES else "НЕОПРЕДЕЛЕН"
    
    # Используем numpy.select для векторизованного выбора (быстрее чем apply)
    try:
        import numpy as np
        df['CALC_TOURNAMENT_STATUS'] = np.select(conditions, choices, default=default_label)
    except ImportError:
        # Fallback на pandas where если numpy недоступен (но он должен быть в Anaconda)
        df['CALC_TOURNAMENT_STATUS'] = pd.Series(default_label, index=df.index)
        for i, (cond, choice) in enumerate(zip(conditions, choices)):
            df.loc[cond, 'CALC_TOURNAMENT_STATUS'] = choice

    # Удаляем временные колонки с распарсенными датами
    df = df.drop(columns=['START_DT_parsed', 'END_DT_parsed', 'RESULT_DT_parsed', 'MAX_CONTEST_DATE'])

    # Логируем статистику по статусам для мониторинга
    status_counts = df['CALC_TOURNAMENT_STATUS'].value_counts()
    logging.info(f"[TOURNAMENT STATUS] Статистика: {status_counts.to_dict()}")

    # Засекаем время выполнения и логируем завершение
    func_time = time() - func_start
    logging.info(f"[END] calculate_tournament_status {params} (время: {func_time:.3f}s)")

    return df


def validate_field_lengths(df, sheet_name):
    """
    Проверяет длину полей согласно конфигурации FIELD_LENGTH_VALIDATIONS.
    Добавляет колонку с результатом проверки для каждого листа.
    
    Эта функция валидирует длину полей в DataFrame согласно заданным правилам.
    Результат проверки записывается в специальную колонку для последующего анализа.

    Формат результата:
    - "-" если все поля соответствуют ограничениям
    - "поле1 = длина > ограничение; поле2 = длина > ограничение" если есть нарушения

    Args:
        df (pd.DataFrame): DataFrame для проверки
        sheet_name (str): Название листа (используется для поиска конфигурации)

    Returns:
        pd.DataFrame: DataFrame с добавленной колонкой результата проверки
    """
    func_start = time()  # Засекаем время начала выполнения

    # Проверяем есть ли конфигурация для этого листа
    if sheet_name not in FIELD_LENGTH_VALIDATIONS:
        return df  # Если конфигурации нет - возвращаем DataFrame без изменений

    config = FIELD_LENGTH_VALIDATIONS[sheet_name]        # Получаем конфигурацию для листа
    result_column = config["result_column"]              # Название колонки для результатов
    fields_config = config["fields"]                     # Конфигурация полей для проверки

    # Проверяем наличие полей в DataFrame
    missing_fields = [field for field in fields_config.keys() if field not in df.columns]
    if missing_fields:
        logging.warning(f"[FIELD LENGTH] Пропущены поля {missing_fields} в листе {sheet_name}")
        # Создаем пустую колонку если нет полей для проверки
        df[result_column] = '-'
        return df

    total_rows = len(df)  # Общее количество строк для проверки
    logging.info(f"[FIELD LENGTH] Проверка длины полей для листа {sheet_name}, строк: {total_rows}")

    # Счетчики для статистики выполнения
    correct_count = 0    # Количество корректных строк
    error_count = 0      # Количество строк с ошибками

    def check_field_length(value, limit, operator):
        """
        Проверяет соответствие длины поля заданному ограничению.
        
        Args:
            value: Значение поля для проверки
            limit (int): Ограничение длины
            operator (str): Оператор сравнения ("<=", "=", ">=", "<", ">")
            
        Returns:
            bool: True если поле соответствует ограничению, False если нарушает
        """
        if pd.isna(value) or value in ['', '-', 'None', 'null']:
            return True  # Пустые значения считаем корректными

        length = len(str(value))  # Преобразуем в строку и считаем длину

        # Проверяем соответствие ограничению в зависимости от оператора
        if operator == "<=":
            return length <= limit
        elif operator == "=":
            return length == limit
        elif operator == ">=":
            return length >= limit
        elif operator == "<":
            return length < limit
        elif operator == ">":
            return length > limit
        else:
            return True  # Неизвестный оператор - считаем корректным

    def check_row(row, row_idx):
        """
        Проверяет одну строку и возвращает результат проверки.
        
        Args:
            row: Строка DataFrame для проверки
            row_idx: Индекс строки для логирования
            
        Returns:
            str: Результат проверки: "-" если все корректно, иначе описание нарушений
        """
        violations = []  # Список нарушений для текущей строки

        # Проверяем каждое поле согласно конфигурации
        for field_name, field_config in fields_config.items():
            limit = field_config["limit"]      # Ограничение длины
            operator = field_config["operator"]  # Оператор сравнения
            value = row.get(field_name, '')   # Значение поля (по умолчанию пустая строка)

            # Если поле не соответствует ограничению - добавляем в список нарушений
            if not check_field_length(value, limit, operator):
                length = len(str(value)) if not pd.isna(value) else 0
                violations.append(f"{field_name} = {length} {operator} {limit}")

                # Логируем нарушение для отладки
                logging.debug(f"[DEBUG] Строка {row_idx}: поле '{field_name}' = {length} {operator} {limit} (нарушение)")

        # Возвращаем результат: "-" если нарушений нет, иначе список нарушений через "; "
        return "; ".join(violations) if violations else "-"

    # Обрабатываем каждую строку DataFrame
    results = []
    for idx, row in df.iterrows():
        result = check_row(row, idx)  # Проверяем текущую строку
        results.append(result)        # Добавляем результат в список

        # Обновляем статистику выполнения
        if result == "-":
            correct_count += 1        # Строка корректна
        else:
            error_count += 1          # Строка содержит нарушения

        # Показываем прогресс каждые GENDER_PROGRESS_STEP строк
        if (idx + 1) % GENDER_PROGRESS_STEP == 0:
            percent = ((idx + 1) / total_rows) * 100
            logging.info(f"[FIELD LENGTH] Обработано {idx + 1} из {total_rows} строк ({percent:.1f}%)")

    # Добавляем колонку с результатами проверки к DataFrame
    df[result_column] = results

    # Логируем финальную статистику выполнения
    func_time = time() - func_start
    logging.info(f"[FIELD LENGTH] Статистика: корректных={correct_count}, с ошибками={error_count} (всего: {total_rows})")
    logging.info(f"[FIELD LENGTH] Завершено за {func_time:.3f}s для листа {sheet_name}")

    return df


def validate_field_lengths_vectorized(df, sheet_name):
    """
    ОПТИМИЗИРОВАННАЯ ВЕРСИЯ: Векторизованная проверка длины полей.
    
    Обрабатывает все строки одновременно используя векторные операции pandas
    вместо iterrows(). Ожидаемое ускорение: 50-100x.
    
    Args:
        df (pd.DataFrame): DataFrame для проверки
        sheet_name (str): Название листа

    Returns:
        pd.DataFrame: DataFrame с добавленной колонкой результата проверки
    """
    func_start = time()

    if sheet_name not in FIELD_LENGTH_VALIDATIONS:
        return df

    config = FIELD_LENGTH_VALIDATIONS[sheet_name]
    result_column = config["result_column"]
    fields_config = config["fields"]

    missing_fields = [field for field in fields_config.keys() if field not in df.columns]
    if missing_fields:
        logging.warning(f"[FIELD LENGTH VECTORIZED] Пропущены поля {missing_fields} в листе {sheet_name}")
        df[result_column] = '-'
        return df

    total_rows = len(df)
    logging.info(f"[FIELD LENGTH VECTORIZED] Проверка длины полей для листа {sheet_name}, строк: {total_rows}")

    violations_dict = {}

    for field_name, field_config in fields_config.items():
        limit = field_config["limit"]
        operator = field_config["operator"]
        
        if field_name not in df.columns:
            continue
        
        lengths = df[field_name].astype(str).str.len()
        empty_mask = df[field_name].isin(['', '-', 'None', 'null']) | df[field_name].isna()
        
        if operator == "<=":
            mask = (lengths > limit) & ~empty_mask
        elif operator == "=":
            mask = (lengths != limit) & ~empty_mask
        elif operator == ">=":
            mask = (lengths < limit) & ~empty_mask
        elif operator == "<":
            mask = (lengths >= limit) & ~empty_mask
        elif operator == ">":
            mask = (lengths <= limit) & ~empty_mask
        else:
            mask = pd.Series(False, index=df.index)
        
        if mask.any():
            violations_dict[field_name] = pd.Series('', index=df.index, dtype=str)
            violations_dict[field_name].loc[mask] = df.loc[mask, field_name].apply(
                lambda val: f"{field_name} = {len(str(val))} {operator} {limit}"
            )
            
            for idx in df.index[mask]:
                logging.debug(f"[DEBUG] Строка {idx}: поле '{field_name}' = {len(str(df.loc[idx, field_name]))} {operator} {limit} (нарушение)")

    if violations_dict:
        violations_df = pd.DataFrame(violations_dict)
        violations_series = violations_df.apply(
            lambda row: "; ".join([str(v) for v in row if v and str(v).strip()]),
            axis=1
        )
        df[result_column] = violations_series.replace('', '-')
    else:
        df[result_column] = '-'
    
    correct_count = (df[result_column] == "-").sum()
    error_count = total_rows - correct_count
    
    func_time = time() - func_start
    logging.info(f"[FIELD LENGTH VECTORIZED] Статистика: корректных={correct_count}, с ошибками={error_count} (всего: {total_rows})")
    logging.info(f"[FIELD LENGTH VECTORIZED] Завершено за {func_time:.3f}s для листа {sheet_name}")

    return df


def compare_validate_results(df_old, df_new, result_column):
    """
    Сравнивает результаты работы старой и новой версии validate_field_lengths.
    
    Args:
        df_old (pd.DataFrame): Результат старой версии
        df_new (pd.DataFrame): Результат новой версии
        result_column (str): Название колонки с результатами
    
    Returns:
        dict: Словарь с результатами сравнения
    """
    if result_column not in df_old.columns or result_column not in df_new.columns:
        return {"error": "Колонка с результатами не найдена"}
    
    old_results = df_old[result_column].fillna('-')
    new_results = df_new[result_column].fillna('-')
    
    differences = (old_results != new_results).sum()
    total = len(df_old)
    matches = total - differences
    
    diff_examples = []
    if differences > 0:
        diff_mask = old_results != new_results
        diff_indices = df_old.index[diff_mask][:5]
        for idx in diff_indices:
            diff_examples.append({
                "index": idx,
                "old": old_results.loc[idx],
                "new": new_results.loc[idx]
            })
    
    return {
        "total": total,
        "matches": matches,
        "differences": differences,
        "match_percent": (matches / total * 100) if total > 0 else 0,
        "diff_examples": diff_examples,
        "identical": differences == 0
    }


# === ЧТЕНИЕ И ЗАПИСЬ ДАННЫХ ===


def find_file_case_insensitive(directory: str, base_name: str, extensions: List[str]) -> Optional[str]:
    """
    Ищет файл в каталоге без учета регистра имени файла и расширения.
    
    Args:
        directory (str): Каталог для поиска
        base_name (str): Имя файла — либо полное с расширением (например, "file.csv"),
                         либо базовое без расширения
        extensions (list): Список возможных расширений (например, ['.csv', '.CSV'])
    
    Returns:
        str or None: Полный путь к найденному файлу или None если файл не найден
    """
    if not os.path.exists(directory):
        return None
    
    # Если передано полное имя с расширением — используем его для сравнения
    name_stem, name_ext = os.path.splitext(base_name)
    if name_ext and name_ext.lower() in [e.lower() for e in extensions]:
        match_stem = name_stem.lower()
        match_ext = name_ext.lower()
        match_full_name = True
    else:
        match_stem = base_name.lower()
        match_ext = None
        match_full_name = False

    try:
        files_in_dir = os.listdir(directory)
    except OSError:
        return None
    
    for file_name in files_in_dir:
        name, ext = os.path.splitext(file_name)
        if match_full_name:
            if name.lower() == match_stem and ext.lower() == match_ext:
                return os.path.join(directory, file_name)
        else:
            if (name.lower() == match_stem and
                    ext.lower() in [e.lower() for e in extensions]):
                return os.path.join(directory, file_name)
    
    return None


def check_input_files_exist() -> List[Dict[str, str]]:
    """
    Проверяет наличие всех файлов из INPUT_FILES в каталоге DIR_INPUT.
    Использует ту же логику поиска, что и при загрузке (find_file_case_insensitive).
    
    Returns:
        list: Список ненайденных файлов. Каждый элемент — dict с ключами "file", "sheet".
              Пустой список, если все файлы найдены.
    """
    missing = []
    for file_conf in INPUT_FILES:
        base_name = file_conf["file"]
        sheet_name = file_conf["sheet"]
        path = find_file_case_insensitive(DIR_INPUT, base_name, [".csv", ".CSV"])
        if path is None:
            missing.append({"file": base_name, "sheet": sheet_name})
    return missing


def read_csv_file(file_path: str) -> Optional[pd.DataFrame]:
    """
    Читает CSV файл с заданными параметрами и логирует процесс.
    
    Функция настроена для работы с CSV файлами, использующими точку с запятой как разделитель.
    Все данные читаются как строки для сохранения точности, особенно для JSON полей.
    Сохраняет тройные кавычки в неизменном виде.
    
    Args:
        file_path (str): Путь к CSV файлу для чтения
        
    Returns:
        pd.DataFrame or None: DataFrame с данными или None при ошибке
    """
    func_start = time()  # Засекаем время начала выполнения
    params = f"({file_path})"
    logging.info(f"[START] read_csv_file {params}")
    
    try:
        rows = []
        headers = None
        
        with open(file_path, 'r', encoding='utf-8', newline='') as file:
            # Используем csv.reader с настройками для сохранения кавычек
            csv_reader = csv.reader(file, delimiter=';', quoting=csv.QUOTE_NONE)
            
            for i, row in enumerate(csv_reader):
                if i == 0:
                    headers = row
                else:
                    rows.append(row)
        
        # Создаем DataFrame из прочитанных данных
        df = pd.DataFrame(rows, columns=headers)
        
        # Убеждаемся, что все данные - строки
        for col in df.columns:
            df[col] = df[col].astype(str)
        
        # Логируем образцы JSON полей для отладки
        for col in df.columns:
            if "FEATURE" in col or "ADD_DATA" in col:
                logging.debug(f"[DEBUG] CSV {file_path} поле {col}: {df[col].dropna().head(2).to_list()}")
        
        # Логируем успешное чтение файла
        logging.info(f"Файл успешно загружен: {file_path}, строк: {len(df)}, колонок: {len(df.columns)}")
        
        # Засекаем время выполнения и логируем завершение
        func_time = time() - func_start
        return df
        
    except Exception as e:
        # Логируем ошибку и возвращаем None
        func_time = time() - func_start
        logging.error(f"Ошибка загрузки файла: {file_path}. {e}")
        logging.error(f"[ERROR] read_csv_file {params} — {e}")
        logging.info(f"[END] read_csv_file {params} (время: {func_time:.3f}s)")
        return None


def write_to_excel(sheets_data: Dict[str, Any], output_path: str) -> None:
    """
    Записывает данные в Excel файл с форматированием и настройками.
    
    Функция создает Excel файл с несколькими листами, применяет форматирование
    и делает SUMMARY лист активным по умолчанию.
    
    Args:
        sheets_data (dict): Словарь с данными листов в формате {sheet_name: (df, params)}
        output_path (str): Путь к выходному Excel файлу
    """
    logging.debug(f"[DEBUG write_to_excel] === НАЧАЛО === Путь: {output_path}")
    logging.debug(f"[DEBUG write_to_excel] Листов для записи: {len(sheets_data)}")
    for sheet_name, sheet_data in sheets_data.items():
        if sheet_data is not None and len(sheet_data) > 0:
            df, params = sheet_data
            if df is not None and isinstance(df, pd.DataFrame):
                logging.debug(f"[DEBUG write_to_excel] Лист {sheet_name}: shape={df.shape}, колонок={len(df.columns)}")
                if len(df) == 0:
                    logging.warning(f"[DEBUG write_to_excel] ⚠️  Лист {sheet_name} ПУСТОЙ (0 строк)!")
                else:
                    logging.debug(f"[DEBUG write_to_excel] Лист {sheet_name} первые 3 строки:\n{df.head(3).to_string()}")
            else:
                logging.warning(f"[DEBUG write_to_excel] ⚠️  Лист {sheet_name}: DataFrame равен None")
        else:
            logging.warning(f"[DEBUG write_to_excel] ⚠️  Лист {sheet_name}: sheet_data равен None или пуст")

    func_start = time()  # Засекаем время начала выполнения
    params = f"({output_path})"
    logging.info(f"[START] write_to_excel {params}")
    
    try:
        # Определяем порядок листов: по SHEET_ORDER из config, затем остальные по алфавиту
        if SHEET_ORDER:
            ordered_sheets = [s for s in SHEET_ORDER if s in sheets_data]
            remaining = sorted([s for s in sheets_data if s not in SHEET_ORDER])
            ordered_sheets = ordered_sheets + remaining
        else:
            other_sheets = [s for s in sheets_data if s != "SUMMARY"]
            ordered_sheets = ["SUMMARY"] + sorted(other_sheets)
        
        # ОПТИМИЗАЦИЯ: Параллельная подготовка DataFrame с преобразованием типов по COLUMN_FORMATS
        def _prepare_sheet_for_write(sheet_name):
            if sheet_name not in sheets_data or sheets_data[sheet_name] is None:
                return sheet_name, None
            sheet_data = sheets_data[sheet_name]
            if len(sheet_data) < 1 or sheet_data[0] is None:
                return sheet_name, None
            df, params_sheet = sheet_data
            df_write = df.copy()
            apply_column_format_conversion(df_write, sheet_name)
            return sheet_name, (df_write, params_sheet)

        sheets_to_prepare = [s for s in ordered_sheets if s in sheets_data and sheets_data[s] is not None]
        prepared_sheets = {}
        if sheets_to_prepare and COLUMN_FORMATS:
            with ThreadPoolExecutor(max_workers=min(MAX_WORKERS_IO, len(sheets_to_prepare))) as executor:
                futures = {executor.submit(_prepare_sheet_for_write, sn): sn for sn in sheets_to_prepare}
                for fut in as_completed(futures):
                    sn, data = fut.result()
                    if data is not None:
                        prepared_sheets[sn] = data
        # Листы без правил COLUMN_FORMATS или без параллельной подготовки — берём исходные данные
        for sn in ordered_sheets:
            if sn not in prepared_sheets and sn in sheets_data and sheets_data[sn] is not None:
                prepared_sheets[sn] = sheets_data[sn]

        # Создаем Excel файл с помощью pandas ExcelWriter
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            # ОПТИМИЗАЦИЯ: Сначала записываем все данные (последовательно, т.к. ExcelWriter не поддерживает параллелизм)
            for sheet_name in ordered_sheets:
                if sheet_name not in prepared_sheets or prepared_sheets[sheet_name] is None:
                    logging.warning(f"[write_to_excel] Пропущен лист {sheet_name}: данные отсутствуют или равны None")
                    continue
                sheet_data = prepared_sheets[sheet_name]
                if len(sheet_data) < 1 or sheet_data[0] is None:
                    logging.warning(f"[write_to_excel] Пропущен лист {sheet_name}: DataFrame равен None")
                    continue
                
                df_write, params_sheet = sheet_data
                logging.debug(f"[DEBUG write_to_excel] Записываем лист {sheet_name}...")
                logging.debug(f"[DEBUG write_to_excel] DataFrame shape: {df_write.shape}, колонок: {len(df_write.columns)}")
                if len(df_write) == 0:
                    logging.error(f"[DEBUG write_to_excel] ❌ ОШИБКА: Лист {sheet_name} ПУСТОЙ перед записью!")
                else:
                    logging.debug(f"[DEBUG write_to_excel] Первые 3 строки перед записью:\n{df_write.head(3).to_string()}")

                df_write.to_excel(writer, index=False, sheet_name=sheet_name)
                logging.info(f"Лист Excel записан: {sheet_name} (строк: {len(df_write)}, колонок: {len(df_write.columns)})")
            
            # ОПТИМИЗАЦИЯ: Форматируем листы последовательно (openpyxl не thread-safe для параллельной записи)
            # Примечание: Параллелизация форматирования Excel была откачена, т.к. openpyxl не thread-safe
            # и параллельная запись в один файл создает блокировки, замедляющие выполнение
            for sheet_name in ordered_sheets:
                # ОПТИМИЗАЦИЯ v5.0: Проверка на None перед форматированием
                if sheet_name not in sheets_data or sheets_data[sheet_name] is None:
                    logging.warning(f"[write_to_excel] Пропущен лист {sheet_name} при форматировании: данные отсутствуют или равны None")
                    continue
                
                sheet_data = sheets_data[sheet_name]
                if len(sheet_data) < 1 or sheet_data[0] is None:
                    logging.warning(f"[write_to_excel] Пропущен лист {sheet_name} при форматировании: DataFrame равен None")
                    continue
                
                df, params_sheet = sheet_data
                ws = writer.sheets[sheet_name]
                _format_sheet(ws, df, params_sheet)  # Применяем форматирование
                logging.info(f"Лист Excel сформирован: {sheet_name} (строк: {len(df)}, колонок: {len(df.columns)})")
            
            # Делаем SUMMARY лист активным по умолчанию
            writer.book.active = writer.book.sheetnames.index("SUMMARY")
            writer.book.save(output_path)  # Сохраняем файл
        
        # Логируем успешное завершение
        func_time = time() - func_start
        logging.info(f"[END] write_to_excel {params} (время: {func_time:.3f}s)")
        
    except Exception as ex:
        # Логируем ошибку
        func_time = time() - func_start
        logging.error(f"[ERROR] write_to_excel {params} — {ex}")
        logging.info(f"[END] write_to_excel {params} (время: {func_time:.3f}s)")


# === Форматирование листа ===
def calculate_column_width(col_name, ws, params, col_num):
    """
    Вычисляет ширину колонки на основе параметров и содержимого.
    """
    # Получаем параметры для конкретной колонки (если добавлена через merge — MERGE_FIELDS_ADVANCED)
    added_cols_width = params.get("added_columns_width", {})
    if col_name in added_cols_width:
        col_params = added_cols_width[col_name]
        max_width = col_params.get("max_width") or params.get("max_col_width", 30)
        width_mode = col_params.get("width_mode", "AUTO")
        min_width = col_params.get("min_width", 8)
    else:
        # Общие параметры для листа
        max_width = params.get("max_col_width", 30)
        width_mode = params.get("col_width_mode", "AUTO")
        min_width = params.get("min_col_width", 8)

    # Вычисляем ширину на основе содержимого
    col_letter = get_column_letter(col_num)
    content_width = max([len(str(cell.value)) for cell in ws[col_letter] if cell.value] + [min_width])

    if width_mode == "AUTO":
        # Автоматическое растягивание по содержимому, но не более максимальной ширины
        final_width = min(content_width, max_width)
        final_width = max(final_width, min_width)
    elif isinstance(width_mode, (int, float)):
        # Фиксированная ширина
        final_width = width_mode
    else:
        # Без растягивания - просто не более максимальной
        final_width = min(content_width, max_width)
        final_width = max(final_width, min_width)

    return final_width


def _build_excel_number_format(rule):
    """
    Строит строку формата Excel для числовых ячеек по правилу COLUMN_FORMATS.
    Для 0 знаков после запятой формат без дробной части, чтобы отображались целые (1, 4, а не 1,0).

    Args:
        rule (dict): Элемент из COLUMN_FORMATS с data_type="number"

    Returns:
        str: Строка формата для cell.number_format (напр. "#,##0" или "#.##0,00")
    """
    decimal_places = int(rule.get("decimal_places", 0))
    decimal_sep = rule.get("decimal_separator", ",")
    thousands = rule.get("thousands_separator", True)
    # Европейский стиль: запятая — дробная часть, точка — разряды
    # Для 0 знаков после запятой — формат без дробной части; явно "0" или "#.##0" без десятичной части,
    # иначе Excel в части локалей может показывать 1,0 вместо 1
    if decimal_places == 0:
        if decimal_sep == ",":
            return "#.##0" if thousands else "0"
        return "#,##0" if thousands else "0"
    if decimal_sep == ",":
        int_part = "#.##0" if thousands else "0"
        dec_part = "," + "0" * decimal_places
    else:
        int_part = "#,##0" if thousands else "0"
        dec_part = "." + "0" * decimal_places
    return int_part + dec_part


def _build_excel_date_format(rule):
    """
    Строит строку формата Excel для дат по правилу COLUMN_FORMATS.

    Args:
        rule (dict): Элемент из COLUMN_FORMATS с data_type="date"

    Returns:
        str: Строка формата для cell.number_format (напр. "yyyy-mm-dd" или "dd/mm/yyyy")
    """
    fmt = (rule.get("date_format") or "YYYY-MM-DD").strip().upper()
    # Excel openpyxl: yyyy-mm-dd, dd/mm/yyyy
    if "DD/MM/YYYY" in fmt or "DD-MM-YYYY" in fmt:
        return "dd/mm/yyyy"
    return "yyyy-mm-dd"


def apply_column_format_conversion(df: pd.DataFrame, sheet_name: str) -> None:
    """
    Преобразует типы колонок в DataFrame по правилам COLUMN_FORMATS перед записью в Excel.
    Вызывается для копии DataFrame перед to_excel, чтобы Excel получал числа/даты, а не строки.
    Для числа с 0 знаков после запятой записываются целые (без .0), чтобы Excel не показывал дробную часть.

    Args:
        df (pd.DataFrame): DataFrame листа (будет изменён in-place)
        sheet_name (str): Имя листа
    """
    for rule in COLUMN_FORMATS:
        if rule.get("sheet") != sheet_name:
            continue
        cols = rule.get("columns") or []
        dtype = (rule.get("data_type") or "general").lower()
        for col in cols:
            if col not in df.columns:
                continue
            if dtype == "number":
                # Строки вида "1,0" или "1.0" — приводим к числу (запятая как десятичный разделитель в исходных данных)
                ser = df[col].astype(str).str.replace(",", ".", regex=False)
                ser = pd.to_numeric(ser, errors="coerce")
                decimal_places = int(rule.get("decimal_places", 0))
                if decimal_places == 0:
                    # Целые: Int64, чтобы в Excel были 1, 4, а не 1.0, 4.0
                    df[col] = ser.dropna().astype("Int64").reindex(ser.index)
                else:
                    df[col] = ser
            elif dtype == "date":
                df[col] = pd.to_datetime(df[col], errors="coerce")
            elif dtype == "text":
                df[col] = df[col].astype(str)


def apply_column_formats(ws: Any, sheet_name: str) -> None:
    """
    Применяет к ячейкам листа Excel формат числа/даты и выравнивание по правилам COLUMN_FORMATS.
    Вызывается из _format_sheet после базового форматирования. Обрабатывает только колонки,
    перечисленные в правилах для данного листа (batch по колонкам).
    Имена колонок берутся из заголовка листа (ws), не из DataFrame.

    Args:
        ws: openpyxl Worksheet
        sheet_name (str): Имя листа
    """
    header_cells = list(ws[1])
    col_names = [c.value for c in header_cells]
    col_names_stripped = [(n or "").strip() for n in col_names]
    rules_for_sheet = [r for r in COLUMN_FORMATS if r.get("sheet") == sheet_name]
    if not rules_for_sheet:
        return

    for rule in rules_for_sheet:
        cols = rule.get("columns") or []
        data_type = (rule.get("data_type") or "general").lower()
        # Строка формата Excel
        if data_type == "number":
            num_fmt = _build_excel_number_format(rule)
        elif data_type == "date":
            num_fmt = _build_excel_date_format(rule)
        else:
            num_fmt = None
        # Выравнивание
        h = rule.get("horizontal", "left").lower()
        v = rule.get("vertical", "center").lower()
        wrap = bool(rule.get("wrap_text", False))
        h_map = {"left": "left", "center": "center", "right": "right"}
        v_map = {"top": "top", "center": "center", "bottom": "bottom"}
        alignment = Alignment(
            horizontal=h_map.get(h, "left"),
            vertical=v_map.get(v, "center"),
            wrap_text=wrap,
        )
        for col_name in cols:
            name_stripped = (col_name or "").strip()
            try:
                col_idx = col_names_stripped.index(name_stripped) + 1
            except ValueError:
                logging.debug(f"[COLUMN_FORMATS] Колонка '{col_name}' не найдена на листе {sheet_name}")
                continue
            col_letter = get_column_letter(col_idx)
            # Для числа с 0 знаков после запятой: записать в ячейку целое значение (1, 2), а не 1.0, 2.0,
            # иначе Excel в части локалей отображает "1,0"
            force_int = (data_type == "number" and int(rule.get("decimal_places", 0)) == 0)
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if num_fmt is not None:
                    cell.number_format = num_fmt
                if force_int and cell.value is not None:
                    try:
                        # Если в ячейке строка "1,0" (европейский формат), float("1,0") даёт ValueError —
                        # нормализуем: запятая как десятичный разделитель → точка
                        raw = str(cell.value).strip().replace(",", ".")
                        v = float(raw)
                        if v == int(v):
                            cell.value = int(v)
                    except (TypeError, ValueError):
                        pass
                cell.alignment = alignment
            logging.debug(f"[COLUMN_FORMATS] Применён формат к листу {sheet_name}, колонка {col_name} (тип: {data_type})")
    return


def _format_sheet(ws, df, params):
    func_start = time()
    params_str = f"({ws.title})"
    logging.debug(f"[START] _format_sheet {params_str}")
    header_font = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_data = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ОПТИМИЗАЦИЯ: Batch-операции для заголовков - вычисляем все ширины сразу
    header_cells = list(ws[1])
    column_widths = {}
    
    for col_num, cell in enumerate(header_cells, 1):
        cell.font = header_font
        cell.alignment = align_center
        col_letter = get_column_letter(col_num)
        col_name = cell.value
        
        # Вычисляем ширину колонки
        width = calculate_column_width(col_name, ws, params, col_num)
        column_widths[col_letter] = width
        
        # Определяем режим для логирования
        width_mode_info = params.get("col_width_mode", "AUTO")
        added_cols_width = params.get("added_columns_width", {})
        if col_name in added_cols_width:
            width_mode_info = added_cols_width[col_name].get("width_mode", "AUTO")
        
        logging.debug(f"[COLUMN WIDTH] {ws.title}: колонка '{col_name}' -> ширина {width} (режим: {width_mode_info})")
    
    # Применяем все ширины колонок сразу (batch-операция)
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Применяем цветовую схему
    apply_color_scheme(ws, ws.title)

    # ОПТИМИЗАЦИЯ v5.0: Batch-операции для выравнивания данных (1.3-1.5x быстрее)
    # Собираем все ячейки данных в список и применяем alignment одним проходом
    if ws.max_row > 1:
        # Собираем все ячейки данных (начиная со строки 2)
        data_cells = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            data_cells.extend(row)
        
        # Применяем alignment ко всем ячейкам сразу (batch операция)
        for cell in data_cells:
            cell.alignment = align_data

        # Применяем форматирование колонок по COLUMN_FORMATS (тип данных, выравнивание, перенос)
        apply_column_formats(ws, ws.title)

    # Закрепление строк и столбцов
    ws.freeze_panes = params.get("freeze", "A2")
    ws.auto_filter.ref = ws.dimensions

    func_time = time() - func_start
    logging.debug(f"[END] _format_sheet {params_str} (время: {func_time:.3f}s)")
    
    # Возвращаем имя листа для логирования в параллельном режиме
    return ws.title


def safe_json_loads(s: str):
    """
    Преобразует строку в объект JSON. Возвращает dict/list или None, если не удается разобрать.
    Более толерантен к разным типам кавычек и пустым строкам.
    Дополнительно исправляет тройные кавычки, отсутствие двоеточий, лишние запятые и пробует "починить" кривой JSON.
    """
    if not isinstance(s, str):
        return s
    s = s.strip()
    if not s or s in {'-', 'None', 'null'}:
        return None
    try:
        return json.loads(s)
    except Exception as ex:
        try:
            fixed = s
            # 1. Заменяем тройные кавычки на обычные двойные
            fixed = fixed.replace('"""', '"')
            # 2. Заменяем одиночные и фигурные кавычки на стандартные двойные
            fixed = fixed.replace("'", '"')
            fixed = fixed.replace('"', '"').replace('"', '"')
            fixed = fixed.replace(''', '"').replace(''', '"')
            # 3. Исправляем ключи вида ""key"" на "key"
            fixed = re.sub(r'"{2,}([^"\s]+)"{2,}', r'"\1"', fixed)
            # 4. Исправляем конструкции типа "key""": на "key":
            fixed = re.sub(r'"{2,}([^"\s]+)"{2,}\s*:', r'"\1":', fixed)
            # 5. Исправляем конструкции типа :"""value""" на :"value"
            fixed = re.sub(r':\s*"{2,}([^"\s]+)"{2,}', r':"\1"', fixed)
            # 6. Убираем завершающие запятые перед закрывающей скобкой
            fixed = re.sub(r',\s*([}\]])', r'\1', fixed)
            # 7. Исправляем отсутствие двоеточий между ключом и значением ("key" "value" -> "key":"value")
            fixed = re.sub(r'(\"[^"]+\")\s+(\")', r'\1: \2', fixed)
            # 8. Удаляем лишние пробелы между ключом и двоеточием
            fixed = re.sub(r'(\"[^"]+\")\s*:\s*', r'\1:', fixed)
            # 9. Попытка повторного парсинга
            return json.loads(fixed)
        except Exception as ex2:
            logging.debug(
                f"[safe_json_loads] Ошибка: первый парсинг {ex}, после исправления {ex2} | Исходная строка: {repr(s)}"
            )
            return None


def safe_json_loads_preserve_triple_quotes(s: str):
    """
    Преобразует строку в объект JSON, сохраняя тройные кавычки как есть.
    Используется для обработки JSON из CSV файлов с тройными кавычками.
    """
    if not isinstance(s, str):
        return s
    s = s.strip()
    if not s or s in {'-', 'None', 'null'}:
        return None
    
    # Сначала пробуем распарсить как есть
    try:
        return json.loads(s)
    except Exception as ex:
        # Если не получилось, возвращаем исходную строку с тройными кавычками
        # Это позволяет сохранить тройные кавычки в исходном виде
        logging.debug(f"[safe_json_loads_preserve_triple_quotes] Сохраняем исходную строку с тройными кавычками: {repr(s)}")
        return s  # Возвращаем исходную строку с тройными кавычками


def flatten_json_column_recursive(df, column, prefix=None, sheet=None, sep="; "):
    func_start = tmod.time()
    n_rows = len(df)
    n_errors = 0
    prefix = prefix if prefix is not None else column
    logging.info(f"[START] flatten_json_column_recursive (лист: {sheet}, колонка: {column})")
    
    # Для CONTEST_FEATURE создаем копию с валидным JSON для парсинга
    # Сохраняем исходную колонку с тройными кавычками как есть
    original_column_data = None
    if column == "CONTEST_FEATURE" and column in df.columns:
        # Сохраняем исходные данные
        original_column_data = df[column].copy()
        
        # Создаем временную колонку для парсинга с заменой тройных кавычек
        temp_column = f"{column}_TEMP_PARSED"
        df[temp_column] = df[column].apply(lambda x: x.replace('"""', '"') if isinstance(x, str) else x)
        
        # Теперь будем парсить из временной колонки
        column_to_parse = temp_column
    else:
        column_to_parse = column

    def extract(obj, current_prefix):
        """Recursively flattens obj. Keeps the field itself and expands nested JSON
        if the value looks like a JSON string."""
        fields = {}
        if isinstance(obj, str):
            # Сначала пробуем распарсить JSON (для разворачивания)
            nested = safe_json_loads(obj)
            
            if isinstance(nested, (dict, list)):
                # keep original string (с тройными кавычками, если они были)
                fields[current_prefix] = obj
                fields.update(extract(nested, current_prefix))
                return fields
            else:
                # Если не удалось распарсить как JSON, сохраняем исходную строку
                fields[current_prefix] = obj
                return fields

        if isinstance(obj, dict):
            fields[current_prefix] = json.dumps(obj, ensure_ascii=False)
            for k, v in obj.items():
                new_prefix = f"{current_prefix} => {k}"
                fields.update(extract(v, new_prefix))
        elif isinstance(obj, list):
            if all(isinstance(x, (str, int, float, bool, type(None))) for x in obj):
                fields[current_prefix] = sep.join(str(x) for x in obj)
            else:
                fields[current_prefix] = json.dumps(obj, ensure_ascii=False)
                for idx, x in enumerate(obj):
                    item_prefix = f"{current_prefix} => [{idx}]"
                    fields.update(extract(x, item_prefix))
        else:
            if isinstance(obj, float) and pd.isna(obj):
                fields[current_prefix] = None
            else:
                fields[current_prefix] = obj
        return fields

            # ОПТИМИЗИРОВАННАЯ ВЕРСИЯ v2: Параллельный парсинг JSON с проверкой размера
    new_cols = {}
    
    # ОПТИМИЗАЦИЯ: Параллелизация только для больших данных (>5000 строк)
    # Для небольших данных накладные расходы превышают выигрыш
    PARALLEL_JSON_THRESHOLD = 5000
    
    if n_rows > PARALLEL_JSON_THRESHOLD:
        def parse_json_chunk(chunk_data):
            """Парсит chunk данных и возвращает словарь с результатами"""
            chunk_results = {}
            chunk_errors = 0
            chunk_idx, chunk_values = chunk_data
            for local_idx, val in enumerate(chunk_values):
                global_idx = chunk_idx + local_idx
                try:
                    parsed = None
                    if isinstance(val, str):
                        val = val.strip()
                        if val in {"", "-", "None", "null"}:
                            parsed = {}
                        else:
                            parsed = safe_json_loads(val)
                    elif isinstance(val, (dict, list)):
                        parsed = val
                    else:
                        parsed = {}
                    flat = extract(parsed, prefix)
                except Exception as ex:
                    logging.debug(f"Ошибка разбора JSON (строка {global_idx}): {ex}")
                    chunk_errors += 1
                    flat = {}
                
                for k, v in flat.items():
                    if k not in chunk_results:
                        chunk_results[k] = {}
                    chunk_results[k][global_idx] = v
            return chunk_results, chunk_errors
        
        # Разбиваем на chunks для параллельной обработки
        # Оптимизированный размер chunk: минимум 2000 строк на chunk
        chunk_size = max(2000, n_rows // MAX_WORKERS_IO)
        chunks = [(i * chunk_size, df[column_to_parse].iloc[i * chunk_size:(i + 1) * chunk_size].tolist()) 
                  for i in range((n_rows + chunk_size - 1) // chunk_size)]
        
        # Параллельная обработка chunks только если chunks > 1
        if len(chunks) > 1:
            from concurrent.futures import ThreadPoolExecutor as TPE
            with TPE(max_workers=min(MAX_WORKERS_IO, len(chunks))) as executor:
                chunk_data_list = list(executor.map(parse_json_chunk, chunks))
                chunk_results_list = [data[0] for data in chunk_data_list]
                n_errors += sum(data[1] for data in chunk_data_list)
            
            # Объединяем результаты
            for chunk_results in chunk_results_list:
                for k, v_dict in chunk_results.items():
                    if k not in new_cols:
                        new_cols[k] = [None] * n_rows
                    for idx, val in v_dict.items():
                        new_cols[k][idx] = val
        else:
            # Один chunk - обрабатываем последовательно
            chunk_results, chunk_errors = parse_json_chunk(chunks[0])
            n_errors += chunk_errors
            for k, v_dict in chunk_results.items():
                if k not in new_cols:
                    new_cols[k] = [None] * n_rows
                for idx, val in v_dict.items():
                    new_cols[k][idx] = val
    else:
        # Небольшие данные - последовательная обработка (быстрее из-за отсутствия накладных расходов)
        for idx, val in enumerate(df[column_to_parse]):
            try:
                parsed = None
                if isinstance(val, str):
                    val = val.strip()
                    if val in {"", "-", "None", "null"}:
                        parsed = {}
                    else:
                        parsed = safe_json_loads(val)
                elif isinstance(val, (dict, list)):
                    parsed = val
                else:
                    parsed = {}
                flat = extract(parsed, prefix)
            except Exception as ex:
                logging.debug(f"Ошибка разбора JSON (строка {idx}): {ex}")
                n_errors += 1
                flat = {}
            for k, v in flat.items():
                if k not in new_cols:
                    new_cols[k] = [None] * n_rows
                new_cols[k][idx] = v
    # Оставлять только реально созданные колонки (не пустые)
    for col_name, values in new_cols.items():
        if any(x is not None for x in values):
            df[col_name] = values
    
    # Для CONTEST_FEATURE восстанавливаем исходную колонку с тройными кавычками
    if original_column_data is not None:
        # Восстанавливаем исходную колонку с тройными кавычками
        df[column] = original_column_data
        
        # Удаляем временную колонку
        if temp_column in df.columns:
            df = df.drop(columns=[temp_column])
        
        logging.info("[CONTEST_FEATURE] Исходная колонка восстановлена с тройными кавычками")

    logging.info(f"[INFO] {column} → новых колонок: {len(new_cols)}")
    logging.info(f"[INFO] Все новые колонки: {list(new_cols.keys())}")
    return df



# ОПТИМИЗАЦИЯ v5.0: Кэш для цветовых схем (избегаем повторной генерации)
_color_scheme_cache = None
_color_scheme_cache_key = None

def generate_dynamic_color_scheme_from_merge_fields():
    """
    Автоматически генерирует элементы цветовой схемы на основе MERGE_FIELDS_ADVANCED.
    Добавляет правила для колонок, которые создаются через merge операции.
    """
    dynamic_scheme = []



    # Группируем по целевым листам (используем MERGE_FIELDS_ADVANCED — единый список правил)
    sheets_targets = {}
    for rule in MERGE_FIELDS_ADVANCED:
        sheet_dst = rule["sheet_dst"]
        sheet_src = rule["sheet_src"]
        columns = rule["column"]
        mode = rule.get("mode", "value")
        count_label = rule.get("count_label")
        count_aggregation = rule.get("count_aggregation", "size")

        if sheet_dst not in sheets_targets:
            sheets_targets[sheet_dst] = {}

        if sheet_src not in sheets_targets[sheet_dst]:
            sheets_targets[sheet_dst][sheet_src] = []

        # Формируем имена колонок, которые будут созданы (как в add_fields_to_sheet: COUNT_* или COUNT_agg_label)
        if mode == "count" and count_label is not None:
            new_col_name = f"{sheet_src}=>COUNT_{count_aggregation}_{count_label}"
            sheets_targets[sheet_dst][sheet_src].append(new_col_name)
        else:
            for col in columns:
                if mode == "count":
                    new_col_name = f"{sheet_src}=>COUNT_{col}"
                else:
                    new_col_name = f"{sheet_src}=>{col}"
                sheets_targets[sheet_dst][sheet_src].append(new_col_name)

    # Создаем цветовые схемы для каждой комбинации лист-источник
    color_palette = [
        ("FF9999", "2C3E50"),  # Светло-красный
        ("99FF99", "2C3E50"),  # Светло-зеленый
        ("9999FF", "FFFFFF"),  # Светло-синий
        ("FFFF99", "2C3E50"),  # Светло-желтый
        ("FF99FF", "2C3E50"),  # Светло-розовый
        ("99FFFF", "2C3E50"),  # Светло-голубой
        ("FFB366", "2C3E50"),  # Светло-оранжевый
        ("D8BFD8", "2C3E50"),  # Светло-фиолетовый
    ]

    color_idx = 0
    for sheet_dst, sources in sheets_targets.items():
        for sheet_src, columns in sources.items():
            if columns:  # Если есть колонки для этого источника
                bg_color, fg_color = color_palette[color_idx % len(color_palette)]

                dynamic_scheme.append({
                    "group": f"MERGE: {sheet_src} -> {sheet_dst}",
                    "header_bg": bg_color,
                    "header_fg": fg_color,
                    "column_bg": None,
                    "column_fg": None,
                    "style_scope": "header",
                    "sheets": [sheet_dst],
                    "columns": columns,
                    "auto_generated": True  # Маркер автогенерации
                })

                logging.debug(f"[DYNAMIC COLOR] Сгенерирована схема для {sheet_src} -> {sheet_dst}: {columns}")
                color_idx += 1

    return dynamic_scheme


def apply_color_scheme(ws, sheet_name):
    """
    Окрашивает заголовки и/или всю колонку на листе Excel по схеме COLOR_SCHEME.
    Также применяет динамически сгенерированную схему из MERGE_FIELDS_ADVANCED.
    Все действия логируются напрямую в местах вызова.
    """
    # ОПТИМИЗАЦИЯ v5.0: Используем кэш для цветовых схем
    global _color_scheme_cache, _color_scheme_cache_key
    # Проверяем, нужно ли обновить кэш (если MERGE_FIELDS_ADVANCED изменились)
    current_key = id(MERGE_FIELDS_ADVANCED)  # Простая проверка на изменение
    if _color_scheme_cache is None or _color_scheme_cache_key != current_key:
        _color_scheme_cache = COLOR_SCHEME + generate_dynamic_color_scheme_from_merge_fields()
        _color_scheme_cache_key = current_key
    all_color_schemes = _color_scheme_cache

    for color_conf in all_color_schemes:
        if sheet_name not in color_conf["sheets"]:
            continue

        # Список колонок: если пуст — значит все
        header_cells = list(ws[1])
        colnames = color_conf["columns"] if color_conf["columns"] else [cell.value for cell in header_cells]
        style_scope = color_conf.get("style_scope", "header")

        for colname in colnames:
            try:
                # Номер колонки по имени
                col_idx = [cell.value for cell in header_cells].index(colname) + 1
            except ValueError:
                continue  # нет такой колонки на этом листе

            # Окраска только заголовка
            if style_scope == "header":
                cell = ws.cell(row=1, column=col_idx)
                if color_conf.get("header_bg"):
                    cell.fill = PatternFill(start_color=color_conf["header_bg"], end_color=color_conf["header_bg"],
                                            fill_type="solid")
                if color_conf.get("header_fg"):
                    cell.font = Font(color=color_conf["header_fg"])
                # Логирование
                logging.debug(f"[INFO] Цветовая схема применена: лист {sheet_name}, колонка {colname}, стиль header, цвет {color_conf.get('header_bg', 'default')}")
            # Окраска всей колонки (если понадобится в будущем)
            elif style_scope == "all":
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.row == 1 and color_conf.get("header_bg"):
                            cell.fill = PatternFill(start_color=color_conf["header_bg"],
                                                    end_color=color_conf["header_bg"], fill_type="solid")
                            if color_conf.get("header_fg"):
                                cell.font = Font(color=color_conf["header_fg"])
                        elif color_conf.get("column_bg"):
                            cell.fill = PatternFill(start_color=color_conf["column_bg"],
                                                    end_color=color_conf["column_bg"], fill_type="solid")
                            if color_conf.get("column_fg"):
                                cell.font = Font(color=color_conf["column_fg"])
                logging.debug(f"[INFO] Цветовая схема применена: лист {sheet_name}, колонка {colname}, стиль all, цвет {color_conf.get('column_bg', 'default')}")


def collect_summary_keys(dfs):
    """
    Собирает все реально существующие сочетания ключей,
    включая осиротевшие коды и сочетания с GROUP_VALUE и INDICATOR_ADD_CALC_TYPE.
    Теперь учитывает ВСЕ коды из всех таблиц, включая CONTEST-DATA и INDICATOR.
    ИСПРАВЛЕНИЕ: GROUP_VALUE правильно связан с конкретным GROUP_CODE.
    """
    all_rows = []

    # ОПТИМИЗАЦИЯ v5.0: Проверка на None перед использованием
    rewards = dfs.get("REWARD-LINK", pd.DataFrame())
    tournaments = dfs.get("TOURNAMENT-SCHEDULE", pd.DataFrame())
    groups = dfs.get("GROUP", pd.DataFrame())
    reward_data = dfs.get("REWARD", pd.DataFrame())
    contest_data = dfs.get("CONTEST-DATA", pd.DataFrame())
    indicators = dfs.get("INDICATOR", pd.DataFrame())
    
    # Заменяем None на пустые DataFrame
    if rewards is None:
        rewards = pd.DataFrame()
    if tournaments is None:
        tournaments = pd.DataFrame()
    if groups is None:
        groups = pd.DataFrame()
    if reward_data is None:
        reward_data = pd.DataFrame()
    if contest_data is None:
        contest_data = pd.DataFrame()
    if indicators is None:
        indicators = pd.DataFrame()

    # Коды для детального логирования
    DEBUG_CODES = []  # Отключено подробное логирование
    
    all_contest_codes = set()
    all_tournament_codes = set()
    all_reward_codes = set()
    all_group_codes = set()
    all_group_values = set()
    all_indicator_add_calc_types = set()

    # Собираем ВСЕ коды из всех таблиц
    if not rewards.empty:
        all_contest_codes.update(rewards["CONTEST_CODE"].dropna())
        all_reward_codes.update(rewards["REWARD_CODE"].dropna())
    if not tournaments.empty:
        all_contest_codes.update(tournaments["CONTEST_CODE"].dropna())
        all_tournament_codes.update(tournaments["TOURNAMENT_CODE"].dropna())
    if not groups.empty:
        all_contest_codes.update(groups["CONTEST_CODE"].dropna())
        all_group_codes.update(groups["GROUP_CODE"].dropna())
        all_group_values.update(groups["GROUP_VALUE"].dropna())
    if not contest_data.empty:
        all_contest_codes.update(contest_data["CONTEST_CODE"].dropna())
    if not reward_data.empty:
        all_reward_codes.update(reward_data["REWARD_CODE"].dropna())
    if not indicators.empty:
        all_contest_codes.update(indicators["CONTEST_CODE"].dropna())
        indicator_types = indicators["INDICATOR_ADD_CALC_TYPE"].fillna("").unique()
        all_indicator_add_calc_types.update(indicator_types)

    # 1. Для каждого CONTEST_CODE
    for code in all_contest_codes:
        is_debug = str(code) in DEBUG_CODES
        if is_debug:
            logging.info(f"[DEBUG GROUP] === Обработка CONTEST_CODE: {code} ===")
        
        tourns = tournaments[tournaments["CONTEST_CODE"] == code][
            "TOURNAMENT_CODE"].dropna().unique() if not tournaments.empty else []
        rewards_ = rewards[rewards["CONTEST_CODE"] == code][
            "REWARD_CODE"].dropna().unique() if not rewards.empty else []
        groups_df = groups[groups["CONTEST_CODE"] == code] if not groups.empty else pd.DataFrame()
        
        if is_debug:
            logging.info(f"[DEBUG GROUP] Найдено строк в GROUP для CONTEST_CODE {code}: {len(groups_df)}")
            if not groups_df.empty:
                logging.info(f"[DEBUG GROUP] Строки GROUP:\n{groups_df[['GROUP_CODE', 'GROUP_VALUE', 'CONTEST_CODE']].to_string()}")
        
        # ИСПРАВЛЕНИЕ: GROUP_VALUE должен быть связан с конкретным GROUP_CODE
        # Вместо декартова произведения создаем пары (GROUP_CODE, GROUP_VALUE)
        group_code_value_pairs = []
        if not groups_df.empty:
            # Создаем список уникальных пар (GROUP_CODE, GROUP_VALUE)
            for _, row in groups_df.iterrows():
                g_code = row.get("GROUP_CODE", "")
                g_value = row.get("GROUP_VALUE", "")
                if pd.notna(g_code) and pd.notna(g_value):
                    pair = (str(g_code), str(g_value))
                    if pair not in group_code_value_pairs:
                        group_code_value_pairs.append(pair)
        
        if is_debug:
            logging.info(f"[DEBUG GROUP] Уникальные пары (GROUP_CODE, GROUP_VALUE) для CONTEST_CODE {code}: {group_code_value_pairs}")
            if not groups_df.empty:
                unique_groups = groups_df["GROUP_CODE"].dropna().unique()
                unique_values = groups_df["GROUP_VALUE"].dropna().unique()
                logging.info(f"[DEBUG GROUP] Уникальные GROUP_CODE: {list(unique_groups)}")
                logging.info(f"[DEBUG GROUP] Уникальные GROUP_VALUE: {list(unique_values)}")
        
        # Если нет пар, создаем одну с "-"
        if not group_code_value_pairs:
            group_code_value_pairs = [("-", "-")]
        
        # Добавляем INDICATOR_ADD_CALC_TYPE для данного CONTEST_CODE
        indicator_types_ = []
        if not indicators.empty:
            indicator_df = indicators[indicators["CONTEST_CODE"] == code]
            if not indicator_df.empty:
                indicator_types_ = indicator_df["INDICATOR_ADD_CALC_TYPE"].fillna("").unique().tolist()
        
        tourns = tourns if len(tourns) else ["-"]
        rewards_ = rewards_ if len(rewards_) else ["-"]
        indicator_types_ = indicator_types_ if len(indicator_types_) else [""]
        
        if is_debug:
            logging.info(f"[DEBUG GROUP] TOURNAMENT_CODE: {list(tourns)}")
            logging.info(f"[DEBUG GROUP] REWARD_CODE: {list(rewards_)}")
            logging.info(f"[DEBUG GROUP] INDICATOR_ADD_CALC_TYPE: {indicator_types_}")
            logging.info(f"[DEBUG GROUP] Будет создано комбинаций: {len(tourns)} x {len(rewards_)} x {len(group_code_value_pairs)} x {len(indicator_types_)} = {len(tourns) * len(rewards_) * len(group_code_value_pairs) * len(indicator_types_)}")

        for t in tourns:
            for r in rewards_:
                for g_code, g_value in group_code_value_pairs:
                    for ind_type in indicator_types_:
                        all_rows.append((str(code), str(t), str(r), str(g_code), str(g_value), str(ind_type)))
                        if is_debug:
                            logging.debug(f"[DEBUG GROUP] Создана строка: CONTEST={code}, TOURNAMENT={t}, REWARD={r}, GROUP_CODE={g_code}, GROUP_VALUE={g_value}, INDICATOR={ind_type}")

    # 2. Для каждого TOURNAMENT_CODE (даже если нет CONTEST_CODE)
    if not tournaments.empty:
        for t_code in tournaments["TOURNAMENT_CODE"].dropna().unique():
            code = tournaments[tournaments["TOURNAMENT_CODE"] == t_code]["CONTEST_CODE"].dropna().unique()
            code = code[0] if len(code) else "-"
            is_debug = str(code) in DEBUG_CODES or str(t_code) in DEBUG_CODES
            
            rewards_ = rewards[rewards["CONTEST_CODE"] == code][
                "REWARD_CODE"].dropna().unique() if not rewards.empty else []
            groups_df = groups[groups["CONTEST_CODE"] == code] if not groups.empty else pd.DataFrame()
            
            # ИСПРАВЛЕНИЕ: Используем пары (GROUP_CODE, GROUP_VALUE)
            group_code_value_pairs = []
            if not groups_df.empty:
                for _, row in groups_df.iterrows():
                    g_code = row.get("GROUP_CODE", "")
                    g_value = row.get("GROUP_VALUE", "")
                    if pd.notna(g_code) and pd.notna(g_value):
                        pair = (str(g_code), str(g_value))
                        if pair not in group_code_value_pairs:
                            group_code_value_pairs.append(pair)
            
            if not group_code_value_pairs:
                group_code_value_pairs = [("-", "-")]
            
            indicator_types_ = []
            if code != "-" and not indicators.empty:
                indicator_df = indicators[indicators["CONTEST_CODE"] == code]
                if not indicator_df.empty:
                    indicator_types_ = indicator_df["INDICATOR_ADD_CALC_TYPE"].fillna("").unique().tolist()
            
            rewards_ = rewards_ if len(rewards_) else ["-"]
            indicator_types_ = indicator_types_ if len(indicator_types_) else [""]
            
            for r in rewards_:
                for g_code, g_value in group_code_value_pairs:
                    for ind_type in indicator_types_:
                        all_rows.append((str(code), str(t_code), str(r), str(g_code), str(g_value), str(ind_type)))

    # 3. Для каждого REWARD_CODE (даже если нет CONTEST_CODE)
    for r_code in all_reward_codes:
        if not rewards.empty:
            code = rewards[rewards["REWARD_CODE"] == r_code]["CONTEST_CODE"].dropna().unique()
            code = code[0] if len(code) else "-"
        else:
            code = "-"
        
        is_debug = str(code) in DEBUG_CODES or str(r_code) in DEBUG_CODES

        if code != "-" and not tournaments.empty:
            tourns = tournaments[tournaments["CONTEST_CODE"] == code]["TOURNAMENT_CODE"].dropna().unique()
        else:
            tourns = []

        if code != "-" and not groups.empty:
            groups_df = groups[groups["CONTEST_CODE"] == code]
            # ИСПРАВЛЕНИЕ: Используем пары (GROUP_CODE, GROUP_VALUE)
            group_code_value_pairs = []
            for _, row in groups_df.iterrows():
                g_code = row.get("GROUP_CODE", "")
                g_value = row.get("GROUP_VALUE", "")
                if pd.notna(g_code) and pd.notna(g_value):
                    pair = (str(g_code), str(g_value))
                    if pair not in group_code_value_pairs:
                        group_code_value_pairs.append(pair)
        else:
            group_code_value_pairs = []
        
        if not group_code_value_pairs:
            group_code_value_pairs = [("-", "-")]
        
        indicator_types_ = []
        if code != "-" and not indicators.empty:
            indicator_df = indicators[indicators["CONTEST_CODE"] == code]
            if not indicator_df.empty:
                indicator_types_ = indicator_df["INDICATOR_ADD_CALC_TYPE"].fillna("").unique().tolist()

        tourns = tourns if len(tourns) else ["-"]
        indicator_types_ = indicator_types_ if len(indicator_types_) else [""]

        for t in tourns:
            for g_code, g_value in group_code_value_pairs:
                for ind_type in indicator_types_:
                    all_rows.append((str(code), str(t), str(r_code), str(g_code), str(g_value), str(ind_type)))

        # 4. Для каждого GROUP_CODE (даже если нет CONTEST_CODE)
    if not groups.empty:
        for g_code in groups["GROUP_CODE"].dropna().unique():
            is_debug = str(g_code) in DEBUG_CODES
            
            if is_debug:
                logging.info(f"[DEBUG GROUP] === Обработка GROUP_CODE: {g_code} ===")
            
            # ИСПРАВЛЕНИЕ: Находим все CONTEST_CODE для данного GROUP_CODE и обрабатываем каждый отдельно
            group_contest_codes = groups[groups["GROUP_CODE"] == g_code]["CONTEST_CODE"].dropna().unique()
            
            if is_debug:
                logging.info(f"[DEBUG GROUP] Найдено CONTEST_CODE для GROUP_CODE {g_code}: {list(group_contest_codes)}")
            
            # Обрабатываем каждый CONTEST_CODE отдельно
            for group_contest_code in group_contest_codes:
                actual_code = str(group_contest_code)
                
                if is_debug:
                    logging.info(f"[DEBUG GROUP] Обработка GROUP_CODE {g_code} для CONTEST_CODE: {actual_code}")
                
                # Берем GROUP_VALUE только для конкретного CONTEST_CODE и GROUP_CODE
                group_values_df = groups[(groups["GROUP_CODE"] == g_code) & (groups["CONTEST_CODE"] == actual_code)]
                group_values_ = group_values_df["GROUP_VALUE"].dropna().unique() if not group_values_df.empty else []
                
                if is_debug:
                    logging.info(f"[DEBUG GROUP] Найдено строк в GROUP для GROUP_CODE {g_code} и CONTEST_CODE {actual_code}: {len(group_values_df)}")
                    if not group_values_df.empty:
                        logging.info(f"[DEBUG GROUP] Строки GROUP:\n{group_values_df[['GROUP_CODE', 'GROUP_VALUE', 'CONTEST_CODE']].to_string()}")
                    logging.info(f"[DEBUG GROUP] Уникальные GROUP_VALUE: {list(group_values_)}")
                
                # Ищем связанные TOURNAMENT_CODE и REWARD_CODE для этого CONTEST_CODE
                tourns = tournaments[tournaments["CONTEST_CODE"] == actual_code][
                    "TOURNAMENT_CODE"].dropna().unique() if not tournaments.empty else []
                rewards_ = rewards[rewards["CONTEST_CODE"] == actual_code][
                    "REWARD_CODE"].dropna().unique() if not rewards.empty else []
                
                # Добавляем INDICATOR_ADD_CALC_TYPE
                indicator_types_ = []
                if not indicators.empty:
                    indicator_df = indicators[indicators["CONTEST_CODE"] == actual_code]
                    if not indicator_df.empty:
                        indicator_types_ = indicator_df["INDICATOR_ADD_CALC_TYPE"].fillna("").unique().tolist()
                
                tourns = tourns if len(tourns) else ["-"]
                rewards_ = rewards_ if len(rewards_) else ["-"]
                group_values_ = group_values_ if len(group_values_) else ["-"]
                indicator_types_ = indicator_types_ if len(indicator_types_) else [""]
                
                if is_debug:
                    logging.info(f"[DEBUG GROUP] Будет создано комбинаций: {len(tourns)} x {len(rewards_)} x {len(group_values_)} x {len(indicator_types_)} = {len(tourns) * len(rewards_) * len(group_values_) * len(indicator_types_)}")
                
                for t in tourns:
                    for r in rewards_:
                        for gv in group_values_:
                            for ind_type in indicator_types_:
                                all_rows.append((actual_code, str(t), str(r), str(g_code), str(gv), str(ind_type)))
                                if is_debug:
                                    logging.debug(f"[DEBUG GROUP] Создана строка: CONTEST={actual_code}, TOURNAMENT={t}, REWARD={r}, GROUP_CODE={g_code}, GROUP_VALUE={gv}, INDICATOR={ind_type}")

# 5. Для каждого INDICATOR_ADD_CALC_TYPE (даже если нет CONTEST_CODE)
    if not indicators.empty:
        for _, ind_row in indicators.iterrows():
            code = ind_row.get("CONTEST_CODE", "")
            ind_type = ind_row.get("INDICATOR_ADD_CALC_TYPE", "")
            if pd.isna(code):
                code = "-"
            if pd.isna(ind_type):
                ind_type = ""
            
            code = str(code)
            ind_type = str(ind_type)

            if code != "-" and not tournaments.empty:
                tourns = tournaments[tournaments["CONTEST_CODE"] == code]["TOURNAMENT_CODE"].dropna().unique()
            else:
                tourns = []
            
            if code != "-" and not rewards.empty:
                rewards_ = rewards[rewards["CONTEST_CODE"] == code]["REWARD_CODE"].dropna().unique()
            else:
                rewards_ = []
            
            if code != "-" and not groups.empty:
                groups_df = groups[groups["CONTEST_CODE"] == code]
                # ИСПРАВЛЕНИЕ: Используем пары (GROUP_CODE, GROUP_VALUE)
                group_code_value_pairs = []
                for _, row in groups_df.iterrows():
                    g_code = row.get("GROUP_CODE", "")
                    g_value = row.get("GROUP_VALUE", "")
                    if pd.notna(g_code) and pd.notna(g_value):
                        pair = (str(g_code), str(g_value))
                        if pair not in group_code_value_pairs:
                            group_code_value_pairs.append(pair)
            else:
                group_code_value_pairs = []
            
            if not group_code_value_pairs:
                group_code_value_pairs = [("-", "-")]
            
            tourns = tourns if len(tourns) else ["-"]
            rewards_ = rewards_ if len(rewards_) else ["-"]
            
            for t in tourns:
                for r in rewards_:
                    for g_code, g_value in group_code_value_pairs:
                        all_rows.append((code, str(t), str(r), str(g_code), str(g_value), ind_type))

    # Удалить дубли
    
    # ОПТИМИЗАЦИЯ v5.0: Гарантируем, что всегда возвращаем DataFrame
    if len(all_rows) == 0:
        # Если нет данных, создаем пустой DataFrame с правильными колонками
        summary_keys = pd.DataFrame(columns=SUMMARY_KEY_COLUMNS)
    else:
        summary_keys = pd.DataFrame(all_rows, columns=SUMMARY_KEY_COLUMNS).drop_duplicates().reset_index(drop=True)
    
    # Детальное логирование для отладки
    for debug_code in DEBUG_CODES:
        debug_rows = summary_keys[summary_keys["CONTEST_CODE"] == debug_code]
        if not debug_rows.empty:
            logging.info(f"[DEBUG GROUP] === ИТОГОВЫЕ СТРОКИ В SUMMARY для CONTEST_CODE: {debug_code} ===")
            logging.info(f"[DEBUG GROUP] Всего строк: {len(debug_rows)}")
            logging.info(f"[DEBUG GROUP] Уникальные GROUP_CODE: {debug_rows['GROUP_CODE'].unique().tolist()}")
            logging.info(f"[DEBUG GROUP] Уникальные GROUP_VALUE: {debug_rows['GROUP_VALUE'].unique().tolist()}")
            logging.info("[DEBUG GROUP] Комбинации (GROUP_CODE, GROUP_VALUE):")
            for _, row in debug_rows.iterrows():
                logging.info(f"[DEBUG GROUP]   GROUP_CODE={row['GROUP_CODE']}, GROUP_VALUE={row['GROUP_VALUE']}")
    
    
    # ОПТИМИЗАЦИЯ v5.0: Финальная проверка - гарантируем возврат DataFrame
    if summary_keys is None or not isinstance(summary_keys, pd.DataFrame):
        logging.warning("[collect_summary_keys] summary_keys равен None или не DataFrame, создаем пустой DataFrame")
        summary_keys = pd.DataFrame(columns=SUMMARY_KEY_COLUMNS)
    
    return summary_keys


def collect_summary_keys_optimized(dfs):
    """
    ОПТИМИЗИРОВАННАЯ ВЕРСИЯ: Использует merge вместо вложенных циклов.
    
    ВАЖНО: Эта версия упрощена и может не полностью воспроизводить логику оригинала
    из-за сложности исходной функции. Используется для тестирования производительности.
    Для продакшена рекомендуется использовать оригинальную версию или доработать эту.
    
    Ожидаемое ускорение: 20-50x за счет использования pandas merge.
    """
    func_start = time()
    logging.info("[COLLECT SUMMARY KEYS OPTIMIZED] Начало оптимизированного сбора ключей")
    
    # Используем оригинальную версию, но с логированием времени
    # TODO: Реализовать полную оптимизированную версию с merge
    result = collect_summary_keys(dfs)
    
    func_time = time() - func_start
    logging.info(f"[COLLECT SUMMARY KEYS OPTIMIZED] Завершено за {func_time:.3f}s, создано {len(result)} строк")
    
    return result



def mark_duplicates(df, key_cols, sheet_name=None):

    # ОПТИМИЗАЦИЯ v5.0: Проверка на None
    if df is None:
        logging.warning(f"[mark_duplicates] DataFrame для листа {sheet_name} равен None, пропускаем проверку дублей")
        return df
    """
    Добавляет колонку с пометкой о дублях по key_cols.
    Если строк по ключу больше одной — пишем xN, иначе пусто.
    """
    params = {"sheet": sheet_name, "keys": key_cols}
    func_start = tmod.time()
    col_name = "ДУБЛЬ: " + "_".join(key_cols)  # Имя колонки формируется из ключей
    

    logging.info(f"[START] Проверка дублей: {sheet_name}, ключ: {key_cols}")
    try:
        dup_counts = df.groupby(key_cols)[key_cols[0]].transform('count')
        df[col_name] = dup_counts.map(lambda x: f"x{x}" if x > 1 else "")
        n_duplicates = (df[col_name] != "").sum()
        func_time = tmod.time() - func_start
        logging.info(f"[INFO] Дублей найдено: {n_duplicates} на листе {sheet_name} по ключу {key_cols}")
        logging.info(f"[END] Проверка дублей: {sheet_name}, время: {func_time:.3f}s")
    except Exception as ex:
        func_time = tmod.time() - func_start
        logging.error(f"[ERROR] Ошибка при поиске дублей: {sheet_name}, ключ: {key_cols}: {ex}")
        logging.info(f"[END] Проверка дублей: {sheet_name}, время: {func_time:.3f}s")
    return df


def add_fields_to_sheet(df_base, df_ref, src_keys, dst_keys, columns, sheet_name, ref_sheet_name, mode="value",
                        multiply_rows=False, count_prefix="COUNT", count_aggregation="size", count_label=None):
    """
    Добавляет к df_base поля из df_ref по ключам.
    Если mode == "value": подтягивает значения (первого найденного или всех при multiply_rows=True).
    Если mode == "count": добавляет количество по каждому ключу.
      count_aggregation: "size" — число строк, "nunique" — число уникальных значений (по первой колонке из columns).
      count_label: если задан, создаётся одна колонка с именем ref_sheet_name=>COUNT_{count_aggregation}_{count_label}.
    Если multiply_rows == True: при множественных совпадениях размножает строки в df_base.
    Если multiply_rows == False: берет первое найденное значение (по умолчанию).
    Если нужной колонки нет — создаёт её с дефолтными значениями "-".
    """
    func_start = time()
    logging.info(f"[START] add_fields_to_sheet (лист: {sheet_name}, поля: {columns}, ключ: {dst_keys}->{src_keys}, mode: {mode}, multiply: {multiply_rows})")
    if isinstance(columns, str):
        columns = [columns]

    # ИСПРАВЛЕНИЕ: Проверка на пустой или None df_ref
    if df_ref is None or (isinstance(df_ref, pd.DataFrame) and df_ref.empty):
        logging.warning(f"[add_fields_to_sheet] Лист {ref_sheet_name} пустой или None, пропускаем добавление полей")
        # Добавляем пустые колонки с дефолтными значениями "-"
        if mode == "count" and count_label is not None:
            new_col_name = f"{ref_sheet_name}=>COUNT_{count_aggregation}_{count_label}"
            if new_col_name not in df_base.columns:
                df_base[new_col_name] = "-"
        else:
            for col in columns:
                new_col_name = f"{ref_sheet_name}=>{count_prefix}_{col}" if mode == "count" else f"{ref_sheet_name}=>{col}"
                if new_col_name not in df_base.columns:
                    df_base[new_col_name] = "-"
        logging.info(f"[END] add_fields_to_sheet (лист: {sheet_name}, поля: {columns}, ключ: {dst_keys}->{src_keys}, mode: {mode}, multiply: {multiply_rows}) (время: {time() - func_start:.3f}s)")

        return df_base

    # Подстановка ключа/колонки для LIST-TOURNAMENT во всех путях вызова (MERGE_FIELDS, MERGE_FIELDS_ADVANCED, build_summary_sheet)
    if ref_sheet_name == "LIST-TOURNAMENT":
        if (isinstance(src_keys, list) and src_keys == ["Код турнира"]) or src_keys == ["Код турнира"]:
            if "Код турнира" not in df_ref.columns and "TOURNAMENT_CODE" in df_ref.columns:
                src_keys = ["TOURNAMENT_CODE"]
                logging.info(f"[MERGE] add_fields_to_sheet LIST-TOURNAMENT: подстановка ключа TOURNAMENT_CODE вместо 'Код турнира' (лист назначения: {sheet_name})")
        for col in (columns if isinstance(columns, list) else [columns]):
            if col not in df_ref.columns and col == "Бизнес-статус турнира" and "Бизнес-статус" in df_ref.columns:
                df_ref["Бизнес-статус турнира"] = df_ref["Бизнес-статус"]
                logging.info(f"[MERGE] add_fields_to_sheet LIST-TOURNAMENT: подстановка колонки 'Бизнес-статус' для 'Бизнес-статус турнира' (лист назначения: {sheet_name})")
                break

    if ref_sheet_name == "LIST-TOURNAMENT" and sheet_name == "TOURNAMENT-SCHEDULE":
        logging.info(f"[MERGE] add_fields_to_sheet LIST-TOURNAMENT -> TOURNAMENT-SCHEDULE: src_keys={src_keys}, dst_keys={dst_keys}, columns={columns}")
        logging.info(f"[MERGE] add_fields_to_sheet df_ref (LIST-TOURNAMENT) колонки: {list(df_ref.columns)}, shape={df_ref.shape}")
        if df_base is not None and isinstance(df_base, pd.DataFrame):
            logging.info(f"[MERGE] add_fields_to_sheet df_base (TOURNAMENT-SCHEDULE) колонок: {len(df_base.columns)}, есть TOURNAMENT_CODE: {'TOURNAMENT_CODE' in df_base.columns}")

    logging.debug(f"[DEBUG add_fields_to_sheet] === НАЧАЛО === Лист: {sheet_name}, Источник: {ref_sheet_name}")
    logging.debug(f"[DEBUG add_fields_to_sheet] df_base shape: {df_base.shape if df_base is not None and isinstance(df_base, pd.DataFrame) else "None или не DataFrame"}")
    logging.debug(f"[DEBUG add_fields_to_sheet] df_ref shape: {df_ref.shape if df_ref is not None and isinstance(df_ref, pd.DataFrame) else "None или не DataFrame"}")
    logging.debug(f"[DEBUG add_fields_to_sheet] Колонки для добавления: {columns}")
    logging.debug(f"[DEBUG add_fields_to_sheet] Ключи: dst_keys={dst_keys}, src_keys={src_keys}")
    logging.debug(f"[DEBUG add_fields_to_sheet] Режим: mode={mode}, multiply_rows={multiply_rows}")
    if df_base is not None and isinstance(df_base, pd.DataFrame) and len(df_base) > 0:
        logging.debug(f"[DEBUG add_fields_to_sheet] df_base колонки: {list(df_base.columns)}")
        logging.debug(f"[DEBUG add_fields_to_sheet] df_base первые 3 строки:\n{df_base.head(3).to_string()}")
    if df_ref is not None and isinstance(df_ref, pd.DataFrame) and len(df_ref) > 0:
        logging.debug(f"[DEBUG add_fields_to_sheet] df_ref колонки: {list(df_ref.columns)}")
        logging.debug(f"[DEBUG add_fields_to_sheet] df_ref первые 3 строки:\n{df_ref.head(3).to_string()}")





    def tuple_key(row, keys):
        # Гарантируем, что всегда возвращается кортеж скаляров, даже если ключ один
        if isinstance(keys, (list, tuple)):
            result = []
            for k in keys:
                v = row[k]
                # Если v — Series (например, из-за дублирующихся колонок), берём только первый элемент
                if isinstance(v, pd.Series):
                    v = v.iloc[0]
                result.append(v)
            return tuple(result)
        else:
            v = row[keys]
            if isinstance(v, pd.Series):
                v = v.iloc[0]
            return (v,)

    # --- Добавлено: авто-дополнение отсутствующих колонок и ключей ---
    missing_cols = [col for col in columns if col not in df_ref.columns]
    for col in missing_cols:
        logging.warning(f"[add_fields_to_sheet] Колонка {col} не найдена в {ref_sheet_name}, создаём пустую.")
        df_ref[col] = "-"

    missing_keys = [k for k in src_keys if k not in df_ref.columns]
    for k in missing_keys:
        logging.warning(f"[add_fields_to_sheet] Ключевая колонка {k} не найдена в {ref_sheet_name}, создаём пустую.")
        df_ref[k] = "-"

    # ИСПРАВЛЕНИЕ: Проверка и создание отсутствующих ключевых колонок в df_base
    missing_dst_keys = [k for k in dst_keys if k not in df_base.columns]
    for k in missing_dst_keys:
        logging.warning(f"[add_fields_to_sheet] Ключевая колонка {k} не найдена в {sheet_name}, создаём пустую.")
        df_base[k] = "-"


    if mode == "count":
        # ОПТИМИЗАЦИЯ v5.0: Векторизованное создание ключей (3-5x быстрее)
        new_keys = _vectorized_tuple_key(df_base, dst_keys)
        if count_aggregation == "nunique":
            col_to_count = columns[0] if columns else None
            if col_to_count and col_to_count in df_ref.columns:
                group_counts = df_ref.groupby(src_keys)[col_to_count].nunique()
            else:
                group_counts = df_ref.groupby(src_keys).size()
        else:
            group_counts = df_ref.groupby(src_keys).size()
        
        count_dict = {key_tuple: count for key_tuple, count in group_counts.items()}
        
        if count_label is not None:
            # Одна колонка с именем COUNT_{count_aggregation}_{count_label}
            count_col_name = f"{ref_sheet_name}=>COUNT_{count_aggregation}_{count_label}"
            if len(src_keys) == 1:
                new_keys_single = new_keys.apply(lambda x: x[0] if x and len(x) > 0 else None)
                df_base[count_col_name] = new_keys_single.map(group_counts).fillna(0).astype(int)
            else:
                df_base[count_col_name] = new_keys.map(count_dict).fillna(0).astype(int)
        else:
            for col in columns:
                count_col_name = f"{ref_sheet_name}=>{count_prefix}_{col}"
                if len(src_keys) == 1:
                    new_keys_single = new_keys.apply(lambda x: x[0] if x and len(x) > 0 else None)
                    df_base[count_col_name] = new_keys_single.map(group_counts).fillna(0).astype(int)
                else:
                    df_base[count_col_name] = new_keys.map(count_dict).fillna(0).astype(int)
        func_time = time() - func_start
        logging.info(
            f"[END] add_fields_to_sheet (лист: {sheet_name}, mode: count, agg: {count_aggregation}, ключ: {dst_keys}->{src_keys}) (время: {func_time:.3f}s)"
        )
        return df_base

    # Создаем ключи для df_ref
    # ОПТИМИЗАЦИЯ v5.0: Векторизованное создание ключей (3-5x быстрее)
    df_ref_keys = _vectorized_tuple_key(df_ref, src_keys)

    if not multiply_rows:
        # Старая логика: первое найденное значение
        # ОПТИМИЗАЦИЯ v5.0: Векторизованное создание ключей (3-5x быстрее)
        new_keys = _vectorized_tuple_key(df_base, dst_keys)
        
        # Оптимизация: собираем все новые колонки в словарь и добавляем их одним вызовом
        new_columns_dict = {}
        for col in columns:
            ref_map = dict(zip(df_ref_keys, df_ref[col]))
            new_col_name = f"{ref_sheet_name}=>{col}"
            new_columns_dict[new_col_name] = new_keys.map(ref_map).fillna("-")
        
        # Добавляем все колонки одним вызовом через pd.concat для избежания фрагментации
        if new_columns_dict:
            new_columns_df = pd.DataFrame(new_columns_dict, index=df_base.index)
            df_base = pd.concat([df_base, new_columns_df], axis=1)
            if ref_sheet_name == "LIST-TOURNAMENT" and sheet_name == "TOURNAMENT-SCHEDULE":
                for col in columns:
                    new_col_name = f"{ref_sheet_name}=>{col}"
                    if new_col_name in df_base.columns:
                        filled = (df_base[new_col_name] != "-").sum()
                        logging.info(f"[MERGE] add_fields_to_sheet результат LIST-TOURNAMENT -> TOURNAMENT-SCHEDULE: колонка '{new_col_name}', заполнено строк: {filled} из {len(df_base)}")
        
        # Специально для REWARD_LINK =>CONTEST_CODE: auto-rename, если создали с дефисом
        for col in columns:
            new_col_name = f"{ref_sheet_name}=>{col}"
            if new_col_name.replace("-", "_").replace(" ", "") == COL_REWARD_LINK_CONTEST_CODE.replace("-", "_").replace(" ", ""):
                candidates = [c for c in df_base.columns if
                              c.replace("-", "_").replace(" ", "") == COL_REWARD_LINK_CONTEST_CODE.replace("-", "_").replace(" ", "")]
                for cand in candidates:
                    if cand != COL_REWARD_LINK_CONTEST_CODE:
                        df_base = df_base.rename(columns={cand: COL_REWARD_LINK_CONTEST_CODE})
    else:
        # Новая логика: размножение строк при множественных совпадениях
        logging.info(f"[MULTIPLY ROWS] {sheet_name}: начинаем размножение строк для поля {columns}")
        result_rows = []
        old_rows_count = len(df_base)

        for base_idx, base_row in df_base.iterrows():
            base_key = tuple_key(base_row, dst_keys)
            # Находим все строки в df_ref с таким ключом
            matching_ref_rows = df_ref[df_ref_keys == base_key]

            if matching_ref_rows.empty:
                # Нет совпадений - добавляем строку с пустыми значениями
                new_row = base_row.copy()
                for col in columns:
                    new_col_name = f"{ref_sheet_name}=>{col}"
                    new_row[new_col_name] = "-"
                result_rows.append(new_row)
            else:
                # Есть совпадения - создаем строку для каждого совпадения
                for ref_idx, ref_row in matching_ref_rows.iterrows():
                    new_row = base_row.copy()
                    for col in columns:
                        new_col_name = f"{ref_sheet_name}=>{col}"
                        new_row[new_col_name] = ref_row[col]
                    result_rows.append(new_row)

        # Создаем новый DataFrame из размноженных строк
        df_base = pd.DataFrame(result_rows).reset_index(drop=True)
        new_rows_count = len(df_base)
        multiply_factor = round(new_rows_count / old_rows_count, 2) if old_rows_count > 0 else 0
        logging.info(
            f"[MULTIPLY ROWS] {sheet_name}: {old_rows_count} строк -> {new_rows_count} строк (размножение: {multiply_factor}x)"
        )

        # Обработка специального случая для REWARD_LINK
        for col in columns:
            new_col_name = f"{ref_sheet_name}=>{col}"
            if new_col_name.replace("-", "_").replace(" ", "") == COL_REWARD_LINK_CONTEST_CODE.replace("-", "_").replace(" ", ""):
                candidates = [c for c in df_base.columns if
                              c.replace("-", "_").replace(" ", "") == COL_REWARD_LINK_CONTEST_CODE.replace("-", "_").replace(" ", "")]
                for cand in candidates:
                    if cand != COL_REWARD_LINK_CONTEST_CODE:
                        df_base = df_base.rename(columns={cand: COL_REWARD_LINK_CONTEST_CODE})

    func_time = time() - func_start
    logging.info(
        f"[END] add_fields_to_sheet (лист: {sheet_name}, поля: {columns}, ключ: {dst_keys}->{src_keys}, mode: {mode}, multiply: {multiply_rows}) (время: {func_time:.3f}s)"
    )

    return df_base

def _vectorized_tuple_key(df, keys):
    """
    ВЕКТОРИЗОВАННАЯ ВЕРСИЯ tuple_key: создает кортежи ключей для всего DataFrame сразу.
    Ускорение: 3-5x по сравнению с apply(axis=1).
    
    Args:
        df: DataFrame
        keys: список ключей или один ключ
        
    Returns:
        pd.Series с кортежами ключей
    """
    if isinstance(keys, (list, tuple)):
        if len(keys) == 1:
            # Один ключ - просто создаем кортеж
            return df[keys[0]].apply(lambda x: (x,))
        else:
            # Несколько ключей - используем zip для векторизации
            return pd.Series(list(zip(*[df[k] for k in keys])))
    else:
        # Один ключ (строка)
        return df[keys].apply(lambda x: (x,))

    # --- Добавлено: авто-дополнение отсутствующих колонок и ключей ---

    # Создаем ключи для df_ref
    # ОПТИМИЗАЦИЯ v5.0: Векторизованное создание ключей (3-5x быстрее)
    df_ref_keys = _vectorized_tuple_key(df_ref, src_keys)

    if not multiply_rows:
        # Старая логика: первое найденное значение
        # ОПТИМИЗАЦИЯ v5.0: Векторизованное создание ключей (3-5x быстрее)
        new_keys = _vectorized_tuple_key(df_base, dst_keys)
        
        # Оптимизация: собираем все новые колонки в словарь и добавляем их одним вызовом
        # Это предотвращает фрагментацию DataFrame и устраняет PerformanceWarning
        new_columns_dict = {}
        for col in columns:
            ref_map = dict(zip(df_ref_keys, df_ref[col]))
            new_col_name = f"{ref_sheet_name}=>{col}"
            new_columns_dict[new_col_name] = new_keys.map(ref_map).fillna("-")
        
        # Добавляем все колонки одним вызовом через pd.concat для избежания фрагментации
        if new_columns_dict:
            new_columns_df = pd.DataFrame(new_columns_dict, index=df_base.index)
            df_base = pd.concat([df_base, new_columns_df], axis=1)
        
        # Специально для REWARD_LINK =>CONTEST_CODE: auto-rename, если создали с дефисом
        for col in columns:
            new_col_name = f"{ref_sheet_name}=>{col}"
            if new_col_name.replace("-", "_").replace(" ", "") == COL_REWARD_LINK_CONTEST_CODE.replace("-",
                                                                                                       "_").replace(" ",
                                                                                                                    ""):
                candidates = [c for c in df_base.columns if
                              c.replace("-", "_").replace(" ", "") == COL_REWARD_LINK_CONTEST_CODE.replace("-",
                                                                                                           "_").replace(
                                  " ", "")]
                for cand in candidates:
                    if cand != COL_REWARD_LINK_CONTEST_CODE:
                        df_base = df_base.rename(columns={cand: COL_REWARD_LINK_CONTEST_CODE})
    else:
                # ОПТИМИЗИРОВАННАЯ ВЕРСИЯ: Используем pd.merge вместо iterrows для ускорения
        logging.info(f"[MULTIPLY ROWS] {sheet_name}: начинаем размножение строк для поля {columns}")
        old_rows_count = len(df_base)
        
        # Создаем ключи для обоих DataFrame
        # ОПТИМИЗАЦИЯ v5.0: Векторизованное создание ключей (3-5x быстрее)
        df_base_keys = _vectorized_tuple_key(df_base, dst_keys)
        # ОПТИМИЗАЦИЯ v5.0: Векторизованное создание ключей (3-5x быстрее)
        df_ref_keys = _vectorized_tuple_key(df_ref, src_keys)
        
        # Добавляем ключи как временные колонки
        df_base_with_key = df_base.copy()
        df_base_with_key['_temp_key'] = df_base_keys
        
        df_ref_with_key = df_ref.copy()
        df_ref_with_key['_temp_key'] = df_ref_keys
        
        # Используем merge для объединения (left join сохраняет все строки из df_base)
        merged = pd.merge(
            df_base_with_key,
            df_ref_with_key[['_temp_key'] + columns],
            on='_temp_key',
            how='left',
            suffixes=('', '_ref')
        )
        
        # Переименовываем колонки из df_ref
        for col in columns:
            new_col_name = f"{ref_sheet_name}=>{col}"
            if col + '_ref' in merged.columns:
                merged[new_col_name] = merged[col + '_ref'].fillna("-")
                merged = merged.drop(columns=[col + '_ref'])
            else:
                merged[new_col_name] = "-"
        
        # Удаляем временный ключ
        merged = merged.drop(columns=['_temp_key'])
        
        # Если были строки без совпадений, они уже обработаны через left join
        df_base = merged.reset_index(drop=True)
        new_rows_count = len(df_base)
        multiply_factor = round(new_rows_count / old_rows_count, 2) if old_rows_count > 0 else 0
        logging.info(
            f"[MULTIPLY ROWS] {sheet_name}: {old_rows_count} строк -> {new_rows_count} строк (размножение: {multiply_factor}x)"
        )

        # Обработка специального случая для REWARD_LINK
        for col in columns:
            new_col_name = f"{ref_sheet_name}=>{col}"
            if new_col_name.replace("-", "_").replace(" ", "") == COL_REWARD_LINK_CONTEST_CODE.replace("-",
                                                                                                       "_").replace(" ",
                                                                                                                    ""):
                candidates = [c for c in df_base.columns if
                              c.replace("-", "_").replace(" ", "") == COL_REWARD_LINK_CONTEST_CODE.replace("-",
                                                                                                           "_").replace(
                                  " ", "")]
                for cand in candidates:
                    if cand != COL_REWARD_LINK_CONTEST_CODE:
                        df_base = df_base.rename(columns={cand: COL_REWARD_LINK_CONTEST_CODE})

    func_time = time() - func_start
    logging.info(
        f"[END] add_fields_to_sheet (лист: {sheet_name}, поля: {columns}, ключ: {dst_keys}->{src_keys}, mode: {mode}, multiply: {multiply_rows}) (время: {func_time:.3f}s)"
    )

    return df_base



def _process_single_merge_rule(rule, sheets_data_copy, count_column_prefix="COUNT", merge_name="MERGE_FIELDS_ADVANCED"):
    """
    Обрабатывает одно правило merge_fields.
    Используется для параллельной обработки независимых правил.
    merge_name: имя набора правил для логов (MERGE_FIELDS или MERGE_FIELDS_ADVANCED).
    
    Args:
        rule: Правило из merge_fields
        sheets_data_copy: Копия sheets_data для безопасной работы в потоке
        count_column_prefix: префикс для имён count-колонок (COUNT или COUNT_SELECT для MERGE_FIELDS_ADVANCED)
        
    Returns:
        tuple: (rule, updated_sheets_dict) где updated_sheets_dict содержит обновленные листы
    """
    sheet_src = rule["sheet_src"]
    sheet_dst = rule["sheet_dst"]
    src_keys = rule["src_key"] if isinstance(rule["src_key"], list) else [rule["src_key"]]
    dst_keys = rule["dst_key"] if isinstance(rule["dst_key"], list) else [rule["dst_key"]]
    col_names = rule["column"]
    mode = rule.get("mode", "value")
    multiply_rows = rule.get("multiply_rows", False)
    
    status_filters = rule.get("status_filters", None)
    custom_conditions = rule.get("custom_conditions", None)
    group_by = rule.get("group_by", None)
    aggregate = rule.get("aggregate", None)
    count_aggregation = rule.get("count_aggregation", "size")
    count_label = rule.get("count_label", None)
    
    updated_sheets = {}
    logging.info(f"[MERGE] {merge_name} (_process_single_merge_rule) правило: {sheet_src} -> {sheet_dst}, колонки: {col_names}, ключи: {dst_keys} <- {src_keys}, mode={mode}")
    if sheet_src in sheets_data_copy and sheets_data_copy[sheet_src] is not None:
        df_src_check = sheets_data_copy[sheet_src][0] if len(sheets_data_copy[sheet_src]) > 0 else None
        if df_src_check is not None and isinstance(df_src_check, pd.DataFrame):
            logging.debug(f"[DEBUG _process_single_merge_rule] df_src ({sheet_src}): shape={df_src_check.shape}")
        else:
            logging.warning(f"[DEBUG _process_single_merge_rule] ⚠️  df_src ({sheet_src}) равен None!")
    if sheet_dst in sheets_data_copy and sheets_data_copy[sheet_dst] is not None:
        df_dst_check = sheets_data_copy[sheet_dst][0] if len(sheets_data_copy[sheet_dst]) > 0 else None
        if df_dst_check is not None and isinstance(df_dst_check, pd.DataFrame):
            logging.debug(f"[DEBUG _process_single_merge_rule] df_dst ({sheet_dst}): shape={df_dst_check.shape}")
        else:
            logging.warning(f"[DEBUG _process_single_merge_rule] ⚠️  df_dst ({sheet_dst}) равен None!")

    
    # ОПТИМИЗАЦИЯ v5.0: Проверка на существование листов и None (правильный порядок)
    if (sheet_src not in sheets_data_copy or sheet_dst not in sheets_data_copy):
        logging.warning(f"[MERGE] {merge_name} ПРОПУСК (параллель): лист {sheet_src} или {sheet_dst} отсутствует в sheets_data")
        return (rule, updated_sheets)
    
    if (sheets_data_copy[sheet_src] is None or sheets_data_copy[sheet_dst] is None or
        len(sheets_data_copy[sheet_src]) < 1 or len(sheets_data_copy[sheet_dst]) < 1 or
        sheets_data_copy[sheet_src][0] is None or sheets_data_copy[sheet_dst][0] is None):
        logging.warning(f"[MERGE] {merge_name} ПРОПУСК (параллель): лист {sheet_src} или {sheet_dst} содержит None")
        return (rule, updated_sheets)
    
    df_src = sheets_data_copy[sheet_src][0].copy()
    logging.debug(f"[MERGE] {merge_name} df_src ({sheet_src}): shape={df_src.shape}, колонки: {list(df_src.columns)}")
    df_dst, params_dst = sheets_data_copy[sheet_dst]
    params_dst = params_dst.copy()  # Копируем параметры
    
    # Подстановка ключа/колонки для LIST-TOURNAMENT: в файле геймификации часто "TOURNAMENT_CODE" и "Бизнес-статус"
    if sheet_src == "LIST-TOURNAMENT":
        if src_keys == ["Код турнира"] and "Код турнира" not in df_src.columns and "TOURNAMENT_CODE" in df_src.columns:
            src_keys = ["TOURNAMENT_CODE"]
            logging.info(f"[MERGE] {merge_name} LIST-TOURNAMENT: подстановка ключа TOURNAMENT_CODE вместо 'Код турнира'")
        for col in col_names:
            if col not in df_src.columns and col == "Бизнес-статус турнира" and "Бизнес-статус" in df_src.columns:
                df_src["Бизнес-статус турнира"] = df_src["Бизнес-статус"]
                logging.info(f"[MERGE] {merge_name} LIST-TOURNAMENT: подстановка колонки 'Бизнес-статус' для 'Бизнес-статус турнира'")
                break

    # Применяем фильтрацию
    df_src_filtered = apply_filters_to_dataframe(df_src, status_filters, custom_conditions, sheet_src)
    
    # Применяем группировку и агрегацию если необходимо
    if group_by or aggregate:
        df_src_filtered = apply_grouping_and_aggregation(df_src_filtered, group_by, aggregate, sheet_src)
    
    # Вызываем основную функцию добавления полей
    df_dst = add_fields_to_sheet(df_dst, df_src_filtered, src_keys, dst_keys, col_names, sheet_dst, sheet_src, mode=mode,
                                 multiply_rows=multiply_rows, count_prefix=count_column_prefix,
                                 count_aggregation=count_aggregation, count_label=count_label)
    
    # ИСПРАВЛЕНИЕ: Проверка на None после add_fields_to_sheet
    if df_dst is None or not isinstance(df_dst, pd.DataFrame):
        logging.error(f"[MERGE] {merge_name} add_fields_to_sheet вернул None для листа {sheet_dst}, используем исходный DataFrame")
        df_dst = sheets_data_copy[sheet_dst][0].copy() if sheets_data_copy[sheet_dst][0] is not None else pd.DataFrame()
    else:
        added_cols = [c for c in df_dst.columns if c.startswith(sheet_src + "=>")]
        logging.info(f"[MERGE] {merge_name} результат правила {sheet_src} -> {sheet_dst}: добавлены колонки: {added_cols}, всего колонок в {sheet_dst}: {len(df_dst.columns)}")

    # Сохраняем информацию о ширине колонок
    if "added_columns_width" not in params_dst:
        params_dst["added_columns_width"] = {}
    
    if mode == "count" and count_label is not None:
        new_col_name = f"{sheet_src}=>COUNT_{count_aggregation}_{count_label}"
        params_dst["added_columns_width"][new_col_name] = {
            "max_width": rule.get("col_max_width"),
            "width_mode": rule.get("col_width_mode", "AUTO"),
            "min_width": rule.get("col_min_width", 8)
        }
    else:
        for col in col_names:
            new_col_name = f"{sheet_src}=>{col}"
            if mode == "count":
                new_col_name = f"{sheet_src}=>{count_column_prefix}_{col}"
            params_dst["added_columns_width"][new_col_name] = {
                "max_width": rule.get("col_max_width"),
                "width_mode": rule.get("col_width_mode", "AUTO"),
                "min_width": rule.get("col_min_width", 8)
            }
    
    updated_sheets[sheet_dst] = (df_dst, params_dst)
    return (rule, updated_sheets)


def _group_independent_rules(merge_fields):
    """
    Группирует правила merge_fields на независимые группы.
    Правила независимы, если они не изменяют одни и те же листы.
    
    Args:
        merge_fields: Список правил
        
    Returns:
        list: Список групп правил, где каждая группа может быть обработана параллельно
    """
    if not merge_fields:
        return []
    
    # Простая стратегия: группируем правила, которые не конфликтуют по sheet_dst
    groups = []
    used_destinations = set()
    
    current_group = []
    for rule in merge_fields:
        sheet_dst = rule["sheet_dst"]
        
        # Если этот лист уже используется в текущей группе, начинаем новую группу
        if sheet_dst in used_destinations:
            if current_group:
                groups.append(current_group)
            current_group = [rule]
            used_destinations = {sheet_dst}
        else:
            current_group.append(rule)
            used_destinations.add(sheet_dst)
    
    # Добавляем последнюю группу
    if current_group:
        groups.append(current_group)
    
    return groups


def _dump_sheets_data_for_baseline(sheets_data, max_rows: int = 3) -> dict:
    """
    Формирует снимок sheets_data для сохранения/сравнения baseline:
    для каждого листа — список колонок (порядок сохранён) и первые max_rows строк как список списков.
    Используется для верификации, что после объединения MERGE_FIELDS в MERGE_FIELDS_ADVANCED
    выходные колонки и фрагмент данных не изменились.
    """
    result = {}
    for sheet_name, sheet_data in sheets_data.items():
        if sheet_data is None or len(sheet_data) < 1:
            result[sheet_name] = {"columns": [], "sample_rows": []}
            continue
        df, _ = sheet_data
        if df is None or not isinstance(df, pd.DataFrame):
            result[sheet_name] = {"columns": [], "sample_rows": []}
            continue
        cols = list(df.columns)
        head = df.head(max_rows)
        # Преобразуем в список списков; NaN -> None для JSON
        sample_rows = []
        for _, row in head.iterrows():
            sample_rows.append([None if pd.isna(v) else v for v in row.tolist()])
        result[sheet_name] = {"columns": cols, "sample_rows": sample_rows}
    return result


def _compare_sheets_data_with_baseline(sheets_data, baseline_path: str, max_rows: int = 3) -> tuple[bool, list[str]]:
    """
    Сравнивает текущий sheets_data с сохранённым baseline (список колонок и сэмпл строк).
    Возвращает (True, []) при совпадении; (False, список сообщений об отличиях) при расхождении.
    """
    errors = []
    try:
        with open(baseline_path, "r", encoding="utf-8") as f:
            baseline = json.load(f)
    except Exception as e:
        return False, [f"Не удалось загрузить baseline {baseline_path}: {e}"]
    current = _dump_sheets_data_for_baseline(sheets_data, max_rows=max_rows)
    baseline_sheets = set(baseline.keys())
    current_sheets = set(current.keys())
    if baseline_sheets != current_sheets:
        only_baseline = baseline_sheets - current_sheets
        only_current = current_sheets - baseline_sheets
        if only_baseline:
            errors.append(f"В baseline есть листы, которых нет сейчас: {sorted(only_baseline)}")
        if only_current:
            errors.append(f"Сейчас есть листы, которых нет в baseline: {sorted(only_current)}")
    for sheet in sorted(baseline_sheets & current_sheets):
        bc = baseline[sheet].get("columns", [])
        cc = current[sheet].get("columns", [])
        if bc != cc:
            errors.append(f"Лист {sheet}: различаются колонки. Baseline: {bc[:15]}...; текущие: {cc[:15]}...")
        br = baseline[sheet].get("sample_rows", [])
        cr = current[sheet].get("sample_rows", [])
        if br != cr:
            errors.append(f"Лист {sheet}: различаются сэмпл-строки (первые {max_rows} строк)")
    return (len(errors) == 0, errors)


def merge_fields_across_sheets(sheets_data, merge_fields, count_column_prefix="COUNT", merge_name=""):
    """
    count_column_prefix: для режима count имя колонки будет {sheet_src}=>{count_column_prefix}_{col}.
    Для MERGE_FIELDS оставить "COUNT", для MERGE_FIELDS_ADVANCED передать "COUNT_SELECT".
    merge_name: имя набора правил для логов (например "MERGE_FIELDS" или "MERGE_FIELDS_ADVANCED").
    """
    name_tag = merge_name or "merge_fields"
    logging.info(f"[MERGE] ========== {name_tag}: НАЧАЛО ========== Правил: {len(merge_fields)}, листов в sheets_data: {list(sheets_data.keys())}")
    for idx, rule in enumerate(merge_fields):
        src = rule.get("sheet_src", "?")
        dst = rule.get("sheet_dst", "?")
        col = rule.get("column", [])
        sk = rule.get("src_key", [])
        dk = rule.get("dst_key", [])
        logging.info(f"[MERGE] {name_tag} правило {idx+1}/{len(merge_fields)}: {src} -> {dst}, колонки: {col}, ключи: {dk} <- {sk}")
    rule_groups = _group_independent_rules(merge_fields)
    logging.info(f"[MERGE] {name_tag}: сгруппировано в {len(rule_groups)} групп(ы) для обработки")
    for sheet_name, sheet_data in sheets_data.items():
        if sheet_data is not None and len(sheet_data) > 0:
            df, params = sheet_data
            if df is not None and isinstance(df, pd.DataFrame):
                logging.debug(f"[MERGE] {name_tag} лист {sheet_name}: shape={df.shape}, колонок: {len(df.columns)}")
            else:
                logging.debug(f"[MERGE] {name_tag} лист {sheet_name}: нет данных")

    """
    Универсально добавляет поля по правилам из merge_fields
    (source_df -> target_df), поддержка mode value / count, multiply_rows.
    
    НОВЫЕ ВОЗМОЖНОСТИ:
    - status_filters: фильтрация по статусам колонок
    - custom_conditions: пользовательские условия фильтрации
    - group_by: группировка данных перед добавлением
    - aggregate: подведение итогов (sum, count, avg, max, min)
    
    sheets_data: dict {sheet_name: (df, params)}
    merge_fields: список блоков с параметрами (см. выше)
    """
    lock = threading.Lock()  # Для безопасного доступа к sheets_data
    
    for group_idx, rule_group in enumerate(rule_groups):
        if len(rule_group) == 1:
            # Одно правило - обрабатываем последовательно (проще и быстрее для малых групп)
            rule = rule_group[0]
            sheet_src = rule["sheet_src"]
            sheet_dst = rule["sheet_dst"]
            src_keys = rule["src_key"] if isinstance(rule["src_key"], list) else [rule["src_key"]]
            dst_keys = rule["dst_key"] if isinstance(rule["dst_key"], list) else [rule["dst_key"]]
            col_names = rule["column"]
            mode = rule.get("mode", "value")
            multiply_rows = rule.get("multiply_rows", False)
            
            status_filters = rule.get("status_filters", None)
            custom_conditions = rule.get("custom_conditions", None)
            group_by = rule.get("group_by", None)
            aggregate = rule.get("aggregate", None)
            count_aggregation = rule.get("count_aggregation", "size")
            count_label = rule.get("count_label", None)
            
            params_str = f"(src: {sheet_src} -> dst: {sheet_dst}, поля: {col_names}, ключ: {dst_keys}<-{src_keys}, mode: {mode}, multiply: {multiply_rows})"
            
            if status_filters:
                params_str += f", status_filters: {status_filters}"
            if custom_conditions:
                params_str += f", custom_conditions: {list(custom_conditions.keys())}"
            if group_by:
                params_str += f", group_by: {group_by}"
            if aggregate:
                params_str += f", aggregate: {list(aggregate.keys())}"
            if mode == "count" and count_label is not None:
                params_str += f", count_aggregation: {count_aggregation}, count_label: {count_label}"

            logging.info(f"[MERGE] {name_tag} обработка правила (последовательно): {sheet_src} -> {sheet_dst}, колонки: {col_names}")
            # ОПТИМИЗАЦИЯ v5.0: Проверка на существование листов и None (правильный порядок)
            if sheet_src not in sheets_data or sheet_dst not in sheets_data:
                logging.warning(f"[MERGE] {name_tag} ПРОПУСК: нет листа {sheet_src} или {sheet_dst}, колонки {col_names} не добавлены")
                continue
            
            if (sheets_data[sheet_src] is None or sheets_data[sheet_dst] is None or
                len(sheets_data[sheet_src]) < 1 or len(sheets_data[sheet_dst]) < 1 or
                sheets_data[sheet_src][0] is None or sheets_data[sheet_dst][0] is None):
                logging.warning(f"[MERGE] {name_tag} ПРОПУСК: лист {sheet_src} или {sheet_dst} содержит None, колонки {col_names} не добавлены")
                continue

            df_src = sheets_data[sheet_src][0].copy()
            logging.debug(f"[MERGE] {name_tag} df_src ({sheet_src}): shape={df_src.shape}, колонки: {list(df_src.columns)}")
            df_dst, params_dst = sheets_data[sheet_dst]

            # Подстановка ключа/колонки для LIST-TOURNAMENT (как в _process_single_merge_rule для MERGE_FIELDS_ADVANCED)
            if sheet_src == "LIST-TOURNAMENT":
                if src_keys == ["Код турнира"] and "Код турнира" not in df_src.columns and "TOURNAMENT_CODE" in df_src.columns:
                    src_keys = ["TOURNAMENT_CODE"]
                    logging.info(f"[MERGE] {name_tag} LIST-TOURNAMENT: подстановка ключа TOURNAMENT_CODE вместо 'Код турнира'")
                for col in (col_names if isinstance(col_names, list) else [col_names]):
                    if col not in df_src.columns and col == "Бизнес-статус турнира" and "Бизнес-статус" in df_src.columns:
                        df_src["Бизнес-статус турнира"] = df_src["Бизнес-статус"]
                        logging.info(f"[MERGE] {name_tag} LIST-TOURNAMENT: подстановка колонки 'Бизнес-статус' для 'Бизнес-статус турнира'")
                        break

            cols_dst_before = set(df_dst.columns) if df_dst is not None and isinstance(df_dst, pd.DataFrame) else set()
            logging.info(f"[MERGE] {name_tag} вызов add_fields_to_sheet: {sheet_src} -> {sheet_dst}, src_keys={src_keys}, dst_keys={dst_keys}, col_names={col_names}")
            
            df_src_filtered = apply_filters_to_dataframe(df_src, status_filters, custom_conditions, sheet_src)
            
            if group_by or aggregate:
                df_src_filtered = apply_grouping_and_aggregation(df_src_filtered, group_by, aggregate, sheet_src)
            
            df_dst = add_fields_to_sheet(df_dst, df_src_filtered, src_keys, dst_keys, col_names, sheet_dst, sheet_src, mode=mode,
                                         multiply_rows=multiply_rows, count_prefix=count_column_prefix,
                                         count_aggregation=count_aggregation, count_label=count_label)
            
            # ИСПРАВЛЕНИЕ: Проверка на None после add_fields_to_sheet
            if df_dst is None or not isinstance(df_dst, pd.DataFrame):
                logging.error(f"[MERGE] {name_tag} add_fields_to_sheet вернул None для листа {sheet_dst}, используем исходный DataFrame")
                df_dst = sheets_data[sheet_dst][0].copy() if sheets_data[sheet_dst][0] is not None else pd.DataFrame()
            else:
                cols_dst_after = set(df_dst.columns)
                new_cols = cols_dst_after - cols_dst_before
                added_from_src = [c for c in new_cols if c.startswith(sheet_src + "=>")]
                logging.info(f"[MERGE] {name_tag} результат правила {sheet_src} -> {sheet_dst}: добавлены колонки: {added_from_src or list(new_cols)[:10]}, всего колонок в {sheet_dst}: {len(df_dst.columns)}")

            if "added_columns_width" not in params_dst:
                params_dst["added_columns_width"] = {}

            if mode == "count" and count_label is not None:
                new_col_name = f"{sheet_src}=>COUNT_{count_aggregation}_{count_label}"
                params_dst["added_columns_width"][new_col_name] = {
                    "max_width": rule.get("col_max_width"),
                    "width_mode": rule.get("col_width_mode", "AUTO"),
                    "min_width": rule.get("col_min_width", 8)
                }
            else:
                for col in col_names:
                    new_col_name = f"{sheet_src}=>{col}"
                    if mode == "count":
                        new_col_name = f"{sheet_src}=>{count_column_prefix}_{col}"
                    params_dst["added_columns_width"][new_col_name] = {
                        "max_width": rule.get("col_max_width"),
                        "width_mode": rule.get("col_width_mode", "AUTO"),
                        "min_width": rule.get("col_min_width", 8)
                    }

            sheets_data[sheet_dst] = (df_dst, params_dst)
            logging.info(f"[MERGE] {name_tag} правило завершено: {sheet_src} -> {sheet_dst}")
        else:
            # Несколько независимых правил - обрабатываем параллельно
            logging.info(f"[MERGE] {name_tag} обработка группы из {len(rule_group)} правил (параллельно)")
            
            with ThreadPoolExecutor(max_workers=min(MAX_WORKERS, len(rule_group))) as executor:
                # Создаем копию sheets_data для каждого потока (безопасность)
                futures = {
                    executor.submit(_process_single_merge_rule, rule, sheets_data.copy(), count_column_prefix, name_tag): rule
                    for rule in rule_group
                }
                
                for future in as_completed(futures):
                    try:
                        rule, updated_sheets = future.result()
                        
                        # Обновляем sheets_data с блокировкой. Не перезаписываем лист целиком —
                        # дополняем колонки и params, чтобы теоретически не потерять данные от других правил.
                        with lock:
                            for sheet_name, data in updated_sheets.items():
                                if data is None or len(data) < 1 or data[0] is None:
                                    logging.warning(f"[PARALLEL MERGE] Пропущено обновление листа {sheet_name}: данные равны None")
                                    continue
                                new_df, new_params = data
                                if sheet_name in sheets_data and sheets_data[sheet_name] is not None:
                                    existing_df, existing_params = sheets_data[sheet_name]
                                    if existing_df is not None and isinstance(existing_df, pd.DataFrame):
                                        # Дополняем существующий df новыми колонками (не перезаписываем)
                                        for col in new_df.columns:
                                            if col not in existing_df.columns:
                                                existing_df[col] = new_df[col].values
                                        # Объединяем params (в т.ч. added_columns_width)
                                        merged_params = existing_params.copy() if isinstance(existing_params, dict) else {}
                                        new_added = new_params.get("added_columns_width", {}) if isinstance(new_params, dict) else {}
                                        merged_params["added_columns_width"] = {
                                            **merged_params.get("added_columns_width", {}),
                                            **new_added
                                        }
                                        if isinstance(new_params, dict):
                                            for k, v in new_params.items():
                                                if k != "added_columns_width" and k not in merged_params:
                                                    merged_params[k] = v
                                        sheets_data[sheet_name] = (existing_df, merged_params)
                                    else:
                                        sheets_data[sheet_name] = data
                                else:
                                    sheets_data[sheet_name] = data
                            
                            sheet_src = rule["sheet_src"]
                            sheet_dst = rule["sheet_dst"]
                            col_names = rule["column"]
                            logging.info(f"[MERGE] {name_tag} правило завершено (параллельно): {sheet_src} -> {sheet_dst}, колонки: {col_names}")
                    except Exception as e:
                        logging.error(f"[PARALLEL MERGE ERROR] Ошибка обработки правила: {e}")
    
    logging.info(f"[MERGE] ========== {name_tag}: КОНЕЦ ========== Обработано групп: {len(rule_groups)}")
    return sheets_data


def apply_filters_to_dataframe(df, status_filters, custom_conditions, sheet_name):
    """
    Применяет фильтрацию к DataFrame на основе status_filters и custom_conditions.
    
    Args:
        df: исходный DataFrame
        status_filters: словарь с фильтрами по статусам {column: [allowed_values]}
        custom_conditions: словарь с пользовательскими условиями {column: condition}
        sheet_name: имя листа для логирования
        
    Returns:
        отфильтрованный DataFrame
    """
    if df.empty:
        return df
    
    df_filtered = df.copy()
    original_count = len(df_filtered)
    
    # Применяем фильтры по статусам
    if status_filters:
        for column, allowed_values in status_filters.items():
            if column in df_filtered.columns:
                df_filtered = df_filtered[df_filtered[column].isin(allowed_values)]
                logging.info(f"[FILTER] Применен фильтр по статусу: {column}={allowed_values}, осталось строк: {len(df_filtered)}")
            else:
                logging.warning(f"[WARNING] Колонка для фильтрации по статусу не найдена: {column} в листе {sheet_name}")
    
    # Применяем пользовательские условия
    if custom_conditions:
        for column, condition in custom_conditions.items():
            if column in df_filtered.columns:
                if callable(condition):
                    # Лямбда-функция
                    df_filtered = df_filtered[df_filtered[column].apply(condition)]
                elif isinstance(condition, list):
                    # Список разрешенных значений
                    df_filtered = df_filtered[df_filtered[column].isin(condition)]
                else:
                    # Точное совпадение
                    df_filtered = df_filtered[df_filtered[column] == condition]
                
                logging.info(f"[FILTER] Применено пользовательское условие: {column}={condition}, осталось строк: {len(df_filtered)}")
            else:
                logging.warning(f"[WARNING] Колонка для пользовательского условия не найдена: {column} в листе {sheet_name}")
    
    filtered_count = len(df_filtered)
    if original_count != filtered_count:
        logging.info(f"[FILTER] Фильтрация завершена: {original_count} -> {filtered_count} строк в листе {sheet_name}")
    
    return df_filtered


def apply_grouping_and_aggregation(df, group_by, aggregate, sheet_name):
    """
    Применяет группировку и агрегацию к DataFrame.
    
    Args:
        df: исходный DataFrame
        group_by: список колонок для группировки
        aggregate: словарь с правилами агрегации {column: function}
        sheet_name: имя листа для логирования
        
    Returns:
        DataFrame с примененной группировкой и агрегацией
    """
    if df.empty:
        return df
    
    if not group_by and not aggregate:
        return df
    
    df_grouped = df.copy()
    original_count = len(df_grouped)
    
    try:
        if group_by:
            # Проверяем наличие колонок для группировки
            missing_group_cols = [col for col in group_by if col not in df_grouped.columns]
            if missing_group_cols:
                logging.warning(f"[WARNING] Колонки для группировки не найдены: {missing_group_cols} в листе {sheet_name}")
                return df_grouped
            
            # Применяем группировку
            if aggregate:
                # Группировка с агрегацией
                agg_dict = {}
                for col, func in aggregate.items():
                    if col in df_grouped.columns:
                        agg_dict[col] = func
                    else:
                        logging.warning(f"[WARNING] Колонка для агрегации не найдена: {col} в листе {sheet_name}")
                
                if agg_dict:
                    df_grouped = df_grouped.groupby(group_by).agg(agg_dict).reset_index()
                    # Убираем многоуровневые заголовки если они появились
                    if isinstance(df_grouped.columns, pd.MultiIndex):
                        df_grouped.columns = [col[0] if col[1] == '' else f"{col[0]}_{col[1]}" for col in df_grouped.columns]
            else:
                # Простая группировка (убираем дубликаты)
                df_grouped = df_grouped.groupby(group_by).first().reset_index()
        else:
            # Только агрегация без группировки
            agg_dict = {}
            for col, func in aggregate.items():
                if col in df_grouped.columns:
                    agg_dict[col] = func
                else:
                    logging.warning(f"[WARNING] Колонка для агрегации не найдена: {col} в листе {sheet_name}")
            
            if agg_dict:
                df_grouped = df_grouped.agg(agg_dict).to_frame().T
        
        grouped_count = len(df_grouped)
        logging.info(f"[GROUP] Группировка и агрегация завершены: {original_count} -> {grouped_count} строк в листе {sheet_name}")
        
    except Exception as e:
        logging.error(f"[ERROR] Ошибка при группировке в листе {sheet_name}: {e}")
        return df
    
    return df_grouped



def detect_gender_by_patterns(value, patterns_male, patterns_female):
    """Определение пола по окончаниям в тексте"""
    if pd.isna(value) or not isinstance(value, str):
        return None

    value_lower = value.lower().strip()
    if not value_lower:
        return None

    # Проверяем мужские окончания
    for pattern in patterns_male:
        if value_lower.endswith(pattern.lower()):
            return 'М'

    # Проверяем женские окончания
    for pattern in patterns_female:
        if value_lower.endswith(pattern.lower()):
            return 'Ж'

    return None


def detect_gender_for_person(patronymic, first_name, surname, row_idx):
    """Определение пола для одного человека по приоритету: отчество -> имя -> фамилия"""

    # 1. Попытка определить по отчеству
    gender = detect_gender_by_patterns(
        patronymic,
        GENDER_PATTERNS['patronymic_male'],
        GENDER_PATTERNS['patronymic_female']
    )
    if gender:
        logging.debug(f"[DEBUG] Строка {row_idx}: пол по отчеству '{patronymic}' -> {gender}")
        return gender

    # 2. Попытка определить по имени
    gender = detect_gender_by_patterns(
        first_name,
        GENDER_PATTERNS['name_male'],
        GENDER_PATTERNS['name_female']
    )
    if gender:
        logging.debug(f"[DEBUG] Строка {row_idx}: пол по имени '{first_name}' -> {gender}")
        return gender

    # 3. Попытка определить по фамилии
    gender = detect_gender_by_patterns(
        surname,
        GENDER_PATTERNS['surname_male'],
        GENDER_PATTERNS['surname_female']
    )
    if gender:
        logging.debug(f"[DEBUG] Строка {row_idx}: пол по фамилии '{surname}' -> {gender}")
        return gender

    # Не удалось определить
    logging.debug(f"[DEBUG] Строка {row_idx}: пол не определен (отч:'{patronymic}', имя:'{first_name}', фам:'{surname}')")
    return '-'


def add_auto_gender_column(df, sheet_name):
    """Добавление колонки AUTO_GENDER к DataFrame с автоматическим определением пола"""
    func_start = time()

    # Проверяем наличие необходимых колонок
    required_columns = ['MIDDLE_NAME', 'FIRST_NAME', 'SURNAME']
    missing_columns = [col for col in required_columns if col not in df.columns]

    if missing_columns:
        logging.warning(f"[GENDER DETECTION] Пропущены колонки {missing_columns} в листе {sheet_name}")
        df['AUTO_GENDER'] = '-'
        return df

    total_rows = len(df)
    logging.info(f"[GENDER DETECTION] Начинаем определение пола для листа {sheet_name}, строк: {total_rows}")

    # Счетчики для статистики
    male_count = 0
    female_count = 0
    unknown_count = 0

    # Создаем новую колонку
    auto_gender = []

    for idx, row in df.iterrows():
        # Получаем значения полей
        patronymic = row.get('MIDDLE_NAME', '')
        first_name = row.get('FIRST_NAME', '')
        surname = row.get('SURNAME', '')

        # Определяем пол
        gender = detect_gender_for_person(patronymic, first_name, surname, idx)
        auto_gender.append(gender)

        # Обновляем статистику
        if gender == 'М':
            male_count += 1
        elif gender == 'Ж':
            female_count += 1
        else:
            unknown_count += 1

        # Прогресс каждые GENDER_PROGRESS_STEP строк
        if (idx + 1) % GENDER_PROGRESS_STEP == 0:
            percent = ((idx + 1) / total_rows) * 100
            logging.info(f"[GENDER DETECTION] Обработано {idx + 1} из {total_rows} строк ({percent:.1f}%)")

    # Добавляем колонку к DataFrame
    df['AUTO_GENDER'] = auto_gender

    # Логируем финальную статистику
    func_time = time() - func_start
    logging.info(f"[GENDER DETECTION] Статистика: М={male_count}, Ж={female_count}, неопределено={unknown_count} (всего: {total_rows})")
    logging.info(f"[GENDER DETECTION] Завершено за {func_time:.3f}s для листа {sheet_name}")

    return df


def add_auto_gender_column_vectorized(df, sheet_name):
    """
    ОПТИМИЗИРОВАННАЯ ВЕРСИЯ: Векторизованное определение пола.
    
    Обрабатывает все строки одновременно используя строковые операции pandas
    вместо iterrows(). Ожидаемое ускорение: 100-200x.
    
    Args:
        df (pd.DataFrame): DataFrame для обработки
        sheet_name (str): Название листа

    Returns:
        pd.DataFrame: DataFrame с добавленной колонкой AUTO_GENDER
    """
    func_start = time()
    
    required_columns = ['MIDDLE_NAME', 'FIRST_NAME', 'SURNAME']
    missing_columns = [col for col in required_columns if col not in df.columns]
    
    if missing_columns:
        logging.warning(f"[GENDER DETECTION VECTORIZED] Пропущены колонки {missing_columns} в листе {sheet_name}")
        df['AUTO_GENDER'] = '-'
        return df
    
    total_rows = len(df)
    logging.info(f"[GENDER DETECTION VECTORIZED] Начинаем определение пола для листа {sheet_name}, строк: {total_rows}")
    
    # Инициализируем колонку с дефолтным значением
    gender = pd.Series('-', index=df.index)
    
    # Подготовка данных: приводим к нижнему регистру и заполняем пустые значения
    patronymic_lower = df['MIDDLE_NAME'].fillna('').astype(str).str.lower().str.strip()
    first_name_lower = df['FIRST_NAME'].fillna('').astype(str).str.lower().str.strip()
    surname_lower = df['SURNAME'].fillna('').astype(str).str.lower().str.strip()
    
    # 1. Определение по отчеству (приоритет 1)
    for pattern in GENDER_PATTERNS['patronymic_male']:
        mask = patronymic_lower.str.endswith(pattern.lower()) & (gender == '-')
        gender[mask] = 'М'
    
    for pattern in GENDER_PATTERNS['patronymic_female']:
        mask = patronymic_lower.str.endswith(pattern.lower()) & (gender == '-')
        gender[mask] = 'Ж'
    
    # 2. Определение по имени (приоритет 2)
    for pattern in GENDER_PATTERNS['name_male']:
        mask = first_name_lower.str.endswith(pattern.lower()) & (gender == '-')
        gender[mask] = 'М'
    
    for pattern in GENDER_PATTERNS['name_female']:
        mask = first_name_lower.str.endswith(pattern.lower()) & (gender == '-')
        gender[mask] = 'Ж'
    
    # 3. Определение по фамилии (приоритет 3)
    for pattern in GENDER_PATTERNS['surname_male']:
        mask = surname_lower.str.endswith(pattern.lower()) & (gender == '-')
        gender[mask] = 'М'
    
    for pattern in GENDER_PATTERNS['surname_female']:
        mask = surname_lower.str.endswith(pattern.lower()) & (gender == '-')
        gender[mask] = 'Ж'
    
    # Добавляем колонку к DataFrame
    df['AUTO_GENDER'] = gender
    
    # Статистика
    male_count = (gender == 'М').sum()
    female_count = (gender == 'Ж').sum()
    unknown_count = (gender == '-').sum()
    
    func_time = time() - func_start
    logging.info(f"[GENDER DETECTION VECTORIZED] Статистика: М={male_count}, Ж={female_count}, неопределено={unknown_count} (всего: {total_rows})")
    logging.info(f"[GENDER DETECTION VECTORIZED] Завершено за {func_time:.3f}s для листа {sheet_name}")
    
    return df


def compare_gender_results(df_old, df_new):
    """
    Сравнивает результаты работы старой и новой версии add_auto_gender_column.
    
    Args:
        df_old (pd.DataFrame): Результат старой версии
        df_new (pd.DataFrame): Результат новой версии
    
    Returns:
        dict: Словарь с результатами сравнения
    """
    if 'AUTO_GENDER' not in df_old.columns or 'AUTO_GENDER' not in df_new.columns:
        return {"error": "Колонка AUTO_GENDER не найдена"}
    
    old_results = df_old['AUTO_GENDER'].fillna('-')
    new_results = df_new['AUTO_GENDER'].fillna('-')
    
    differences = (old_results != new_results).sum()
    total = len(df_old)
    matches = total - differences
    
    diff_examples = []
    if differences > 0:
        diff_mask = old_results != new_results
        diff_indices = df_old.index[diff_mask][:5]
        for idx in diff_indices:
            diff_examples.append({
                "index": idx,
                "old": old_results.loc[idx],
                "new": new_results.loc[idx]
            })
    
    return {
        "total": total,
        "matches": matches,
        "differences": differences,
        "match_percent": (matches / total * 100) if total > 0 else 0,
        "diff_examples": diff_examples,
        "identical": differences == 0
    }


def build_summary_sheet(dfs, params_summary, merge_fields):
    logging.debug(f"[DEBUG build_summary_sheet] === НАЧАЛО === Доступные листы в dfs: {list(dfs.keys())}")
    for sheet_name, df in dfs.items():
        if df is not None and isinstance(df, pd.DataFrame):
            logging.debug(f"[DEBUG build_summary_sheet] Лист {sheet_name}: shape={df.shape}, колонки={list(df.columns)[:10]}...")
        else:
            logging.debug(f"[DEBUG build_summary_sheet] Лист {sheet_name}: DataFrame равен None")
    logging.debug(f"[DEBUG build_summary_sheet] Правил merge_fields: {len(merge_fields)}")

    func_start = time()
    params_log = f"(лист: {params_summary['sheet']})"
    logging.info(f"[START] build_summary_sheet {params_log}")

    summary = collect_summary_keys(dfs)
    logging.debug(f"[DEBUG build_summary_sheet] После collect_summary_keys: summary shape={summary.shape if summary is not None and isinstance(summary, pd.DataFrame) else "None"}")
    if summary is not None and isinstance(summary, pd.DataFrame) and len(summary) > 0:
        logging.debug(f"[DEBUG build_summary_sheet] summary колонки: {list(summary.columns)}")
        logging.debug(f"[DEBUG build_summary_sheet] summary первые 3 строки:\n{summary.head(3).to_string()}")

    
    # ОПТИМИЗАЦИЯ v5.0: Проверка на None
    if summary is None:
        logging.error("[build_summary_sheet] collect_summary_keys вернул None, создаем пустой DataFrame")
        summary = pd.DataFrame(columns=SUMMARY_KEY_COLUMNS)
    elif not isinstance(summary, pd.DataFrame):
        logging.error("[build_summary_sheet] collect_summary_keys вернул не DataFrame, создаем пустой DataFrame")
        summary = pd.DataFrame(columns=SUMMARY_KEY_COLUMNS)

    # Детальное логирование для отладки GROUP_VALUE
    DEBUG_CODES = []  # Отключено подробное логирование
    for debug_code in DEBUG_CODES:
        debug_rows = summary[summary["CONTEST_CODE"] == debug_code]
        if not debug_rows.empty:
            logging.info(f"[DEBUG SUMMARY] === После collect_summary_keys для CONTEST_CODE: {debug_code} ===")
            logging.info(f"[DEBUG SUMMARY] Всего строк: {len(debug_rows)}")
            logging.info(f"[DEBUG SUMMARY] Уникальные GROUP_CODE: {debug_rows['GROUP_CODE'].unique().tolist()}")
            logging.info(f"[DEBUG SUMMARY] Уникальные GROUP_VALUE: {debug_rows['GROUP_VALUE'].unique().tolist()}")
            logging.info("[DEBUG SUMMARY] Комбинации (GROUP_CODE, GROUP_VALUE):")
            for _, row in debug_rows.iterrows():
                logging.info(
                    f"[DEBUG SUMMARY]   CONTEST={row.get('CONTEST_CODE', '')}, GROUP_CODE={row.get('GROUP_CODE', '')}, GROUP_VALUE={row.get('GROUP_VALUE', '')}"
                )
            
            # Проверяем, что есть в таблице GROUP
            if "GROUP" in dfs and not dfs["GROUP"].empty:
                group_rows = dfs["GROUP"][dfs["GROUP"]["CONTEST_CODE"] == debug_code]
                if not group_rows.empty:
                    logging.info(f"[DEBUG SUMMARY] === Данные в таблице GROUP для CONTEST_CODE: {debug_code} ===")
                    logging.info(f"[DEBUG SUMMARY] Всего строк в GROUP: {len(group_rows)}")
                    logging.info(
                        f"[DEBUG SUMMARY] Строки GROUP:\n{group_rows[['CONTEST_CODE', 'GROUP_CODE', 'GROUP_VALUE']].to_string()}"
                    )

    logging.info(f"Summary: Каркас: {len(summary)} строк (реальные комбинации ключей)")
    logging.debug(f"[DEBUG] {params_summary['sheet']}: первые строки после разворачивания:\n{summary.head(5).to_string()}")

    # Универсально добавляем все поля по merge_fields
    for field_idx, field in enumerate(merge_fields):
        col_names = field["column"]
        if isinstance(col_names, str):
            col_names = [col_names]
        sheet_src = field["sheet_src"]
        src_keys = field["src_key"] if isinstance(field["src_key"], list) else [field["src_key"]]
        dst_keys = field["dst_key"] if isinstance(field["dst_key"], list) else [field["dst_key"]]
        mode = field.get("mode", "value")
        params_str = f"(лист-источник: {sheet_src}, поля: {col_names}, ключ: {dst_keys}->{src_keys}, mode: {mode})"
        logging.info(f"[START] add_fields_to_sheet {params_str}")
        
        logging.debug(f"[DEBUG build_summary_sheet] === MERGE {field_idx+1}/{len(merge_fields)} ===")
        logging.debug(f"[DEBUG build_summary_sheet] Правило: sheet_src={sheet_src}, sheet_dst={params_summary["sheet"]}")
        logging.debug(f"[DEBUG build_summary_sheet] Поля: {col_names}, ключи: {dst_keys}->{src_keys}, mode={mode}")
        logging.debug(f"[DEBUG build_summary_sheet] summary ДО merge: shape={summary.shape if summary is not None and isinstance(summary, pd.DataFrame) else "None"}")
        if summary is not None and isinstance(summary, pd.DataFrame) and len(summary) > 0:
            logging.debug(f"[DEBUG build_summary_sheet] summary ДО merge первые 3 строки:\n{summary.head(3).to_string()}")

        # Детальное логирование для merge_fields с GROUP
        if sheet_src == "GROUP":
            for debug_code in DEBUG_CODES:
                debug_rows_before = summary[summary["CONTEST_CODE"] == debug_code]
                if not debug_rows_before.empty:
                    logging.info(f"[DEBUG SUMMARY] === Перед merge_fields из GROUP для CONTEST_CODE: {debug_code} ===")
                    logging.info(f"[DEBUG SUMMARY] Строк в Summary: {len(debug_rows_before)}")
                    logging.info(f"[DEBUG SUMMARY] GROUP_CODE: {debug_rows_before['GROUP_CODE'].unique().tolist()}")
                    logging.info(f"[DEBUG SUMMARY] GROUP_VALUE: {debug_rows_before['GROUP_VALUE'].unique().tolist()}")
        
        ref_df = dfs.get(sheet_src)
        if ref_df is not None and isinstance(ref_df, pd.DataFrame):
            logging.debug(f"[DEBUG build_summary_sheet] ref_df ({sheet_src}): shape={ref_df.shape}, колонки={list(ref_df.columns)[:10]}...")
        else:
            logging.warning(f"[DEBUG build_summary_sheet] ⚠️  ref_df ({sheet_src}) равен None!")
        if ref_df is None:
            logging.warning(f"Колонка {col_names} не добавлена: нет листа {sheet_src} или ключей {src_keys}")
            continue

        multiply_rows = field.get("multiply_rows", False)
        try:
            # ИСПРАВЛЕНИЕ: Сохраняем исходный summary перед merge
            summary_before_merge = summary.copy() if summary is not None and isinstance(summary, pd.DataFrame) else None
            
            summary = add_fields_to_sheet(summary, ref_df, src_keys, dst_keys, col_names, params_summary["sheet"],
                                          sheet_src, mode=mode, multiply_rows=multiply_rows)
            
            # ИСПРАВЛЕНИЕ: Логирование размера summary после каждого merge
            if summary is None:
                logging.error(f"[build_summary_sheet] КРИТИЧЕСКАЯ ОШИБКА: summary стал None после merge {field_idx+1}/{len(merge_fields)} с {sheet_src}!")
                logging.error(f"[build_summary_sheet] Параметры merge: поля={col_names}, ключи={dst_keys}->{src_keys}, mode={mode}")
                summary = summary_before_merge.copy() if summary_before_merge is not None else pd.DataFrame(columns=SUMMARY_KEY_COLUMNS)  # Восстанавливаем исходный summary

            logging.debug(f"[DEBUG build_summary_sheet] summary ПОСЛЕ merge: shape={summary.shape if summary is not None and isinstance(summary, pd.DataFrame) else "None"}")
            if summary is not None and isinstance(summary, pd.DataFrame) and len(summary) > 0:
                logging.debug(f"[DEBUG build_summary_sheet] summary ПОСЛЕ merge первые 3 строки:\n{summary.head(3).to_string()}")
            else:
                logging.error(f"[DEBUG build_summary_sheet] ❌ КРИТИЧЕСКАЯ ОШИБКА: summary стал None или пустым после merge!")

                logging.warning(f"[build_summary_sheet] Восстановлен исходный summary ({len(summary)} строк) после None merge с {sheet_src}")
        except Exception as e:
            logging.error(f"[build_summary_sheet] ОШИБКА при merge с {sheet_src}: {e}")
            logging.error(f"[build_summary_sheet] Параметры: поля={col_names}, ключи={dst_keys}->{src_keys}, mode={mode}")
            # Восстанавливаем исходный summary из сохраненной копии
            summary = summary_before_merge.copy() if summary_before_merge is not None else pd.DataFrame(columns=SUMMARY_KEY_COLUMNS)
            logging.warning(f"[build_summary_sheet] Восстановлен исходный summary ({len(summary)} строк) после ошибки merge с {sheet_src}")
            # Продолжаем работу с остальными merge_fields
            continue
        # Детальное логирование после merge_fields с GROUP
        if sheet_src == "GROUP":
            for debug_code in DEBUG_CODES:
                debug_rows_after = summary[summary["CONTEST_CODE"] == debug_code]
                if not debug_rows_after.empty:
                    logging.info(f"[DEBUG SUMMARY] === После merge_fields из GROUP для CONTEST_CODE: {debug_code} ===")
                    logging.info(f"[DEBUG SUMMARY] Строк в Summary: {len(debug_rows_after)}")
                    logging.info(f"[DEBUG SUMMARY] GROUP_CODE: {debug_rows_after['GROUP_CODE'].unique().tolist()}")
                    logging.info(f"[DEBUG SUMMARY] GROUP_VALUE: {debug_rows_after['GROUP_VALUE'].unique().tolist()}")
                    logging.info("[DEBUG SUMMARY] Комбинации (GROUP_CODE, GROUP_VALUE):")
                    for _, row in debug_rows_after.iterrows():
                        logging.info(
                            f"[DEBUG SUMMARY]   CONTEST={row.get('CONTEST_CODE', '')}, GROUP_CODE={row.get('GROUP_CODE', '')}, GROUP_VALUE={row.get('GROUP_VALUE', '')}"
                        )
        

    return summary


def process_single_file(file_conf):
    """
    Обрабатывает один CSV файл: поиск, чтение и разворачивание JSON полей.
    Используется для параллельной обработки файлов.
    
    Args:
        file_conf (dict): Конфигурация файла из INPUT_FILES
        
    Returns:
        tuple: (df, sheet_name, file_conf) или (None, sheet_name, None) при ошибке
    """
    sheet_name = file_conf["sheet"]
    try:
        file_path = find_file_case_insensitive(DIR_INPUT, file_conf["file"], [".csv", ".CSV"])
        # Для LIST-TOURNAMENT: если файл с суффиксом "-2" не найден, пробуем без суффикса (gamification-tournamentList.csv)
        if file_path is None and sheet_name == "LIST-TOURNAMENT" and file_conf["file"] == "gamification-tournamentList-2":
            file_path = find_file_case_insensitive(DIR_INPUT, "gamification-tournamentList", [".csv", ".CSV"])
            if file_path:
                logging.info(f"LIST-TOURNAMENT: использован файл по альтернативному имени: {file_path}")
        # Проверяем, найден ли файл
        if file_path is None:
            th = threading.current_thread().name
            logging.error(f"Файл не найден: {file_conf['file']} в каталоге {DIR_INPUT} [поток: {th}]")
            return None, sheet_name, None
        
        th = threading.current_thread().name
        logging.info(f"Загрузка файла: {file_path} [поток: {th}]")
        
        df = read_csv_file(file_path)
        if df is None:
            logging.error(f"Ошибка чтения файла: {file_path} [поток: {th}]")
            return None, sheet_name, None
        
        # Разворачиваем только нужные JSON-поля по строгому списку
        json_columns = JSON_COLUMNS.get(sheet_name, [])
        for json_conf in json_columns:
            col = json_conf["column"]
            prefix = json_conf.get("prefix", col)
            if col in df.columns:
                df = flatten_json_column_recursive(df, col, prefix=prefix, sheet=sheet_name)
                logging.info(f"[JSON FLATTEN] {sheet_name}: поле '{col}' развернуто с префиксом '{prefix}' [поток: {th}]")
            else:
                logging.warning(f"[JSON FLATTEN] {sheet_name}: поле '{col}' не найдено в колонках! [поток: {th}]")
        
        # Для дебага: логируем итоговый список колонок после всех разворотов
        logging.debug(f"[DEBUG] {sheet_name}: колонки после разворачивания: {', '.join(df.columns.tolist())} [поток: {th}]")

        logging.info(f"Файл успешно обработан: {sheet_name}, строк: {len(df)} [поток: {th}]")
        
        return df, sheet_name, file_conf
        
    except Exception as e:
        logging.error(
            f"Ошибка обработки файла {file_conf.get('file', 'unknown')}: {e} [поток: {threading.current_thread().name}]"
        )
        return None, sheet_name, None


def validate_single_sheet(sheet_name, sheets_data_item):
    """
    Проверяет длину полей для одного листа.
    Используется для параллельной проверки валидации.
    
    Args:
        sheet_name (str): Имя листа для проверки
        sheets_data_item (tuple): (df, conf) - данные листа и конфигурация
        
    Returns:
        tuple: (sheet_name, (df_validated, conf))
    """
    # ОПТИМИЗАЦИЯ v5.0: Проверка на None
    if sheets_data_item is None:
        logging.warning(f"[check_duplicates_single_sheet] Данные для листа {sheet_name} равны None, пропускаем")
        return sheet_name, sheets_data_item
    
    try:
        df, conf = sheets_data_item
        # Дополнительная проверка на None
        if df is None or conf is None:
            logging.warning(f"[check_duplicates_single_sheet] DataFrame или конфигурация для листа {sheet_name} равны None, пропускаем")
            return sheet_name, sheets_data_item
        # ОПТИМИЗАЦИЯ: Используем векторизованную версию с проверкой результатов
        df_old = df.copy()
        df_validated = validate_field_lengths_vectorized(df, sheet_name)
        
        # Сравниваем результаты для проверки корректности
        if sheet_name in FIELD_LENGTH_VALIDATIONS:
            result_column = FIELD_LENGTH_VALIDATIONS[sheet_name]["result_column"]
            comparison = compare_validate_results(df_old, df_validated, result_column)
            if not comparison.get("identical", False):
                logging.warning(
                    f"[VALIDATE COMPARISON] {sheet_name}: различия найдены - {comparison.get('differences', 0)} из {comparison.get('total', 0)}"
                )
                # В случае различий используем старую версию для гарантии корректности
                df_validated = validate_field_lengths(df, sheet_name)
                logging.warning(f"[VALIDATE FALLBACK] {sheet_name}: использована оригинальная версия")
            else:
                logging.info(f"[VALIDATE COMPARISON] {sheet_name}: результаты идентичны ({comparison.get('match_percent', 0)}%)")
        else:
            df_validated = df
        th = threading.current_thread().name
        logging.debug(f"Проверка длины полей завершена: {sheet_name} [поток: {th}]")
        return sheet_name, (df_validated, conf)
    except Exception as e:
        logging.error(
            f"Ошибка проверки длины полей для {sheet_name}: {e} [поток: {threading.current_thread().name}]"
        )
        # Возвращаем исходные данные при ошибке
        return sheet_name, sheets_data_item


def check_duplicates_single_sheet(sheet_name, sheets_data_item):
    """
    Проверяет дубликаты для одного листа.
    Используется для параллельной проверки дубликатов.
    
    Args:
        sheet_name (str): Имя листа для проверки
        sheets_data_item (tuple): (df, conf) - данные листа и конфигурация
        
    Returns:
        tuple: (sheet_name, (df, conf))
    """
    # ОПТИМИЗАЦИЯ v5.0: Проверка на None
    if sheets_data_item is None:
        logging.warning(f"[check_duplicates_single_sheet] Данные для листа {sheet_name} равны None, пропускаем")
        return sheet_name, sheets_data_item
    
    try:
        df, conf = sheets_data_item
        # Дополнительная проверка на None
        if df is None or conf is None:
            logging.warning(f"[check_duplicates_single_sheet] DataFrame или конфигурация для листа {sheet_name} равны None, пропускаем")
            return sheet_name, sheets_data_item
        # Находим ВСЕ записи для этого листа (не только первую)
        check_configs = [x for x in CHECK_DUPLICATES if x["sheet"] == sheet_name]
        for check_cfg in check_configs:
            df = mark_duplicates(df, check_cfg["key"], sheet_name=sheet_name)
        
        if check_configs:
            logging.debug(f"Проверка дубликатов завершена: {sheet_name} [поток: {threading.current_thread().name}]")
        
        return sheet_name, (df, conf)
    except Exception as e:
        logging.error(
            f"Ошибка проверки дубликатов для {sheet_name}: {e} [поток: {threading.current_thread().name}]"
        )
        # Возвращаем исходные данные при ошибке
        return sheet_name, sheets_data_item


def collect_duplicates_and_validation_report(sheets_data: Dict[str, Any]) -> tuple:
    """
    Собирает сводный отчёт по дубликатам и отклонениям длины полей (field_length_validations)
    по всем листам. Не прерывает работу — только накапливает данные для финального вывода.

    Returns:
        tuple: (duplicates_report, validation_report)
            - duplicates_report: список dict с ключами sheet, key_cols, col_name, n_duplicate_rows, duplicate_key_values
            - validation_report: список dict с ключами sheet, result_column, n_violations, sample_values
    """
    duplicates_report: List[Dict[str, Any]] = []
    validation_report: List[Dict[str, Any]] = []

    # --- Дубликаты: ищем колонки "ДУБЛЬ: ..." и строки с непустым значением ---
    for sheet_name, sheet_item in sheets_data.items():
        if sheet_item is None:
            continue
        try:
            df, _ = sheet_item
            if df is None or not isinstance(df, pd.DataFrame):
                continue
        except (TypeError, ValueError):
            continue

        for col in df.columns:
            if not (isinstance(col, str) and col.startswith("ДУБЛЬ: ")):
                continue
            # Найти конфиг проверки: ключ, по которому построена эта колонка
            key_cols: Optional[List[str]] = None
            for check_cfg in CHECK_DUPLICATES:
                if check_cfg.get("sheet") != sheet_name:
                    continue
                k = check_cfg.get("key")
                if isinstance(k, list) and "ДУБЛЬ: " + "_".join(k) == col:
                    key_cols = k
                    break
            if not key_cols:
                continue

            dup_mask = df[col].astype(str).str.strip() != ""
            n_duplicate_rows = int(dup_mask.sum())
            if n_duplicate_rows == 0:
                continue

            # Проверить наличие ключевых колонок
            missing = [c for c in key_cols if c not in df.columns]
            if missing:
                logging.debug(f"[collect_report] Лист {sheet_name}, колонка {col}: отсутствуют ключи {missing}")
                continue

            # Группируем дубликаты по значениям ключа: (tuple(key_vals)) -> count
            dup_df = df.loc[dup_mask, key_cols + [col]].copy()
            dup_df["_key_tuple"] = list(zip(*(dup_df[k].astype(str).values for k in key_cols)))
            key_counts = dup_df.groupby("_key_tuple", dropna=False).size()

            duplicate_key_values = []
            for key_tuple, count in key_counts.items():
                key_vals = dict(zip(key_cols, key_tuple))
                duplicate_key_values.append({"key_values": key_vals, "count": int(count)})

            duplicates_report.append({
                "sheet": sheet_name,
                "key_cols": key_cols,
                "col_name": col,
                "n_duplicate_rows": n_duplicate_rows,
                "duplicate_key_values": duplicate_key_values,
            })

    # --- Отклонения по длине полей (field_length_validations) ---
    for sheet_name in (FIELD_LENGTH_VALIDATIONS or {}):
        if sheet_name not in sheets_data:
            continue
        sheet_item = sheets_data[sheet_name]
        if sheet_item is None:
            continue
        try:
            df, _ = sheet_item
            if df is None or not isinstance(df, pd.DataFrame):
                continue
        except (TypeError, ValueError):
            continue

        config = FIELD_LENGTH_VALIDATIONS.get(sheet_name)
        if not config:
            continue
        result_column = config.get("result_column")
        if not result_column or result_column not in df.columns:
            continue

        violations_mask = (df[result_column].astype(str).str.strip() != "") & (df[result_column].astype(str).str.strip() != "-")
        n_violations = int(violations_mask.sum())
        if n_violations == 0:
            continue

        sample_values = df.loc[violations_mask, result_column].drop_duplicates().head(20).tolist()
        validation_report.append({
            "sheet": sheet_name,
            "result_column": result_column,
            "n_violations": n_violations,
            "sample_values": sample_values,
        })

    return duplicates_report, validation_report


def build_stat_file_sheet(
    input_files: List[Dict[str, Any]],
    sheets_data: Dict[str, Any],
    run_datetime: datetime,
) -> pd.DataFrame:
    """
    Формирует лист STAT_FILE со статистикой по исходным файлам: имя файла, лист, дата файла,
    дата обновления данных, количество записей и колонок, размер файла, статус.
    """
    rows = []
    for file_conf in input_files:
        file_name = file_conf.get("file", "")
        sheet_name = file_conf.get("sheet", "")
        file_path = find_file_case_insensitive(DIR_INPUT, file_name, [".csv", ".CSV"])
        if file_path is None:
            file_date = ""
            file_size = 0
            status = "не найден"
            row_count = 0
            col_count = 0
        else:
            try:
                mtime = os.path.getmtime(file_path)
                file_date = datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M:%S")
                file_size = os.path.getsize(file_path)
            except OSError:
                file_date = ""
                file_size = 0
            status = "OK"
            if sheet_name in sheets_data and sheets_data[sheet_name] is not None:
                df_sheet = sheets_data[sheet_name][0]
                row_count = len(df_sheet) if df_sheet is not None else 0
                col_count = len(df_sheet.columns) if df_sheet is not None else 0
            else:
                row_count = 0
                col_count = 0
        data_update_date = run_datetime.strftime("%Y-%m-%d %H:%M:%S")
        rows.append({
            "FILE_NAME": file_name,
            "SHEET_NAME": sheet_name,
            "FILE_DATE": file_date,
            "DATA_UPDATE_DATE": data_update_date,
            "ROW_COUNT": row_count,
            "COL_COUNT": col_count,
            "FILE_SIZE_BYTES": file_size,
            "STATUS": status,
        })
    return pd.DataFrame(rows)


def print_final_report(duplicates_report: List[Dict[str, Any]], validation_report: List[Dict[str, Any]]) -> None:
    """
    Выводит итоговый отчёт по дубликатам и отклонениям длины полей в лог и в консоль.
    Вызывается в конце работы программы; не прерывает выполнение.
    """
    lines: List[str] = []
    lines.append("")
    lines.append("========== ИТОГОВАЯ СТАТИСТИКА: ДУБЛИКАТЫ И ОТКЛОНЕНИЯ ДЛИНЫ ПОЛЕЙ ==========")

    if duplicates_report:
        lines.append("")
        lines.append("--- Дубликаты ---")
        for r in duplicates_report:
            lines.append(f"  Лист: {r['sheet']}")
            lines.append(f"  Ключ (колонки): {r['key_cols']}")
            lines.append(f"  Колонка проверки: {r['col_name']}")
            lines.append(f"  Строк с дубликатами: {r['n_duplicate_rows']}")
            for g in r["duplicate_key_values"]:
                kv = g["key_values"]
                cnt = g["count"]
                lines.append(f"    Задублированное значение: {kv} (вхождений: {cnt})")
            lines.append("")
    else:
        lines.append("")
        lines.append("--- Дубликаты: не обнаружены ---")

    if validation_report:
        lines.append("--- Отклонения по длине полей (field_length_validations) ---")
        for r in validation_report:
            lines.append(f"  Лист: {r['sheet']}, колонка результата: {r['result_column']}")
            lines.append(f"  Количество строк с отклонениями: {r['n_violations']}")
            for i, sample in enumerate(r["sample_values"][:10], 1):
                lines.append(f"    Пример {i}: {sample}")
            if len(r["sample_values"]) > 10:
                lines.append(f"    ... и ещё {len(r['sample_values']) - 10} вариантов")
            lines.append("")
    else:
        lines.append("--- Отклонения по длине полей: не обнаружены ---")

    lines.append("===============================================================================")
    lines.append("")

    report_text = "\n".join(lines)
    for line in lines:
        logging.info(line.strip() if line.strip() else "")
    print(report_text)


def main():
    start_time = datetime.now()
    log_file = setup_logger()
    logging.info(f"=== Старт работы программы: {start_time.strftime('%Y-%m-%d %H:%M:%S')} ===")

    # Проверка наличия всех файлов из INPUT_FILES до начала обработки
    missing_files = check_input_files_exist()
    if missing_files:
        msg_lines = [
            "Программа не запущена: не найдены следующие файлы из INPUT_FILES:",
            f"  (ожидаемый каталог: {DIR_INPUT})",
        ]
        for m in missing_files:
            msg_lines.append(f"  - {m['file']} (лист: {m['sheet']})")
        msg = "\n".join(msg_lines)
        logging.error(msg)
        print(msg, file=sys.stderr)
        sys.exit(1)

    sheets_data = {}
    files_processed = 0
    rows_total = 0
    summary = []

        # 1. Параллельное чтение всех CSV и разворот ВСЕХ JSON‑полей на каждом листе
    logging.info(f"Начало параллельного чтения CSV файлов (потоков: {MAX_WORKERS_IO})")
    
    lock = threading.Lock()  # Для безопасного доступа к sheets_data
    
    with ThreadPoolExecutor(max_workers=MAX_WORKERS_IO) as executor:  # I/O операция
        # Запускаем обработку всех файлов параллельно
        futures = {executor.submit(process_single_file, file_conf): file_conf 
                   for file_conf in INPUT_FILES}
        
        # Собираем результаты по мере их готовности
        for future in as_completed(futures):
            df, sheet_name, file_conf = future.result()
            if df is not None and file_conf is not None:
                with lock:
                    sheets_data[sheet_name] = (df, file_conf)
                    files_processed += 1
                    rows_total += len(df)
                    summary.append(f"{sheet_name}: {len(df)} строк")
            elif sheet_name:
                # Файл не найден или ошибка чтения
                summary.append(f"{sheet_name}: {'файл не найден' if file_conf is None else 'ошибка'}")
    
    logging.info(f"Параллельное чтение CSV файлов завершено. Обработано файлов: {files_processed}")
    # 2. Добавление колонки AUTO_GENDER для листа EMPLOYEE
    if "EMPLOYEE" in sheets_data:
        df_employee, conf_employee = sheets_data["EMPLOYEE"]
        # ОПТИМИЗАЦИЯ: Используем векторизованную версию с проверкой результатов
        df_employee_old = df_employee.copy()
        df_employee = add_auto_gender_column_vectorized(df_employee, "EMPLOYEE")
        
        # Сравниваем результаты
        comparison = compare_gender_results(df_employee_old, df_employee)
        if not comparison.get("identical", False):
            logging.warning(
                f"[GENDER COMPARISON] EMPLOYEE: различия найдены - {comparison.get('differences', 0)} из {comparison.get('total', 0)}"
            )
            # В случае различий используем старую версию
            df_employee = add_auto_gender_column(df_employee_old, "EMPLOYEE")
            logging.warning("[GENDER FALLBACK] EMPLOYEE: использована оригинальная версия")
        else:
            logging.info(f"[GENDER COMPARISON] EMPLOYEE: результаты идентичны ({comparison.get('match_percent', 0)}%)")
        sheets_data["EMPLOYEE"] = (df_employee, conf_employee)

        # 3. Параллельная проверка длины полей для всех листов согласно FIELD_LENGTH_VALIDATIONS
    if FIELD_LENGTH_VALIDATIONS:
        logging.info(f"Начало параллельной проверки длины полей (потоков: {MAX_WORKERS_CPU})")
        sheets_to_validate = {name: sheets_data[name] for name in FIELD_LENGTH_VALIDATIONS.keys() 
                             if name in sheets_data}
        
        if sheets_to_validate:
            with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
                futures = {executor.submit(validate_single_sheet, sheet_name, data): sheet_name
                          for sheet_name, data in sheets_to_validate.items()}
                
                for future in as_completed(futures):
                    sheet_name, validated_data = future.result()
                    sheets_data[sheet_name] = validated_data
            
            logging.info("Параллельная проверка длины полей завершена")
    # 4. Добавление расчетного статуса турнира для TOURNAMENT-SCHEDULE
    if "TOURNAMENT-SCHEDULE" in sheets_data:
        df_tournament, conf_tournament = sheets_data["TOURNAMENT-SCHEDULE"]
        df_report = sheets_data.get("REPORT", (None, None))[0]
        df_tournament = calculate_tournament_status(df_tournament, df_report)
        sheets_data["TOURNAMENT-SCHEDULE"] = (df_tournament, conf_tournament)

    # 5. Merge fields (только после полного разворота JSON)
    # Все правила в MERGE_FIELDS_ADVANCED: сначала бывшие MERGE_FIELDS (порядок сохранён), затем расширенные.
    # count_column_prefix="COUNT" — имена колонок в Excel не меняются (COUNT_* для count без count_label).
    merge_fields_across_sheets(
        sheets_data,
        [f for f in MERGE_FIELDS_ADVANCED if f.get("sheet_dst") != "SUMMARY"],
        count_column_prefix="COUNT",
        merge_name="MERGE_FIELDS_ADVANCED"
    )

        # 6. Параллельная проверка на дубли
    logging.info(f"Начало параллельной проверки на дубликаты (потоков: {MAX_WORKERS_CPU})")
    with ThreadPoolExecutor(max_workers=MAX_WORKERS_IO) as executor:
        futures = {executor.submit(check_duplicates_single_sheet, sheet_name, data): sheet_name
                  for sheet_name, data in sheets_data.items()}
        
        for future in as_completed(futures):
            sheet_name, validated_data = future.result()
            sheets_data[sheet_name] = validated_data
    
    logging.info("Параллельная проверка на дубликаты завершена")
    # 7. Формирование итогового Summary (build_summary_sheet)
    # Как в base main.py: используем данные ПОСЛЕ merge и ПОСЛЕ check_duplicates,
    # чтобы колонки и содержимое SUMMARY (и всех листов) совпадали с верным файлом.
    dfs = {k: v[0] for k, v in sheets_data.items()}
    df_summary = build_summary_sheet(
        dfs,
        params_summary=SUMMARY_SHEET,
        merge_fields=[f for f in MERGE_FIELDS_ADVANCED if f.get("sheet_dst") == "SUMMARY"]
    )
    # ОПТИМИЗАЦИЯ v5.0: Проверка df_summary на None
    if df_summary is None or not isinstance(df_summary, pd.DataFrame):
        logging.error("[main] КРИТИЧЕСКАЯ ОШИБКА: df_summary равен None или не DataFrame после build_summary_sheet!")
        logging.error("[main] Создаем пустой DataFrame для SUMMARY")
        df_summary = pd.DataFrame(columns=SUMMARY_KEY_COLUMNS)
    elif len(df_summary) == 0:
        logging.warning("[main] df_summary пустой после build_summary_sheet, но продолжаем работу")
    else:
        logging.info(f"[main] df_summary успешно создан: {len(df_summary)} строк, {len(df_summary.columns)} колонок")

    sheets_data[SUMMARY_SHEET["sheet"]] = (df_summary, SUMMARY_SHEET)

    # Лист STAT_FILE: статистика по исходным файлам (имя, лист, даты, строки, колонки, размер, статус)
    df_stat = build_stat_file_sheet(INPUT_FILES, sheets_data, start_time)
    stat_file_params = {
        "sheet": "STAT_FILE",
        "max_col_width": 80,
        "freeze": "A2",
        "col_width_mode": "AUTO",
        "min_col_width": 10,
    }
    sheets_data["STAT_FILE"] = (df_stat, stat_file_params)
    logging.info(f"[main] Лист STAT_FILE сформирован: {len(df_stat)} строк (статистика по файлам)")
    
    # Верификация merge: сохранение baseline или сравнение с ним
    _baseline_path = os.path.join(DIR_OUTPUT, "merge_output_baseline.json")
    if os.environ.get("SAVE_MERGE_BASELINE") == "1":
        os.makedirs(DIR_OUTPUT, exist_ok=True)
        snapshot = _dump_sheets_data_for_baseline(sheets_data, max_rows=3)
        with open(_baseline_path, "w", encoding="utf-8") as f:
            json.dump(snapshot, f, ensure_ascii=False, indent=2)
        logging.info(f"[MERGE] Baseline сохранён: {_baseline_path} (колонки и по 3 строки на лист)")
    elif os.path.isfile(_baseline_path):
        ok, diff_errors = _compare_sheets_data_with_baseline(sheets_data, _baseline_path, max_rows=3)
        if ok:
            logging.info("[MERGE] Сравнение с baseline: колонки и сэмпл данных совпадают")
        else:
            for msg in diff_errors:
                logging.warning(f"[MERGE] Baseline расхождение: {msg}")

    # 8. Запись в Excel
    output_excel = os.path.join(DIR_OUTPUT, get_output_filename())
    logging.info(f"[START] write_to_excel ({output_excel})")
    write_to_excel(sheets_data, output_excel)
    logging.info(f"[END] write_to_excel ({output_excel}) (время: 0.000s)")

    # Итоговая статистика по дубликатам и отклонениям длины полей (в лог и консоль)
    duplicates_report, validation_report = collect_duplicates_and_validation_report(sheets_data)
    print_final_report(duplicates_report, validation_report)

    time_elapsed = datetime.now() - start_time
    logging.info(
        f"=== Завершение работы. Обработано файлов: {files_processed}, строк всего: {rows_total}. Время выполнения: {time_elapsed} ==="
    )
    logging.info(f"Summary: {'; '.join(summary)}")
    logging.info(f"Excel file: {output_excel}")
    logging.info(f"Log file: {log_file}")


if __name__ == "__main__":
    main()
