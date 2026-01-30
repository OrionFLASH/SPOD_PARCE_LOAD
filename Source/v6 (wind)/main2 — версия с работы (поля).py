import os
import sys
import pandas as pd
import logging
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
from time import time
import json
import re

# === Глобальные константы и переменные ===
# Каталоги
DIR_INPUT = r"C:\Users\01803187\Documents\MyProject\SPOD_CSV_PARCER\SPOD"
DIR_OUTPUT = r"C:\Users\01803187\Documents\MyProject\SPOD_CSV_PARCER\OUT"
DIR_LOGS = r"C:\Users\01803187\Documents\MyProject\SPOD_CSV_PARCER\LOGS"

# Входные файлы (имя без расширения)
# Соответствие: Имя листа, максимальная ширина колонки, закрепление
INPUT_FILES = [
    {
        "file": "CONTEST-DATA (PROM) 2025-07-24 v4",
        "sheet": "CONTEST-DATA",
        "max_col_width": 120,
        "freeze": "C2"
    },
    {
        "file": "GROUP (PROM) 2025-07-14 v0",
        "sheet": "GROUP",
        "max_col_width": 20,
        "freeze": "C2"
    },
    {
        "file": "INDICATOR (PROM) 2025-07-14 v0",
        "sheet": "INDICATOR",
        "max_col_width": 20,
        "freeze": "B2"
    },
    {
        "file": "REPORT (PROM-KMKKSB) 2025-07-25 v5",
        "sheet": "REPORT",
        "max_col_width": 50,
        "freeze": "D2"
    },
    {
        "file": "REWARD (PROM) 2025-07-24 v1",
        "sheet": "REWARD",
        "max_col_width": 140,
        "freeze": "B2"
    },
    {
        "file": "REWARD-LINK (PROM) 2025-07-14 v0",
        "sheet": "REWARD-LINK",
        "max_col_width": 30,
        "freeze": "A2"
    },
    {
        "file": "SVD_KB_DM_GAMIFICATION_ORG_UNIT_V20 2025_07_11 v1",
        "sheet": "ORG_UNIT_V20",
        "max_col_width": 60,
        "freeze": "A2"
    },
    {
        "file": "TOURNAMENT-SCHEDULE (PROM) 2025-07-25 v5",
        "sheet": "TOURNAMENT-SCHEDULE",
        "max_col_width": 120,
        "freeze": "B2"
    },
    {
        "file": "PROM_USER_ROLE 2025-07-21 v0",
        "sheet": "USER_ROLE",
        "max_col_width": 60,
        "freeze": "D2"
    },
    {
        "file": "PROM_USER_ROLE SB 2025-07-21 v0",
        "sheet": "USER_ROLE SB",
        "max_col_width": 60,
        "freeze": "D2"
    },
    {
        "file": "employee_PROM_2025-07-21 v0",
        "sheet": "EMPLOYEE",
        "max_col_width": 70,
        "freeze": "F2"
    }
]

SUMMARY_SHEET = {
    "sheet": "SUMMARY",
    "max_col_width": 70,
    "freeze": "F2"
}

# Логирование: уровень, шаблоны, имена
LOG_LEVEL = "INFO"  # "INFO" или "DEBUG"
LOG_BASE_NAME = "LOGS"
LOG_MESSAGES = {
    "start":                "=== Старт работы программы: {time} ===",
    "reading_file":         "Загрузка файла: {file_path}",
    "read_ok":              "Файл успешно загружен: {file_path}, строк: {rows}, колонок: {cols}",
    "read_fail":            "Ошибка загрузки файла: {file_path}. {error}",
    "sheet_written":        "Лист Excel сформирован: {sheet} (строк: {rows}, колонок: {cols})",
    "finish":               "=== Завершение работы. Обработано файлов: {files}, строк всего: {rows_total}. Время выполнения: {time_elapsed} ===",
    "summary":              "Summary: {summary}",
    "func_start":           "[START] {func} {params}",
    "func_end":             "[END] {func} {params} (время: {time:.3f}s)",
    "func_error":           "[ERROR] {func} {params} — {error}",
    "json_flatten_start":   "Разворачивание колонки {column} (строк: {rows})",
    "json_flatten_end":     "Развёрнуто {n_cols} колонок из {n_keys} ключей, ошибок JSON: {n_errors}, строк: {rows}, время: {time:.3f}s",
    "json_flatten_error":   "Ошибка разбора JSON (строка {row}): {error}",
    "debug_columns":        "[DEBUG] {sheet}: колонки после разворачивания: {columns}",
    "debug_head":           "[DEBUG] {sheet}: первые строки после разворачивания:\n{head}",
    "field_joined":         "Колонка {column} присоединена из {src_sheet} по ключу {dst_key} -> {src_key}",
    "field_missing":        "Колонка {column} не добавлена: нет листа {src_sheet} или ключей {src_key}",
    "fields_summary":       "Итоговая структура: {rows} строк, {cols} колонок",
    "duplicates_start":     "[START] Проверка дублей: {sheet}, ключ: {keys}",
    "duplicates_found":     "[INFO] Дублей найдено: {count} на листе {sheet} по ключу {keys}",
    "duplicates_error":     "[ERROR] Ошибка при поиске дублей: {sheet}, ключ: {keys}: {error}",
    "duplicates_end":       "[END] Проверка дублей: {sheet}, время: {time:.3f}s",
    "color_scheme_applied": "[INFO] Цветовая схема применена: лист {sheet}, колонка {col}, стиль {scope}, цвет {color}"
    ,"json_flatten_summary": "[INFO] {column} → новых колонок: {count}"
    ,"json_flatten_keys":    "[INFO] Все новые колонки: {keys}"
    ,"csv_sample":           "[DEBUG] CSV {file} поле {column}: {sample}"
    ,"excel_path":           "Excel file: {path}"
    ,"log_path":             "Log file: {path}"
    ,"json_flatten_done":    "[JSON FLATTEN] {sheet}: поле '{column}' развернуто с префиксом '{prefix}'"
    ,"json_flatten_missing": "[JSON FLATTEN] {sheet}: поле '{column}' не найдено в колонках!"
    ,"missing_column":       "[add_fields_to_sheet] Колонка {column} не найдена в {sheet}, создаём пустую."
    ,"missing_key":          "[add_fields_to_sheet] Ключевая колонка {key} не найдена в {sheet}, создаём пустую."
    ,"safe_json_error":      "[safe_json_loads] Ошибка: {error} | Исходная строка: {line}"
}

# --- Общие префиксы для колонок JSON ---
PREFIX_CONTEST_FEATURE = "CONTEST_FEATURE"
PREFIX_ADD_DATA = "ADD_DATA"
PREFIX_REWARD_LINK = "REWARD_LINK => "
COL_REWARD_LINK_CONTEST_CODE = f"{PREFIX_REWARD_LINK}CONTEST_CODE"

MERGE_FIELDS = [
    # REPORT: добавляем CONTEST_TYPE и FULL_NAME из CONTEST-DATA
    {
        "sheet_src": "CONTEST-DATA",
        "sheet_dst": "REPORT",
        "src_key": ["CONTEST_CODE"],
        "dst_key": ["CONTEST_CODE"],
        "column": ["CONTEST_TYPE", "FULL_NAME", "BUSINESS_STATUS", "BUSINESS_BLOCK"],
        "mode": "value"
    },
    # REPORT: добавляем даты из TOURNAMENT-SCHEDULE
    {
        "sheet_src": "TOURNAMENT-SCHEDULE",
        "sheet_dst": "REPORT",
        "src_key": ["TOURNAMENT_CODE"],
        "dst_key": ["TOURNAMENT_CODE"],
        "column": ["END_DT", "RESULT_DT", "TOURNAMENT_STATUS"],
        "mode": "value"
    },
    # REWARD: добавляем CONTEST_CODE из REWARD-LINK по REWARD_CODE
    {
        "sheet_src": "REWARD-LINK",
        "sheet_dst": "REWARD",
        "src_key": ["REWARD_CODE"],
        "dst_key": ["REWARD_CODE"],
        "column": ["CONTEST_CODE"],
        "mode": "value"
    },
    {
        "sheet_src": "CONTEST-DATA",
        "sheet_dst": "TOURNAMENT-SCHEDULE",
        "src_key": ["CONTEST_CODE"],
        "dst_key": ["CONTEST_CODE"],
        "column": ["FULL_NAME", "BUSINESS_BLOCK", "CONTEST_TYPE",  "BUSINESS_STATUS"],
        "mode": "value"
    },
    {
        "sheet_src": "REPORT",
        "sheet_dst": "TOURNAMENT-SCHEDULE",
        "src_key": ["CONTEST_CODE", "TOURNAMENT_CODE"],
        "dst_key": ["CONTEST_CODE", "TOURNAMENT_CODE"],
        "column": ["CONTEST_DATE"],
        "mode": "value"
    },
    {
        "sheet_src": "REPORT",
        "sheet_dst": "TOURNAMENT-SCHEDULE",
        "src_key": ["CONTEST_CODE", "TOURNAMENT_CODE"],
        "dst_key": ["CONTEST_CODE", "TOURNAMENT_CODE"],
        "column": ["CONTEST_DATE"],
        "mode": "count"
    },

    # SUMMARY: из CONTEST-DATA по CONTEST_CODE — основные поля
    {
        "sheet_src": "CONTEST-DATA",
        "sheet_dst": "SUMMARY",
        "src_key": ["CONTEST_CODE"],
        "dst_key": ["CONTEST_CODE"],
        "column": [
            "FULL_NAME",
            f"{PREFIX_CONTEST_FEATURE} => momentRewarding",
            "FACTOR_MATCH",
            "PLAN_MOD_VALUE",
            "BUSINESS_BLOCK",
            f"{PREFIX_CONTEST_FEATURE} => tournamentStartMailing",
            f"{PREFIX_CONTEST_FEATURE} => tournamentEndMailing",
            f"{PREFIX_CONTEST_FEATURE} => tournamentRewardingMailing",
            f"{PREFIX_CONTEST_FEATURE} => tournamentLikeMailing"
        ],
        "mode": "value"
    },
    # SUMMARY: из GROUP по составному ключу
    {
        "sheet_src": "GROUP",
        "sheet_dst": "SUMMARY",
        "src_key": ["CONTEST_CODE", "GROUP_CODE", "GROUP_VALUE"],
        "dst_key": ["CONTEST_CODE", "GROUP_CODE", "GROUP_VALUE"],
        "column": [
            "GET_CALC_CRITERION",
            "ADD_CALC_CRITERION",
            "ADD_CALC_CRITERION_2"
        ],
        "mode": "value"
    },
    # SUMMARY: из INDICATOR по CONTEST_CODE
    {
        "sheet_src": "INDICATOR",
        "sheet_dst": "SUMMARY",
        "src_key": ["CONTEST_CODE"],
        "dst_key": ["CONTEST_CODE"],
        "column": [
            "INDICATOR_MARK_TYPE",
            "INDICATOR_MATCH",
            "INDICATOR_VALUE"
        ],
        "mode": "value"
    },
    # SUMMARY: из TOURNAMENT-SCHEDULE по TOURNAMENT_CODE
    {
        "sheet_src": "TOURNAMENT-SCHEDULE",
        "sheet_dst": "SUMMARY",
        "src_key": ["TOURNAMENT_CODE"],
        "dst_key": ["TOURNAMENT_CODE"],
        "column": [
            "START_DT",
            "END_DT",
            "RESULT_DT",
            "TOURNAMENT_STATUS",
            "TARGET_TYPE"
        ],
        "mode": "value"
    },
    # SUMMARY: CONTEST_DATE из REPORT по TOURNAMENT_CODE
    {
        "sheet_src": "REPORT",
        "sheet_dst": "SUMMARY",
        "src_key": ["TOURNAMENT_CODE"],
        "dst_key": ["TOURNAMENT_CODE"],
        "column": [
            "CONTEST_DATE"
        ],
        "mode": "value"
    },
    # SUMMARY: сколько в REPORT строк по паре TOURNAMENT_CODE + CONTEST_CODE
    {
        "sheet_src": "REPORT",
        "sheet_dst": "SUMMARY",
        "src_key": ["TOURNAMENT_CODE", "CONTEST_CODE"],
        "dst_key": ["TOURNAMENT_CODE", "CONTEST_CODE"],
        "column": [
            "CONTEST_DATE"
        ],
        "mode": "count"
    },
    # SUMMARY: все нужные поля из REWARD по составному ключу
    {
        "sheet_src": "REWARD",
        "sheet_dst": "SUMMARY",
        "src_key": [COL_REWARD_LINK_CONTEST_CODE, "REWARD_CODE"],  # ПРОБЕЛ после =>
        "dst_key": ["CONTEST_CODE", "REWARD_CODE"],
        "column": [
            f"{PREFIX_ADD_DATA} => rewardAgainGlobal",
            f"{PREFIX_ADD_DATA} => rewardAgainTournament",
            f"{PREFIX_ADD_DATA} => outstanding",
            f"{PREFIX_ADD_DATA} => teamNews",
            f"{PREFIX_ADD_DATA} => singleNews"
        ],
        "mode": "value"
    }
]



SUMMARY_KEY_DEFS = [
    {"sheet": "CONTEST-DATA",    "cols": ["CONTEST_CODE"]},
    {"sheet": "TOURNAMENT-SCHEDULE", "cols": ["TOURNAMENT_CODE", "CONTEST_CODE"]},
    {"sheet": "REWARD-LINK",     "cols": ["REWARD_CODE", "CONTEST_CODE"]},
    {"sheet": "GROUP",           "cols": ["GROUP_CODE", "CONTEST_CODE", "GROUP_VALUE"]},
    {"sheet": "REWARD",          "cols": ["REWARD_CODE"]},
]

# Построить упорядоченный список всех уникальных ключей
SUMMARY_KEY_COLUMNS = []
for entry in SUMMARY_KEY_DEFS:
    for col in entry["cols"]:
        if col not in SUMMARY_KEY_COLUMNS:
            SUMMARY_KEY_COLUMNS.append(col)


COLOR_SCHEME = [
    # --- ИСХОДНЫЕ ДАННЫЕ (загружаются из CSV) — пастельный голубой ---
    {
        "group": "Исходные данные",
        "header_bg": "E6F3FF",  # пастельный голубой - приятный для глаз
        "header_fg": "2C3E50",  # тёмно-серый для лучшей читаемости
        "column_bg": None,
        "column_fg": None,
        "style_scope": "header",
        "sheets": ["CONTEST-DATA", "GROUP", "INDICATOR", "REPORT", "REWARD", "REWARD-LINK", "TOURNAMENT-SCHEDULE", "ORG_UNIT_V20", "USER_ROLE", "USER_ROLE SB", "EMPLOYEE"],
        "columns": [],  # все колонки (если не указано — все)
        # Назначение: базовые поля из CSV файлов
    },

    # --- ИСХОДНЫЕ JSON ПОЛЯ (CONTEST_FEATURE, REWARD_ADD_DATA) — тёмно-оранжевый со светлыми буквами ---
    {
        "group": "JSON source columns",
        "header_bg": "FF8C42",  # тёмно-оранжевый - самый верхний уровень JSON полей
        "header_fg": "FFFFFF",  # белый текст для контраста
        "column_bg": None,
        "column_fg": None,
        "style_scope": "header",
        "sheets": ["CONTEST-DATA", "REWARD"],
        "columns": ["CONTEST_FEATURE", "REWARD_ADD_DATA"],
        # Назначение: исходные поля с JSON, которые разворачиваются
    },

    # --- РАЗВОРАЧИВАЕМЫЕ JSON ПОЛЯ ПЕРВОГО УРОВНЯ — светлее исходных ---
    {
        "group": "JSON expanded level 1",
        "header_bg": "FFB366",  # светло-оранжевый - светлее исходных JSON полей
        "header_fg": "2C3E50",  # тёмно-серый для читаемости
        "column_bg": None,
        "column_fg": None,
        "style_scope": "header",
        "sheets": ["CONTEST-DATA", "REWARD"],
        "columns": [
            # CONTEST_FEATURE развёрнутые поля
            "CONTEST_FEATURE => momentRewarding", "CONTEST_FEATURE => tournamentStartMailing", "CONTEST_FEATURE => tournamentEndMailing",
            "CONTEST_FEATURE => tournamentRewardingMailing", "CONTEST_FEATURE => tournamentLikeMailing", "CONTEST_FEATURE => capacity",
            "CONTEST_FEATURE => tournamentListMailing", "CONTEST_FEATURE => vid", "CONTEST_FEATURE => tbVisible", "CONTEST_FEATURE => tbHidden",
            "CONTEST_FEATURE => persomanNumberVisible", "CONTEST_FEATURE => typeRewarding", "CONTEST_FEATURE => masking",
            "CONTEST_FEATURE => minNumber", "CONTEST_FEATURE => businessBlock", "CONTEST_FEATURE => accuracy", "CONTEST_FEATURE => gosbHidden",
            "CONTEST_FEATURE => preferences", "CONTEST_FEATURE => persomanNumberHidden", "CONTEST_FEATURE => gosbVisible", "CONTEST_FEATURE => feature",
            # ADD_DATA развёрнутые поля первого уровня
            "ADD_DATA => refreshOldNews", "ADD_DATA => fileName", "ADD_DATA => rewardRule", "ADD_DATA => bookingRequired", "ADD_DATA => outstanding",
            "ADD_DATA => teamNews", "ADD_DATA => singleNews", "ADD_DATA => rewardAgainGlobal", "ADD_DATA => rewardAgainTournament",
            "ADD_DATA => isGrouping", "ADD_DATA => tagEndDT", "ADD_DATA => itemAmount", "ADD_DATA => isGroupingTitle",
            "ADD_DATA => itemLimitCount", "ADD_DATA => recommendationLevel", "ADD_DATA => isGroupingName", "ADD_DATA => ignoreConditions",
            "ADD_DATA => masterBadge", "ADD_DATA => priority", "ADD_DATA => nftFlg", "ADD_DATA => itemMinShow", "ADD_DATA => itemFeature",
            "ADD_DATA => itemLimitPeriod", "ADD_DATA => businessBlock", "ADD_DATA => parentRewardCode", "ADD_DATA => deliveryRequired",
            "ADD_DATA => feature", "ADD_DATA => itemGroupAmount", "ADD_DATA => seasonItem", "ADD_DATA => isGroupingTultip", "ADD_DATA => tagColor",
            "ADD_DATA => commingSoon", "ADD_DATA => tournamentTeam", "ADD_DATA => hidden"
        ],
        # Назначение: поля первого уровня развёртывания JSON (светлее исходных)
    },

    # --- РАЗВОРАЧИВАЕМЫЕ JSON ПОЛЯ ВТОРОГО УРОВНЯ — темнее исходных, но светлее других ---
    {
        "group": "JSON expanded level 2",
        "header_bg": "E67E22",  # оранжевый - темнее исходных JSON, но светлее других развёрнутых
        "header_fg": "FFFFFF",  # белый текст для контраста
        "column_bg": None,
        "column_fg": None,
        "style_scope": "header",
        "sheets": ["CONTEST-DATA", "REWARD"],
        "columns": [
            # ADD_DATA => getCondition развёрнутые поля
            "ADD_DATA => getCondition => nonRewards", "ADD_DATA => getCondition => rewards",
            # ADD_DATA => getCondition => employeeRating развёрнутые поля
            "ADD_DATA => getCondition => employeeRating => minRatingTB", "ADD_DATA => getCondition => employeeRating => minRatingGOSB",
            "ADD_DATA => getCondition => employeeRating => minRatingBANK", "ADD_DATA => getCondition => employeeRating => seasonCode",
            "ADD_DATA => getCondition => employeeRating => minCrystalEarnedTotal"
        ],
        # Назначение: поля второго уровня развёртывания JSON (вложенные объекты)
    },

    # --- ПОЛЯ ДОБАВЛЯЕМЫЕ ЧЕРЕЗ MERGE (кроме SUMMARY) — пастельный розовый ---
    {
        "group": "Process added fields",
        "header_bg": "FFE6F2",  # пастельный розовый - приятный для глаз
        "header_fg": "2C3E50",  # тёмно-серый для читаемости
        "column_bg": None,
        "column_fg": None,
        "style_scope": "header",
        "sheets": ["REWARD", "REPORT", "TOURNAMENT-SCHEDULE"],  # поля добавляемые через merge_fields
        "columns": [COL_REWARD_LINK_CONTEST_CODE,
                    "CONTEST-DATA=>CONTEST_TYPE", "CONTEST-DATA=>FULL_NAME", "CONTEST-DATA=>BUSINESS_BLOCK", "CONTEST-DATA=>BUSINESS_STATUS",
                    "TOURNAMENT-SCHEDULE=>END_DT", "TOURNAMENT-SCHEDULE=>RESULT_DT", "TOURNAMENT-SCHEDULE=>TOURNAMENT_STATUS",
                    "REPORT=>CONTEST_DATE", "REPORT=>COUNT_CONTEST_DATE"],
        # Назначение: поля добавляемые через merge_fields_across_sheets
    },

    # --- SUMMARY КЛЮЧЕВЫЕ ПОЛЯ — как исходные данные ---
    {
        "group": "SUMMARY KEYS",
        "header_bg": "E6F3FF",  # пастельный голубой - как исходные данные
        "header_fg": "2C3E50",  # тёмно-серый для читаемости
        "column_bg": None,
        "column_fg": None,
        "style_scope": "header",
        "sheets": ["SUMMARY"],
        "columns": SUMMARY_KEY_COLUMNS,
        # Назначение: ключевые поля в SUMMARY (как исходные данные)
    },

    # --- SUMMARY ПОЛЯ: CONTEST-DATA — пастельный голубой ---
    {
        "group": "SUMMARY FIELDS: CONTEST-DATA",
        "header_bg": "CCE5FF",  # пастельный голубой - оттенок для CONTEST-DATA
        "header_fg": "2C3E50",  # тёмно-серый для читаемости
        "column_bg": None,
        "column_fg": None,
        "style_scope": "header",
        "sheets": ["SUMMARY"],
        "columns": [
            "CONTEST-DATA=>FULL_NAME",
            "CONTEST-DATA=>CONTEST_FEATURE => momentRewarding",
            "CONTEST-DATA=>FACTOR_MATCH",
            "CONTEST-DATA=>PLAN_MOD_VALUE",
            "CONTEST-DATA=>BUSINESS_BLOCK",
            "CONTEST-DATA=>CONTEST_FEATURE => tournamentStartMailing",
            "CONTEST-DATA=>CONTEST_FEATURE => tournamentEndMailing",
            "CONTEST-DATA=>CONTEST_FEATURE => tournamentRewardingMailing",
            "CONTEST-DATA=>CONTEST_FEATURE => tournamentLikeMailing",
        ],
        # Назначение: поля из CONTEST-DATA в SUMMARY
    },

    # --- SUMMARY ПОЛЯ: GROUP — пастельный зелёный ---
    {
        "group": "SUMMARY FIELDS: GROUP",
        "header_bg": "E8F5E8",  # пастельный зелёный - оттенок для GROUP
        "header_fg": "2C3E50",  # тёмно-серый для читаемости
        "column_bg": None,
        "column_fg": None,
        "style_scope": "header",
        "sheets": ["SUMMARY"],
        "columns": ["GROUP=>GET_CALC_CRITERION", "GROUP=>ADD_CALC_CRITERION", "GROUP=>ADD_CALC_CRITERION_2"],
        # Назначение: поля из GROUP в SUMMARY
    },

    # --- SUMMARY ПОЛЯ: INDICATOR — пастельный жёлтый ---
    {
        "group": "SUMMARY FIELDS: INDICATOR",
        "header_bg": "FFF8E1",  # пастельный жёлтый - оттенок для INDICATOR
        "header_fg": "2C3E50",  # тёмно-серый для читаемости
        "column_bg": None,
        "column_fg": None,
        "style_scope": "header",
        "sheets": ["SUMMARY"],
        "columns": ["INDICATOR=>INDICATOR_MARK_TYPE", "INDICATOR=>INDICATOR_MATCH", "INDICATOR=>INDICATOR_VALUE"],
        # Назначение: поля из INDICATOR в SUMMARY
    },

    # --- SUMMARY ПОЛЯ: TOURNAMENT-SCHEDULE — пастельный голубой ---
    {
        "group": "SUMMARY FIELDS: TOURNAMENT-SCHEDULE",
        "header_bg": "E1F5FE",  # пастельный голубой - оттенок для TOURNAMENT-SCHEDULE
        "header_fg": "2C3E50",  # тёмно-серый для читаемости
        "column_bg": None,
        "column_fg": None,
        "style_scope": "header",
        "sheets": ["SUMMARY"],
        "columns": ["TOURNAMENT-SCHEDULE=>START_DT", "TOURNAMENT-SCHEDULE=>END_DT", "TOURNAMENT-SCHEDULE=>RESULT_DT", "TOURNAMENT-SCHEDULE=>TOURNAMENT_STATUS", "TOURNAMENT-SCHEDULE=>TARGET_TYPE"],
        # Назначение: поля из TOURNAMENT-SCHEDULE в SUMMARY
    },

    # --- SUMMARY ПОЛЯ: REPORT — пастельный зелёный ---
    {
        "group": "SUMMARY FIELDS: REPORT",
        "header_bg": "E8F5E8",  # пастельный зелёный - оттенок для REPORT
        "header_fg": "2C3E50",  # тёмно-серый для читаемости
        "column_bg": None,
        "column_fg": None,
        "style_scope": "header",
        "sheets": ["SUMMARY"],
        "columns": ["REPORT=>CONTEST_DATE", "REPORT=>COUNT_CONTEST_DATE"],
        # Назначение: поля из REPORT в SUMMARY
    },

    # --- SUMMARY ПОЛЯ: REWARD — пастельный оранжевый ---
    {
        "group": "SUMMARY FIELDS: REWARD",
        "header_bg": "FFE8CC",  # пастельный оранжевый - оттенок для REWARD
        "header_fg": "2C3E50",  # тёмно-серый для читаемости
        "column_bg": None,
        "column_fg": None,
        "style_scope": "header",
        "sheets": ["SUMMARY"],
        "columns": [
            "REWARD=>ADD_DATA => rewardAgainGlobal",
            "REWARD=>ADD_DATA => rewardAgainTournament",
            "REWARD=>ADD_DATA => outstanding",
            "REWARD=>ADD_DATA => teamNews",
            "REWARD=>ADD_DATA => singleNews",
        ],
        # Назначение: поля из REWARD в SUMMARY
    },

    # --- ДУБЛИ В SUMMARY — пастельный розовый ---
    {
        "group": "SUMMARY DUPLICATES",
        "header_bg": "FFE6F2",  # пастельный розовый - для дублей
        "header_fg": "2C3E50",  # тёмно-серый для читаемости
        "column_bg": None,
        "column_fg": None,
        "style_scope": "header",
        "sheets": ["SUMMARY"],
        "columns": ["ДУБЛЬ: CONTEST_CODE_TOURNAMENT_CODE_REWARD_CODE_GROUP_CODE"],
        # Назначение: поля дублей в SUMMARY
    }
]

# Добавление секции для дублей по CHECK_DUPLICATES
CHECK_DUPLICATES = [
    {"sheet": "CONTEST-DATA", "key": ["CONTEST_CODE"]},
    {"sheet": "GROUP",        "key": ["CONTEST_CODE", "GROUP_CODE", "GROUP_VALUE"]},
    {"sheet": "INDICATOR",    "key": ["CONTEST_CODE", "INDICATOR_ADD_CALC_TYPE"]},
    {"sheet": "REPORT",       "key": ["MANAGER_PERSON_NUMBER", "TOURNAMENT_CODE", "CONTEST_CODE"]},
    {"sheet": "REWARD",       "key": ["REWARD_CODE"]},
    {"sheet": "REWARD-LINK",  "key": ["CONTEST_CODE", "REWARD_CODE"]},
    {"sheet": "TOURNAMENT-SCHEDULE", "key": ["TOURNAMENT_CODE", "CONTEST_CODE"]},
    {"sheet": "ORG_UNIT_V20", "key": ["ORG_UNIT_CODE"]},
    {"sheet": "USER_ROLE", "key": ["RULE_NUM"]},
    {"sheet": "USER_ROLE SB", "key": ["RULE_NUM"]},
    {"sheet": "EMPLOYEE", "key": ["PERSON_NUMBER"]}
]

for check in CHECK_DUPLICATES:
    sheet = check["sheet"]
    keys = check["key"]
    col_name = "ДУБЛЬ: " + "_".join(keys)
    COLOR_SCHEME.append({
        "group": "DUPLICATES",
        "header_bg": "FFE6F2",  # пастельный розовый - для дублей
        "header_fg": "2C3E50",  # тёмно-серый для читаемости
        "column_bg": None,
        "column_fg": None,
        "style_scope": "header",
        "sheets": [sheet],
        "columns": [col_name],
        # Назначение: поля дублей на всех листах
    })

# Какие поля разворачивать, в каком листе, с каким префиксом (строго регламентировано)
JSON_COLUMNS = {
    "CONTEST-DATA": [
        {"column": "CONTEST_FEATURE", "prefix": PREFIX_CONTEST_FEATURE},
    ],
    "REWARD": [
        {"column": "REWARD_ADD_DATA", "prefix": PREFIX_ADD_DATA},
    ],
    # Если появятся другие листы — добавить по аналогии
}


# Выходной файл Excel
def get_output_filename():
    return f'SPOD_ALL_IN_ONE_{datetime.now().strftime("%Y-%m-%d_%H-%M-%S")}.xlsx'

# Лог-файл с учетом уровня
def get_log_filename():
    # Имя лог-файла по дате, например: LOGS_2025-07-23.log
    suffix = f"_{datetime.now().strftime('%Y-%m-%d')}.log"
    return os.path.join(DIR_LOGS, LOG_BASE_NAME + suffix)


# === Логирование ===
def setup_logger():
    log_file = get_log_filename()
    # Если логгер уже инициализирован, не добавляем обработчики ещё раз
    if logging.getLogger().hasHandlers():
        return log_file
    logging.basicConfig(
        level=logging.DEBUG if LOG_LEVEL == "DEBUG" else logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        handlers=[
            logging.FileHandler(log_file, encoding="utf-8", mode="a"),  # append mode
            logging.StreamHandler(sys.stdout)
        ]
    )
    return log_file


# === Чтение CSV ===
def read_csv_file(file_path):
    func_start = time()
    params = f"({file_path})"
    logging.info(LOG_MESSAGES["func_start"].format(func="read_csv_file", params=params))
    try:
        df = pd.read_csv(file_path, sep=";", header=0, dtype=str, quoting=3, encoding="utf-8", keep_default_na=False)
        # Добавь лог первых строк для всех JSON-полей
        for col in df.columns:
            if "FEATURE" in col or "ADD_DATA" in col:
                logging.debug(LOG_MESSAGES["csv_sample"].format(
                    file=file_path,
                    column=col,
                    sample=df[col].dropna().head(2).to_list()
                ))
        logging.info(LOG_MESSAGES["read_ok"].format(file_path=file_path, rows=len(df), cols=len(df.columns)))
        func_time = time() - func_start
        logging.info(LOG_MESSAGES["func_end"].format(func="read_csv_file", params=params, time=func_time))
        return df
    except Exception as e:
        func_time = time() - func_start
        logging.error(LOG_MESSAGES["read_fail"].format(file_path=file_path, error=e))
        logging.error(LOG_MESSAGES["func_error"].format(func="read_csv_file", params=params, error=e))
        logging.info(LOG_MESSAGES["func_end"].format(func="read_csv_file", params=params, time=func_time))
        return None

# === Запись в Excel с форматированием ===
def write_to_excel(sheets_data, output_path):
    func_start = time()
    params = f"({output_path})"
    logging.info(LOG_MESSAGES["func_start"].format(func="write_to_excel", params=params))
    try:
        # SUMMARY первый, остальные — по алфавиту
        ordered_sheets = ["SUMMARY"] + [s for s in sheets_data if s != "SUMMARY"]
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            for sheet_name in ordered_sheets:
                df, params_sheet = sheets_data[sheet_name]
                df.to_excel(writer, index=False, sheet_name=sheet_name)
                ws = writer.sheets[sheet_name]
                _format_sheet(ws, df, params_sheet)
                logging.info(LOG_MESSAGES["sheet_written"].format(sheet=sheet_name, rows=len(df), cols=len(df.columns)))
            # Делать SUMMARY активным
            writer.book.active = writer.book.sheetnames.index("SUMMARY")
            writer.book.save(output_path)
        func_time = time() - func_start
        logging.info(LOG_MESSAGES["func_end"].format(func="write_to_excel", params=params, time=func_time))
    except Exception as ex:
        func_time = time() - func_start
        logging.error(LOG_MESSAGES["func_error"].format(func="write_to_excel", params=params, error=ex))
        logging.info(LOG_MESSAGES["func_end"].format(func="write_to_excel", params=params, time=func_time))

# === Форматирование листа ===
def _format_sheet(ws, df, params):
    func_start = time()
    params_str = f"({ws.title})"
    logging.debug(LOG_MESSAGES["func_start"].format(func="_format_sheet", params=params_str))
    header_font = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_data = Alignment(horizontal="left", vertical="center", wrap_text=True)
    max_col_width = params.get("max_col_width", 30)

    for col_num, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.alignment = align_center
        col_letter = get_column_letter(col_num)
        max_width = min(
            max([len(str(cell.value)) for cell in ws[get_column_letter(col_num)] if cell.value] + [8]),
            max_col_width
        )
        ws.column_dimensions[col_letter].width = max_width
        apply_color_scheme(ws, ws.title)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = align_data

    # Закрепление строк и столбцов
    ws.freeze_panes = params.get("freeze", "A2")
    ws.auto_filter.ref = ws.dimensions

    func_time = time() - func_start
    logging.debug(LOG_MESSAGES["func_end"].format(func="_format_sheet", params=params_str, time=func_time))


    # Данные: перенос строк, выравнивание по левому краю, по вертикали по центру
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = align_data

def safe_json_loads(s: str):
    """
    Преобразует строку в объект JSON. Возвращает dict/list или None, если не удается разобрать.
    Более толерантен к разным типам кавычек и пустым строкам.
    Дополнительно исправляет тройные кавычки, отсутствие двоеточий, лишние запятые и пробует "починить" кривой JSON.
    """
    if not isinstance(s, str):
        return s
    s = s.strip()
    if not s or s in {'-', 'None', 'null'}:
        return None
    try:
        return json.loads(s)
    except Exception as ex:
        try:
            fixed = s
            # 1. Заменяем тройные кавычки на обычные двойные
            fixed = fixed.replace('"""', '"')
            # 2. Заменяем одиночные и фигурные кавычки на стандартные двойные
            fixed = fixed.replace("'", '"')
            fixed = fixed.replace('“', '"').replace('”', '"')
            fixed = fixed.replace('‘', '"').replace('’', '"')
            # 3. Исправляем ключи вида ""key"" на "key"
            import re
            fixed = re.sub(r'"{2,}([^"\s]+)"{2,}', r'"\1"', fixed)
            # 4. Исправляем конструкции типа "key""": на "key":
            fixed = re.sub(r'"{2,}([^"\s]+)"{2,}\s*:', r'"\1":', fixed)
            # 5. Исправляем конструкции типа :"""value""" на :"value"
            fixed = re.sub(r':\s*"{2,}([^"\s]+)"{2,}', r':"\1"', fixed)
            # 6. Убираем завершающие запятые перед закрывающей скобкой
            fixed = re.sub(r',\s*([}\]])', r'\1', fixed)
            # 7. Исправляем отсутствие двоеточий между ключом и значением ("key" "value" -> "key":"value")
            fixed = re.sub(r'(\"[^"]+\")\s+(\")', r'\1: \2', fixed)
            # 8. Удаляем лишние пробелы между ключом и двоеточием
            fixed = re.sub(r'(\"[^"]+\")\s*:\s*', r'\1:', fixed)
            # 9. Попытка повторного парсинга
            return json.loads(fixed)
        except Exception:
            try:
                import ast
                return ast.literal_eval(fixed)
            except Exception:
                logging.debug(
                    f"[safe_json_loads] Ошибка: {ex} | Исходная строка: {repr(s)}"
                )
                return None

def flatten_json_column_recursive(df, column, prefix=None, sheet=None, sep="; "):
    import time as tmod
    func_start = tmod.time()
    n_rows = len(df)
    n_errors = 0
    prefix = prefix if prefix is not None else column
    logging.info(LOG_MESSAGES["func_start"].format(func="flatten_json_column_recursive", params=f"(лист: {sheet}, колонка: {column})"))

    def extract(obj, current_prefix):
        """Recursively flattens obj. Keeps the field itself and expands nested JSON
        if the value looks like a JSON string."""
        fields = {}
        if isinstance(obj, str):
            # try to parse nested json inside string
            nested = safe_json_loads(obj)
            if isinstance(nested, (dict, list)):
                # keep original string
                fields[current_prefix] = obj
                fields.update(extract(nested, current_prefix))
                return fields
            else:
                fields[current_prefix] = obj
                return fields

        if isinstance(obj, dict):
            fields[current_prefix] = json.dumps(obj, ensure_ascii=False)
            for k, v in obj.items():
                new_prefix = f"{current_prefix} => {k}"
                fields.update(extract(v, new_prefix))
        elif isinstance(obj, list):
            if all(isinstance(x, (str, int, float, bool, type(None))) for x in obj):
                fields[current_prefix] = sep.join(str(x) for x in obj)
            else:
                fields[current_prefix] = json.dumps(obj, ensure_ascii=False)
                for idx, x in enumerate(obj):
                    item_prefix = f"{current_prefix} => [{idx}]"
                    fields.update(extract(x, item_prefix))
        else:
            if isinstance(obj, float) and pd.isna(obj):
                fields[current_prefix] = None
            else:
                fields[current_prefix] = obj
        return fields

    new_cols = {}
    for idx, val in enumerate(df[column]):
        try:
            parsed = None
            # Строка — парсим JSON; dict/list — оставляем; иначе пропускаем
            if isinstance(val, str):
                val = val.strip()
                if val in {"", "-", "None", "null"}:
                    parsed = {}
                else:
                    parsed = safe_json_loads(val)
            elif isinstance(val, (dict, list)):
                parsed = val
            else:
                # Необрабатываемые типы (например float('nan'))
                parsed = {}
            flat = extract(parsed, prefix)
        except Exception as ex:
            logging.debug(LOG_MESSAGES["json_flatten_error"].format(row=idx, error=ex))
            n_errors += 1
            flat = {}
        for k, v in flat.items():
            if k not in new_cols:
                new_cols[k] = [None] * n_rows
            new_cols[k][idx] = v

    # Оставлять только реально созданные колонки (не пустые)
    for col_name, values in new_cols.items():
        if any(x is not None for x in values):
            df[col_name] = values

    logging.info(LOG_MESSAGES["json_flatten_summary"].format(column=column, count=len(new_cols)))
    logging.info(LOG_MESSAGES["json_flatten_keys"].format(keys=list(new_cols.keys())))
    return df



def apply_color_scheme(ws, sheet_name):
    """
    Окрашивает заголовки и/или всю колонку на листе Excel по схеме COLOR_SCHEME.
    Все действия логируются через LOG_MESSAGES.
    """
    for color_conf in COLOR_SCHEME:
        if sheet_name not in color_conf["sheets"]:
            continue

        # Список колонок: если пуст — значит все
        header_cells = list(ws[1])
        colnames = color_conf["columns"] if color_conf["columns"] else [cell.value for cell in header_cells]
        style_scope = color_conf.get("style_scope", "header")

        for colname in colnames:
            try:
                # Номер колонки по имени
                col_idx = [cell.value for cell in header_cells].index(colname) + 1
            except ValueError:
                continue  # нет такой колонки на этом листе

            # Окраска только заголовка
            if style_scope == "header":
                cell = ws.cell(row=1, column=col_idx)
                if color_conf.get("header_bg"):
                    cell.fill = PatternFill(start_color=color_conf["header_bg"], end_color=color_conf["header_bg"], fill_type="solid")
                if color_conf.get("header_fg"):
                    cell.font = Font(color=color_conf["header_fg"])
                # Логирование
                logging.debug(LOG_MESSAGES["color_scheme_applied"].format(
                    sheet=sheet_name,
                    col=colname,
                    scope="header",
                    color=color_conf.get("header_bg", "default")
                ))
            # Окраска всей колонки (если понадобится в будущем)
            elif style_scope == "all":
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.row == 1 and color_conf.get("header_bg"):
                            cell.fill = PatternFill(start_color=color_conf["header_bg"], end_color=color_conf["header_bg"], fill_type="solid")
                            if color_conf.get("header_fg"):
                                cell.font = Font(color=color_conf["header_fg"])
                        elif color_conf.get("column_bg"):
                            cell.fill = PatternFill(start_color=color_conf["column_bg"], end_color=color_conf["column_bg"], fill_type="solid")
                            if color_conf.get("column_fg"):
                                cell.font = Font(color=color_conf["column_fg"])
                logging.debug(LOG_MESSAGES["color_scheme_applied"].format(
                    sheet=sheet_name,
                    col=colname,
                    scope="all",
                    color=color_conf.get("column_bg", "default")
                ))



def collect_summary_keys(dfs):
    """
    Собирает все реально существующие сочетания ключей,
    включая осиротевшие коды и сочетания с GROUP_VALUE.
    """
    all_rows = []

    rewards = dfs.get("REWARD-LINK", pd.DataFrame())
    tournaments = dfs.get("TOURNAMENT-SCHEDULE", pd.DataFrame())
    groups = dfs.get("GROUP", pd.DataFrame())

    all_contest_codes = set()
    all_tournament_codes = set()
    all_reward_codes = set()
    all_group_codes = set()
    all_group_values = set()

    if not rewards.empty:
        all_contest_codes.update(rewards["CONTEST_CODE"].dropna())
        all_reward_codes.update(rewards["REWARD_CODE"].dropna())
    if not tournaments.empty:
        all_contest_codes.update(tournaments["CONTEST_CODE"].dropna())
        all_tournament_codes.update(tournaments["TOURNAMENT_CODE"].dropna())
    if not groups.empty:
        all_contest_codes.update(groups["CONTEST_CODE"].dropna())
        all_group_codes.update(groups["GROUP_CODE"].dropna())
        all_group_values.update(groups["GROUP_VALUE"].dropna())

    # 1. Для каждого CONTEST_CODE
    for code in all_contest_codes:
        tourns = tournaments[tournaments["CONTEST_CODE"] == code]["TOURNAMENT_CODE"].dropna().unique() if not tournaments.empty else []
        rewards_ = rewards[rewards["CONTEST_CODE"] == code]["REWARD_CODE"].dropna().unique() if not rewards.empty else []
        groups_df = groups[groups["CONTEST_CODE"] == code] if not groups.empty else pd.DataFrame()
        groups_ = groups_df["GROUP_CODE"].dropna().unique() if not groups_df.empty else []
        group_values_ = groups_df["GROUP_VALUE"].dropna().unique() if not groups_df.empty else []

        tourns = tourns if len(tourns) else ["-"]
        rewards_ = rewards_ if len(rewards_) else ["-"]
        groups_ = groups_ if len(groups_) else ["-"]
        group_values_ = group_values_ if len(group_values_) else ["-"]

        for t in tourns:
            for r in rewards_:
                for g in groups_:
                    for gv in group_values_:
                        all_rows.append((str(code), str(t), str(r), str(g), str(gv)))

    # 2. Для каждого TOURNAMENT_CODE (даже если нет CONTEST_CODE)
    if not tournaments.empty:
        for t_code in tournaments["TOURNAMENT_CODE"].dropna().unique():
            code = tournaments[tournaments["TOURNAMENT_CODE"] == t_code]["CONTEST_CODE"].dropna().unique()
            code = code[0] if len(code) else "-"
            rewards_ = rewards[rewards["CONTEST_CODE"] == code]["REWARD_CODE"].dropna().unique() if not rewards.empty else ["-"]
            groups_df = groups[groups["CONTEST_CODE"] == code] if not groups.empty else pd.DataFrame()
            groups_ = groups_df["GROUP_CODE"].dropna().unique() if not groups_df.empty else ["-"]
            group_values_ = groups_df["GROUP_VALUE"].dropna().unique() if not groups_df.empty else ["-"]
            rewards_ = rewards_ if len(rewards_) else ["-"]
            groups_ = groups_ if len(groups_) else ["-"]
            group_values_ = group_values_ if len(group_values_) else ["-"]
            for r in rewards_:
                for g in groups_:
                    for gv in group_values_:
                        all_rows.append((str(code), str(t_code), str(r), str(g), str(gv)))

    # 3. Для каждого REWARD_CODE (даже если нет CONTEST_CODE)
    if not rewards.empty:
        for r_code in rewards["REWARD_CODE"].dropna().unique():
            code = rewards[rewards["REWARD_CODE"] == r_code]["CONTEST_CODE"].dropna().unique()
            code = code[0] if len(code) else "-"
            tourns = tournaments[tournaments["CONTEST_CODE"] == code]["TOURNAMENT_CODE"].dropna().unique() if not tournaments.empty else ["-"]
            groups_df = groups[groups["CONTEST_CODE"] == code] if not groups.empty else pd.DataFrame()
            groups_ = groups_df["GROUP_CODE"].dropna().unique() if not groups_df.empty else ["-"]
            group_values_ = groups_df["GROUP_VALUE"].dropna().unique() if not groups_df.empty else ["-"]
            tourns = tourns if len(tourns) else ["-"]
            groups_ = groups_ if len(groups_) else ["-"]
            group_values_ = group_values_ if len(group_values_) else ["-"]
            for t in tourns:
                for g in groups_:
                    for gv in group_values_:
                        all_rows.append((str(code), str(t), str(r_code), str(g), str(gv)))

    # 4. Для каждого GROUP_CODE (даже если нет CONTEST_CODE)
    if not groups.empty:
        for g_code in groups["GROUP_CODE"].dropna().unique():
            code = groups[groups["GROUP_CODE"] == g_code]["CONTEST_CODE"].dropna().unique()
            code = code[0] if len(code) else "-"
            tourns = tournaments[tournaments["CONTEST_CODE"] == code]["TOURNAMENT_CODE"].dropna().unique() if not tournaments.empty else ["-"]
            rewards_ = rewards[rewards["CONTEST_CODE"] == code]["REWARD_CODE"].dropna().unique() if not rewards.empty else ["-"]
            group_values_ = groups[groups["GROUP_CODE"] == g_code]["GROUP_VALUE"].dropna().unique() if not groups.empty else ["-"]
            tourns = tourns if len(tourns) else ["-"]
            rewards_ = rewards_ if len(rewards_) else ["-"]
            group_values_ = group_values_ if len(group_values_) else ["-"]
            for t in tourns:
                for r in rewards_:
                    for gv in group_values_:
                        all_rows.append((str(code), str(t), str(r), str(g_code), str(gv)))

    # Удалить дубли
    summary_keys = pd.DataFrame(all_rows, columns=SUMMARY_KEY_COLUMNS).drop_duplicates().reset_index(drop=True)
    return summary_keys

def mark_duplicates(df, key_cols, sheet_name=None):
    """
    Добавляет колонку с пометкой о дублях по key_cols.
    Если строк по ключу больше одной — пишем xN, иначе пусто.
    """
    import time as tmod
    func_start = tmod.time()
    key_str = "_".join(key_cols)
    col_name = f"ДУБЛЬ: {key_str}"
    params = {"sheet": sheet_name, "keys": key_cols}

    logging.info(LOG_MESSAGES["duplicates_start"].format(sheet=sheet_name, keys=key_cols))
    try:
        dup_counts = df.groupby(key_cols)[key_cols[0]].transform('count')
        df[col_name] = dup_counts.apply(lambda x: f"x{x}" if x > 1 else "")
        n_duplicates = (df[col_name] != "").sum()
        func_time = tmod.time() - func_start
        logging.info(LOG_MESSAGES["duplicates_found"].format(count=n_duplicates, sheet=sheet_name, keys=key_cols))
        logging.info(LOG_MESSAGES["duplicates_end"].format(sheet=sheet_name, time=func_time))
    except Exception as ex:
        func_time = tmod.time() - func_start
        logging.error(LOG_MESSAGES["duplicates_error"].format(sheet=sheet_name, keys=key_cols, error=ex))
        logging.info(LOG_MESSAGES["duplicates_end"].format(sheet=sheet_name, time=func_time))
    return df

def add_fields_to_sheet(df_base, df_ref, src_keys, dst_keys, columns, sheet_name, ref_sheet_name, mode="value"):
    """
    Добавляет к df_base поля из df_ref по ключам.
    Если mode == "value": подтягивает значения первого найденного (основной режим).
    Если mode == "count": добавляет количество строк в df_ref по каждому ключу.
    Если нужной колонки нет — создаёт её с дефолтными значениями "-".
    """
    func_start = time()
    logging.info(LOG_MESSAGES["func_start"].format(
        func="add_fields_to_sheet",
        params=f"(лист: {sheet_name}, поля: {columns}, ключ: {dst_keys}->{src_keys}, mode: {mode})"
    ))
    if isinstance(columns, str):
        columns = [columns]

    def tuple_key(row, keys):
        # Гарантируем, что всегда возвращается кортеж скаляров, даже если ключ один
        if isinstance(keys, (list, tuple)):
            result = []
            for k in keys:
                v = row[k]
                # Если v — Series (например, из-за дублирующихся колонок), берём только первый элемент
                if isinstance(v, pd.Series):
                    v = v.iloc[0]
                result.append(v)
            return tuple(result)
        else:
            v = row[keys]
            if isinstance(v, pd.Series):
                v = v.iloc[0]
            return (v,)

    new_keys = df_base.apply(lambda row: tuple_key(row, dst_keys), axis=1)

    # --- Добавлено: авто-дополнение отсутствующих колонок и ключей ---
    missing_cols = [col for col in columns if col not in df_ref.columns]
    for col in missing_cols:
        logging.warning(LOG_MESSAGES["missing_column"].format(column=col, sheet=ref_sheet_name))
        df_ref[col] = "-"

    missing_keys = [k for k in src_keys if k not in df_ref.columns]
    for k in missing_keys:
        logging.warning(LOG_MESSAGES["missing_key"].format(key=k, sheet=ref_sheet_name))
        df_ref[k] = "-"

    if mode == "count":
        group_counts = df_ref.groupby(src_keys).size().to_dict()
        for col in columns:
            count_col_name = f"{ref_sheet_name}=>COUNT_{col}"
            df_base[count_col_name] = new_keys.map(group_counts).fillna(0).astype(int)
        logging.info(LOG_MESSAGES["func_end"].format(
            func="add_fields_to_sheet",
            params=f"(лист: {sheet_name}, mode: count, ключ: {dst_keys}->{src_keys})",
            time=time() - func_start
        ))
        return df_base

    for col in columns:
        ref_map = dict(zip(
            df_ref.apply(lambda row: tuple_key(row, src_keys), axis=1),
            df_ref[col]
        ))
        new_col_name = f"{ref_sheet_name}=>{col}"
        df_base[new_col_name] = new_keys.map(ref_map).fillna("-")
        # Специально для REWARD_LINK =>CONTEST_CODE: auto-rename, если создали с дефисом
        if new_col_name.replace("-", "_").replace(" ", "") == COL_REWARD_LINK_CONTEST_CODE.replace("-", "_").replace(" ", ""):
            candidates = [c for c in df_base.columns if c.replace("-", "_").replace(" ", "") == COL_REWARD_LINK_CONTEST_CODE.replace("-", "_").replace(" ", "")]
            for cand in candidates:
                if cand != COL_REWARD_LINK_CONTEST_CODE:
                    df_base = df_base.rename(columns={cand: COL_REWARD_LINK_CONTEST_CODE})

    logging.info(LOG_MESSAGES["func_end"].format(
        func="add_fields_to_sheet",
        params=f"(лист: {sheet_name}, поля: {columns}, ключ: {dst_keys}->{src_keys}, mode: {mode})",
        time=time() - func_start
    ))
    return df_base


def merge_fields_across_sheets(sheets_data, merge_fields):
    """
    Универсально добавляет поля по правилам из merge_fields
    (source_df -> target_df), поддержка mode value / count.
    sheets_data: dict {sheet_name: (df, params)}
    merge_fields: список блоков с параметрами (см. выше)
    """
    for rule in merge_fields:
        sheet_src = rule["sheet_src"]
        sheet_dst = rule["sheet_dst"]
        src_keys = rule["src_key"] if isinstance(rule["src_key"], list) else [rule["src_key"]]
        dst_keys = rule["dst_key"] if isinstance(rule["dst_key"], list) else [rule["dst_key"]]
        col_names = rule["column"]
        mode = rule.get("mode", "value")
        params_str = f"(src: {sheet_src} -> dst: {sheet_dst}, поля: {col_names}, ключ: {dst_keys}<-{src_keys}, mode: {mode})"

        if sheet_src not in sheets_data or sheet_dst not in sheets_data:
            logging.warning(LOG_MESSAGES.get("field_missing", LOG_MESSAGES["func_error"]).format(
                column=col_names, src_sheet=sheet_src, src_key=src_keys
            ))
            continue

        df_src = sheets_data[sheet_src][0]
        df_dst, params_dst = sheets_data[sheet_dst]

        logging.info(LOG_MESSAGES["func_start"].format(func="merge_fields_across_sheets", params=params_str))
        df_dst = add_fields_to_sheet(df_dst, df_src, src_keys, dst_keys, col_names, sheet_dst, sheet_src, mode=mode)
        sheets_data[sheet_dst] = (df_dst, params_dst)
        logging.info(LOG_MESSAGES["func_end"].format(func="merge_fields_across_sheets", params=params_str, time=0))
    return sheets_data

def build_summary_sheet(dfs, params_summary, merge_fields):
    func_start = time()
    params_log = f"(лист: {params_summary['sheet']})"
    logging.info(LOG_MESSAGES["func_start"].format(func="build_summary_sheet", params=params_log))

    summary = collect_summary_keys(dfs)

    logging.info(LOG_MESSAGES["summary"].format(summary=f"Каркас: {len(summary)} строк (реальные комбинации ключей)"))
    logging.debug(LOG_MESSAGES["debug_head"].format(sheet=params_summary["sheet"], head=summary.head(5).to_string()))

    # Универсально добавляем все поля по merge_fields
    for field in merge_fields:
        col_names = field["column"]
        if isinstance(col_names, str):
            col_names = [col_names]
        sheet_src = field["sheet_src"]
        src_keys = field["src_key"] if isinstance(field["src_key"], list) else [field["src_key"]]
        dst_keys = field["dst_key"] if isinstance(field["dst_key"], list) else [field["dst_key"]]
        mode = field.get("mode", "value")
        params_str = f"(лист-источник: {sheet_src}, поля: {col_names}, ключ: {dst_keys}->{src_keys}, mode: {mode})"
        logging.info(LOG_MESSAGES["func_start"].format(func="add_fields_to_sheet", params=params_str))
        ref_df = dfs.get(sheet_src)
        if ref_df is None:
            logging.warning(LOG_MESSAGES.get("field_missing", LOG_MESSAGES["func_error"]).format(
                column=col_names, src_sheet=sheet_src, src_key=src_keys
            ))
            continue

        summary = add_fields_to_sheet(summary, ref_df, src_keys, dst_keys, col_names, params_summary["sheet"], sheet_src, mode=mode)
        logging.info(LOG_MESSAGES["func_end"].format(func="add_fields_to_sheet", params=params_str, time=0))

    n_rows, n_cols = summary.shape
    func_time = time() - func_start
    logging.info(LOG_MESSAGES["fields_summary"].format(rows=n_rows, cols=n_cols))
    logging.info(LOG_MESSAGES["sheet_written"].format(sheet=params_summary['sheet'], rows=n_rows, cols=n_cols))
    logging.info(LOG_MESSAGES["func_end"].format(func="build_summary_sheet", params=params_log, time=func_time))
    logging.debug(LOG_MESSAGES["debug_columns"].format(sheet=params_summary["sheet"], columns=', '.join(summary.columns.tolist())))
    logging.debug(LOG_MESSAGES["debug_head"].format(sheet=params_summary["sheet"], head=summary.head(5).to_string()))
    return summary

# Функция enrich_reward_with_contest_code удалена - CONTEST_CODE теперь добавляется через merge_fields_across_sheets

def main():
    start_time = datetime.now()
    log_file = setup_logger()
    logging.info(LOG_MESSAGES["start"].format(time=start_time.strftime("%Y-%m-%d %H:%M:%S")))

    sheets_data = {}
    files_processed = 0
    rows_total = 0
    summary = []

    # 1. Чтение всех CSV и разворот ВСЕХ JSON‑полей на каждом листе
    for file_conf in INPUT_FILES:
        file_path = os.path.join(DIR_INPUT, file_conf["file"] + ".CSV")
        sheet_name = file_conf["sheet"]
        logging.info(LOG_MESSAGES["reading_file"].format(file_path=file_path))
        df = read_csv_file(file_path)
        if df is not None:
            # --- Разворачиваем только нужные JSON-поля по строгому списку ---
            json_columns = JSON_COLUMNS.get(sheet_name, [])
            for json_conf in json_columns:
                col = json_conf["column"]
                prefix = json_conf.get("prefix", col)
                if col in df.columns:
                    df = flatten_json_column_recursive(df, col, prefix=prefix, sheet=sheet_name)
                    logging.info(LOG_MESSAGES["json_flatten_done"].format(sheet=sheet_name, column=col, prefix=prefix))
                else:
                    logging.warning(LOG_MESSAGES["json_flatten_missing"].format(sheet=sheet_name, column=col))

            # Для дебага: логируем итоговый список колонок после всех разворотов
            logging.debug(LOG_MESSAGES["debug_columns"].format(sheet=sheet_name, columns=', '.join(df.columns.tolist())))
            sheets_data[sheet_name] = (df, file_conf)
            files_processed += 1
            rows_total += len(df)
            summary.append(f"{sheet_name}: {len(df)} строк")
        else:
            summary.append(f"{sheet_name}: ошибка")
    # CONTEST_CODE добавляется через merge_fields_across_sheets, поэтому enrich_reward_with_contest_code больше не нужен
    # if "REWARD" in sheets_data and "REWARD-LINK" in sheets_data:
    #     df_reward, conf_reward = sheets_data["REWARD"]
    #     df_link, conf_link = sheets_data["REWARD-LINK"]
    #     # Всегда пересоздаём колонку с нужным именем (автоочистка битых вариантов)
    #     df_reward = enrich_reward_with_contest_code(df_reward, df_link)
    #     sheets_data["REWARD"] = (df_reward, conf_reward)

    # 3. Merge fields (только после полного разворота JSON)
    merge_fields_across_sheets(
        sheets_data,
        [f for f in MERGE_FIELDS if f.get("sheet_dst") != "SUMMARY"]
    )

    # 4. Проверка на дубли
    for sheet_name, (df, conf) in sheets_data.items():
        check_cfg = next((x for x in CHECK_DUPLICATES if x["sheet"] == sheet_name), None)
        if check_cfg:
            df = mark_duplicates(df, check_cfg["key"], sheet_name=sheet_name)
            sheets_data[sheet_name] = (df, conf)

    # 5. Формирование итогового Summary (build_summary_sheet)
    dfs = {k: v[0] for k, v in sheets_data.items()}
    df_summary = build_summary_sheet(
        dfs,
        params_summary=SUMMARY_SHEET,
        merge_fields=[f for f in MERGE_FIELDS if f.get("sheet_dst") == "SUMMARY"]
    )
    sheets_data[SUMMARY_SHEET["sheet"]] = (df_summary, SUMMARY_SHEET)

    # 6. Запись в Excel
    output_excel = os.path.join(DIR_OUTPUT, get_output_filename())
    logging.info(LOG_MESSAGES["func_start"].format(func="write_to_excel", params=f"({output_excel})"))
    write_to_excel(sheets_data, output_excel)
    logging.info(LOG_MESSAGES["func_end"].format(func="write_to_excel", params=f"({output_excel})", time=0))

    time_elapsed = datetime.now() - start_time
    logging.info(LOG_MESSAGES["finish"].format(
        files=files_processed,
        rows_total=rows_total,
        time_elapsed=str(time_elapsed)
    ))
    logging.info(LOG_MESSAGES["summary"].format(summary="; ".join(summary)))
    logging.info(LOG_MESSAGES["excel_path"].format(path=output_excel))
    logging.info(LOG_MESSAGES["log_path"].format(path=log_file))


if __name__ == "__main__":
    main()
