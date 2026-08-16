# -*- coding: utf-8 -*-
"""Импорт Excel-формы BADGE → Excel листов SPOD + CSV."""

from __future__ import annotations

import csv
import logging
import os
from datetime import datetime
from typing import Any, Dict, List, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

from src.contest_badge_form import schema
from src.contest_badge_form.form_io import read_form_workbook
from src.contest_badge_form.spod_json import (
    coerce_form_scalar,
    dumps_spod_json,
    list_from_form_cell,
)


def _validate_badges(payload: Dict[str, Any]) -> List[str]:
    """Проверки лимита BADGE; возвращает список предупреждений/ошибок."""
    warnings: List[str] = []
    flat = payload.get("contest_flat") or {}
    contest_type = str(flat.get("CONTEST_TYPE") or "")
    code = str(flat.get("CONTEST_CODE") or "")
    badges = list(payload.get("badges") or [])
    limit = schema.max_badge_slots(contest_type)
    t = contest_type.strip().upper()
    if t in {"ИНДИВИДУАЛЬНЫЙ", "ИНДИВИДУАЛЬНЫЙ НАКОПИТЕЛЬНЫЙ"}:
        if len(badges) != 1:
            warnings.append(
                f"{code}: для типа «{contest_type}» нужна ровно 1 BADGE, сейчас {len(badges)}"
            )
    elif len(badges) > limit:
        warnings.append(
            f"{code}: BADGE={len(badges)} больше лимита {limit} ({schema.expected_badge_count_note(contest_type)})"
        )
    for b in badges:
        rtype = str((b.get("flat") or {}).get("REWARD_TYPE") or "").upper()
        if rtype and rtype != "BADGE":
            warnings.append(
                f"{code}: награда {(b.get('flat') or {}).get('REWARD_CODE')} "
                f"имеет тип {rtype}, ожидается BADGE"
            )
    return warnings


def _assemble_contest_row(payload: Dict[str, Any]) -> Dict[str, str]:
    """Строка CONTEST-DATA со всеми колонками."""
    flat = dict(payload.get("contest_flat") or {})
    arrays = payload.get("contest_arrays") or {}
    feature = payload.get("contest_feature") or {}

    row: Dict[str, str] = {}
    for key, _ in schema.CONTEST_FLAT_FIELDS:
        row[key] = str(flat.get(key, "") or "")

    for key, _ in schema.CONTEST_ARRAY_FIELDS:
        raw = arrays.get(key, flat.get(key, ""))
        items = list_from_form_cell(raw)
        row[key] = dumps_spod_json(items)

    # CONTEST_FEATURE: только известные ключи + сохранение порядка schema
    feat_obj: Dict[str, Any] = {}
    for key, _label, kind in schema.CONTEST_FEATURE_FIELDS:
        if key not in feature:
            continue
        val = feature[key]
        if kind == "list":
            items = val if isinstance(val, list) else list_from_form_cell(val)
            # Элементы массивов FEATURE — строки (сохраняем ведущие нули)
            feat_obj[key] = ["" if x is None else str(x) for x in items]
        else:
            # accuracy / minNumber в SPOD — числа; прочие скаляры — строки
            if key in {"accuracy", "minNumber"}:
                feat_obj[key] = coerce_form_scalar(val)
            else:
                feat_obj[key] = "" if val is None else str(val)
    # доп. ключи из формы, не описанные в schema
    for key, val in feature.items():
        if key in feat_obj:
            continue
        if isinstance(val, list):
            feat_obj[key] = ["" if x is None else str(x) for x in val]
        else:
            feat_obj[key] = coerce_form_scalar(val)
    row["CONTEST_FEATURE"] = dumps_spod_json(feat_obj)
    return row


def _assemble_reward_row(badge: Dict[str, Any]) -> Dict[str, str]:
    """Строка REWARD."""
    flat = badge.get("flat") or {}
    add = badge.get("add_data") or {}
    row: Dict[str, str] = {}
    for key, _ in schema.REWARD_FLAT_FIELDS:
        row[key] = str(flat.get(key, "") or "")
    if not row.get("REWARD_TYPE"):
        row["REWARD_TYPE"] = "BADGE"

    add_obj: Dict[str, Any] = {}
    for key, _label, kind in schema.REWARD_ADD_DATA_FIELDS:
        if key not in add:
            continue
        val = add[key]
        if kind == "list":
            items = val if isinstance(val, list) else list_from_form_cell(val)
            # Элементы массивов ADD_DATA в SPOD — строки
            add_obj[key] = ["" if x is None else str(x) for x in items]
        else:
            # Скаляры ADD_DATA в выгрузке — строки (в т.ч. «1», «Y»)
            add_obj[key] = "" if val is None else str(val)
    for key, val in add.items():
        if key in add_obj:
            continue
        if isinstance(val, list):
            add_obj[key] = ["" if x is None else str(x) for x in val]
        else:
            add_obj[key] = "" if val is None else str(val)
    row["REWARD_ADD_DATA"] = dumps_spod_json(add_obj)
    return row


def assemble_spod_tables(
    payloads: List[Dict[str, Any]],
) -> Tuple[Dict[str, List[Dict[str, str]]], List[str]]:
    """
    Собрать таблицы SPOD из списка payload формы.
    Возвращает (tables, warnings).
    """
    tables: Dict[str, List[Dict[str, str]]] = {
        "contest": [],
        "reward": [],
        "reward_link": [],
        "group": [],
        "indicator": [],
        "schedule": [],
    }
    warnings: List[str] = []
    seen_rewards: set[str] = set()

    for payload in payloads:
        warnings.extend(_validate_badges(payload))
        contest_row = _assemble_contest_row(payload)
        if not contest_row.get("CONTEST_CODE"):
            warnings.append("Лист без CONTEST_CODE — пропуск")
            continue
        tables["contest"].append(contest_row)

        for badge in payload.get("badges") or []:
            rrow = _assemble_reward_row(badge)
            code = rrow.get("REWARD_CODE") or ""
            if code and code not in seen_rewards:
                tables["reward"].append(rrow)
                seen_rewards.add(code)
            elif code in seen_rewards:
                logging.info(
                    "[contest_badge_form] Дубликат REWARD_CODE при импорте пропущен: %s",
                    code,
                )

        for key, colset in (
            ("reward_link", schema.REWARD_LINK_COLUMNS),
            ("group", schema.GROUP_COLUMNS),
            ("indicator", schema.INDICATOR_COLUMNS),
            ("schedule", schema.SCHEDULE_COLUMNS),
        ):
            for item in payload.get(key) or []:
                row = {c: str(item.get(c, "") or "") for c in colset}
                # пустые строки таблиц пропускаем
                if all(not v for v in row.values()):
                    continue
                tables[key].append(row)

    return tables, warnings


def _write_excel_sheets(
    path: str, tables: Dict[str, List[Dict[str, str]]]
) -> None:
    """Книга с листами SPOD."""
    wb = Workbook()
    first = True
    order = [
        ("contest", schema.SPOD_SHEET_NAMES["contest"], [c for c, _ in schema.CONTEST_FLAT_FIELDS]
         + ["CONTEST_FEATURE"]
         + [c for c, _ in schema.CONTEST_ARRAY_FIELDS]),
        (
            "reward",
            schema.SPOD_SHEET_NAMES["reward"],
            [c for c, _ in schema.REWARD_FLAT_FIELDS] + ["REWARD_ADD_DATA"],
        ),
        ("reward_link", schema.SPOD_SHEET_NAMES["reward_link"], list(schema.REWARD_LINK_COLUMNS)),
        ("group", schema.SPOD_SHEET_NAMES["group"], list(schema.GROUP_COLUMNS)),
        ("indicator", schema.SPOD_SHEET_NAMES["indicator"], list(schema.INDICATOR_COLUMNS)),
        ("schedule", schema.SPOD_SHEET_NAMES["schedule"], list(schema.SCHEDULE_COLUMNS)),
    ]
    # Колонки contest: как в CSV — полный порядок исходника
    contest_cols = [
        "CONTEST_CODE",
        "FULL_NAME",
        "CREATE_DT",
        "CLOSE_DT",
        "BUSINESS_STATUS",
        "CONTEST_TYPE",
        "CONTEST_DESCRIPTION",
        "CONTEST_FEATURE",
        "SHOW_INDICATOR",
        "PRODUCT_GROUP",
        "PRODUCT",
        "CONTEST_SUBJECT",
        "FACTOR_MARK_TYPE",
        "CONTEST_INDICATOR_METHOD",
        "CONTEST_FACTOR_METHOD",
        "PLAN_METHOD_CODE",
        "PLAN_MOD_METOD",
        "PLAN_MOD_VALUE",
        "FACTOR_MATCH",
        "CONTEST_PERIOD",
        "TARGET_TYPE",
        "SOURCE_UPD_FREQUENCY",
        "CALC_TYPE",
        "BUSINESS_BLOCK",
        "FACT_POST_PROCESSING",
    ]
    order[0] = ("contest", schema.SPOD_SHEET_NAMES["contest"], contest_cols)

    for key, sheet_name, columns in order:
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True)
        for row_idx, data_row in enumerate(tables.get(key) or [], start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                ws.cell(
                    row=row_idx,
                    column=col_idx,
                    value=str(data_row.get(col_name, "") or ""),
                )
    wb.save(path)


def _write_csv_tables(
    out_dir: str,
    block: str,
    tables: Dict[str, List[Dict[str, str]]],
    contest_cols: Sequence[str],
) -> List[str]:
    """Записать CSV (sep=;) для каждого листа. Возвращает пути."""
    mapping = [
        ("contest", f"CONTEST ({block}) FORM_IMPORT.csv", contest_cols),
        (
            "reward",
            f"REWARD ({block}) FORM_IMPORT.csv",
            [c for c, _ in schema.REWARD_FLAT_FIELDS] + ["REWARD_ADD_DATA"],
        ),
        (
            "reward_link",
            f"REWARD-LINK ({block}) FORM_IMPORT.csv",
            list(schema.REWARD_LINK_COLUMNS),
        ),
        ("group", f"GROUP ({block}) FORM_IMPORT.csv", list(schema.GROUP_COLUMNS)),
        (
            "indicator",
            f"INDICATOR ({block}) FORM_IMPORT.csv",
            list(schema.INDICATOR_COLUMNS),
        ),
        (
            "schedule",
            f"SCHEDULE ({block}) FORM_IMPORT.csv",
            list(schema.SCHEDULE_COLUMNS),
        ),
    ]
    paths: List[str] = []
    for key, filename, columns in mapping:
        path = os.path.join(out_dir, filename)
        df = pd.DataFrame(tables.get(key) or [], columns=list(columns))
        for c in columns:
            if c not in df.columns:
                df[c] = ""
        df = df[list(columns)]
        df.to_csv(path, sep=";", index=False, encoding="utf-8-sig", quoting=csv.QUOTE_NONE, escapechar="\\")
        paths.append(path)
        logging.info(
            "[contest_badge_form] CSV %s: %s строк → %s",
            key,
            len(df),
            path,
        )
    return paths


def import_form_file(
    form_path: str,
    project_base_dir: str,
    cfg: Dict[str, Any],
    block: str,
    output_dir: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Импорт формы → каталог с xlsx + csv.
    Возвращает метаданные: output_dir, excel, csv_paths, warnings.
    """
    payloads = read_form_workbook(form_path)
    if not payloads:
        raise ValueError(f"contest_badge_form_import: пустая форма {form_path}")

    tables, warnings = assemble_spod_tables(payloads)
    for w in warnings:
        logging.warning("[contest_badge_form] %s", w)

    if not output_dir:
        paths = cfg.get("paths") or {}
        out_root = str(paths.get("output") or "OUT")
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_dir = os.path.join(
            project_base_dir,
            out_root,
            block,
            f"CONTEST_BADGE_FORM_IMPORT_{ts}",
        )
    os.makedirs(output_dir, exist_ok=True)

    contest_cols = [
        "CONTEST_CODE",
        "FULL_NAME",
        "CREATE_DT",
        "CLOSE_DT",
        "BUSINESS_STATUS",
        "CONTEST_TYPE",
        "CONTEST_DESCRIPTION",
        "CONTEST_FEATURE",
        "SHOW_INDICATOR",
        "PRODUCT_GROUP",
        "PRODUCT",
        "CONTEST_SUBJECT",
        "FACTOR_MARK_TYPE",
        "CONTEST_INDICATOR_METHOD",
        "CONTEST_FACTOR_METHOD",
        "PLAN_METHOD_CODE",
        "PLAN_MOD_METOD",
        "PLAN_MOD_VALUE",
        "FACTOR_MATCH",
        "CONTEST_PERIOD",
        "TARGET_TYPE",
        "SOURCE_UPD_FREQUENCY",
        "CALC_TYPE",
        "BUSINESS_BLOCK",
        "FACT_POST_PROCESSING",
    ]
    excel_path = os.path.join(
        output_dir, f"SPOD_{block}_CONTEST_BADGE_FORM_IMPORT.xlsx"
    )
    _write_excel_sheets(excel_path, tables)
    csv_paths = _write_csv_tables(output_dir, block, tables, contest_cols)
    logging.info(
        "[contest_badge_form] Импорт готов: %s (конкурсов=%s)",
        output_dir,
        len(tables["contest"]),
    )
    return {
        "output_dir": output_dir,
        "excel": excel_path,
        "csv_paths": csv_paths,
        "warnings": warnings,
        "tables": tables,
    }
