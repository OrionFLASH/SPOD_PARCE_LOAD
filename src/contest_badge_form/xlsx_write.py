# -*- coding: utf-8 -*-
"""Запись Excel-формы BADGE через openpyxl (+ выпадающие списки)."""

from __future__ import annotations

import logging
import os
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from src.contest_badge_form import schema
from src.contest_badge_form.field_meta import (
    INPUT_KIND_COLORS,
    INPUT_KIND_LABELS,
    INPUT_KIND_ORDER,
    TABLE_COLUMN_HINTS,
    TABLE_DROPDOWNS,
    description_for,
    input_kind_for_kv,
    input_kind_for_table_col,
    merge_dropdowns,
)
from src.contest_badge_form.form_io import save_workbook
from src.contest_badge_form.spod_json import form_cell_from_list, list_from_form_cell

_EXCEL_MAX = 32000
_LISTS_SHEET = "Lists"


def _hex(color: str) -> str:
    """ARGB для openpyxl без символа #."""
    return color.lstrip("#").upper()


def _s(value: Any) -> str:
    """Безопасная строка для ячейки."""
    if value is None:
        return ""
    text = str(value)
    text = "".join(ch for ch in text if ord(ch) >= 32 or ch in "\t\n\r")
    if len(text) > _EXCEL_MAX:
        text = text[: _EXCEL_MAX - 15] + "…[обрезано]"
    return text


def _feature_value(feature: Dict[str, Any], key: str, kind: str) -> str:
    raw = feature.get(key, [] if kind == "list" else "")
    if kind == "list":
        if isinstance(raw, list):
            return form_cell_from_list(raw)
        return form_cell_from_list(list_from_form_cell(raw))
    if raw is None:
        return ""
    if isinstance(raw, list):
        return form_cell_from_list(raw)
    return str(raw)


def _clean_list(values: Sequence[Any]) -> List[str]:
    out: List[str] = []
    for v in values:
        s = _s(v).strip()
        if s and s not in out:
            out.append(s)
    return out


def _needs_lists_sheet(values: List[str]) -> bool:
    """Длинные списки или значения с запятой — только через лист Lists."""
    if not values:
        return False
    if any("," in v for v in values):
        return True
    return len(",".join(values)) > 200


def _build_lists_ranges(dropdown_map: Dict[str, List[str]]) -> Dict[str, str]:
    """
    key → формула диапазона (Lists!$A$2:$A$10) для длинных списков.
    Сам лист создаётся позже.
    """
    to_sheet = {
        k: _clean_list(v)
        for k, v in dropdown_map.items()
        if _needs_lists_sheet(_clean_list(v))
    }
    if not to_sheet:
        return {}

    ranges: Dict[str, str] = {}
    for col, key in enumerate(sorted(to_sheet.keys())):
        values = to_sheet[key]
        letter = get_column_letter(col + 1)
        ranges[key] = f"{_LISTS_SHEET}!${letter}$2:${letter}${1 + len(values)}"
    return ranges


def _fill_lists_sheet(wb: Workbook, dropdown_map: Dict[str, List[str]]) -> None:
    """Создать и заполнить скрытый лист Lists (после листов формы)."""
    to_sheet = {
        k: _clean_list(v)
        for k, v in dropdown_map.items()
        if _needs_lists_sheet(_clean_list(v))
    }
    if not to_sheet:
        return
    ws = wb.create_sheet(_LISTS_SHEET)
    ws.sheet_state = "hidden"
    for col, key in enumerate(sorted(to_sheet.keys()), start=1):
        values = to_sheet[key]
        ws.cell(row=1, column=col, value=key[:80])
        for i, val in enumerate(values, start=2):
            ws.cell(row=i, column=col, value=val)


def _dv_formula(
    key: str, values: List[str], list_ranges: Dict[str, str]
) -> Optional[str]:
    """formula1 для DataValidation: inline-список или ссылка на Lists."""
    cleaned = _clean_list(values)
    if not cleaned:
        return None
    if key in list_ranges:
        return list_ranges[key]
    if _needs_lists_sheet(cleaned):
        cleaned = [v.replace(",", " ") for v in cleaned]
    if all("," not in v for v in cleaned) and len(",".join(cleaned)) <= 240:
        return '"' + ",".join(cleaned) + '"'
    joined = ",".join(v.replace(",", " ") for v in cleaned)
    if len(joined) > 240:
        joined = ",".join(cleaned[:15])
    return f'"{joined}"'


def _apply_list_validation(
    ws: Worksheet,
    cells_a1: List[str],
    formula: Optional[str],
) -> None:
    """Навесить list-validation на ячейки (A1-нотация)."""
    if not cells_a1 or not formula:
        return
    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=False,
        showInputMessage=False,
    )
    ws.add_data_validation(dv)
    if len(cells_a1) == 1:
        dv.add(cells_a1[0])
        return
    dv.sqref = " ".join(cells_a1)


def _value_style(kind: str) -> Dict[str, Any]:
    """Стили ячейки значения по типу ввода."""
    color = INPUT_KIND_COLORS.get(kind, INPUT_KIND_COLORS["text"])
    return {
        "fill": PatternFill("solid", fgColor=_hex(color)),
        "alignment": Alignment(wrap_text=True, vertical="center"),
    }


def _apply_cell_style(cell: Any, style: Dict[str, Any]) -> None:
    for attr, val in style.items():
        setattr(cell, attr, val)


_FILL_SECTION = PatternFill("solid", fgColor=_hex("#1F4E79"))
_FONT_SECTION = Font(bold=True, color="FFFFFF", size=12)
_FILL_KV = PatternFill("solid", fgColor=_hex("#D6EAF8"))
_FONT_KEY = Font(color=_hex("#566573"), size=10)
_FILL_HEADER = PatternFill("solid", fgColor=_hex("#D6EAF8"))
_FILL_HEADER_VAL = PatternFill("solid", fgColor=_hex("#FFF2CC"))
_FILL_HEADER_DESC = PatternFill("solid", fgColor=_hex("#F8F9F9"))
_FILL_DESC = PatternFill("solid", fgColor=_hex("#F8F9F9"))
_FONT_DESC = Font(color=_hex("#5D6D7E"), italic=True, size=9)
_FILL_TABLE = PatternFill("solid", fgColor=_hex("#D5F5E3"))
_FILL_HINT = PatternFill("solid", fgColor=_hex("#E8F8F5"))
_FONT_HINT = Font(italic=True, size=8, color=_hex("#1E8449"))
_FONT_LEGEND = Font(bold=True, size=10)


def write_form_xlsx(
    path: str,
    payloads: List[Dict[str, Any]],
    *,
    dropdowns: Optional[Dict[str, List[str]]] = None,
    with_dropdowns: bool = True,
) -> str:
    """
    Записать форму (листы 1..N) через openpyxl.
    with_dropdowns=True — выпадающие списки из field_meta (Y/N, ПРОМ/ТЕСТ и др.).
    """
    parent = os.path.dirname(os.path.abspath(path))
    if parent:
        os.makedirs(parent, exist_ok=True)

    wb = Workbook()
    default_ws = wb.active
    assert default_ws is not None

    dd_map = merge_dropdowns(dropdowns) if with_dropdowns else {}
    for table_key, col_map in TABLE_DROPDOWNS.items():
        for col_name, values in col_map.items():
            dd_map[f"TBL:{table_key}:{col_name}"] = list(values)

    list_ranges = _build_lists_ranges(dd_map) if with_dropdowns else {}

    value_styles = {kind: _value_style(kind) for kind in INPUT_KIND_COLORS}
    legend_styles = {
        kind: {
            "fill": PatternFill("solid", fgColor=_hex(INPUT_KIND_COLORS[kind])),
            "font": Font(bold=True, size=9),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": None,
        }
        for kind in INPUT_KIND_COLORS
    }

    use_payloads = payloads if payloads else [{}]
    first = True
    for idx, payload in enumerate(use_payloads, start=1):
        if first:
            ws = default_ws
            ws.title = str(idx)
            first = False
        else:
            ws = wb.create_sheet(str(idx))
        _write_sheet(
            ws,
            payload,
            value_styles=value_styles,
            legend_styles=legend_styles,
            dropdown_map=dd_map if with_dropdowns else {},
            list_ranges=list_ranges,
            with_dropdowns=with_dropdowns,
        )

    if with_dropdowns:
        _fill_lists_sheet(wb, dd_map)

    save_workbook(wb, path, keep_data_validations=with_dropdowns)
    logging.info(
        "[contest_badge_form] Форма (openpyxl%s): %s",
        ", dropdowns" if with_dropdowns else "",
        path,
    )
    return path


def _write_kv(
    ws: Worksheet,
    row: int,
    key: str,
    label: str,
    value: Any,
    *,
    in_badge_slot: bool,
    schema_kind: Optional[str],
    value_styles: Dict[str, Dict[str, Any]],
    kv_cells: Dict[str, List[str]],
    dropdown_map: Dict[str, List[str]],
) -> int:
    kind = input_kind_for_kv(
        key,
        schema_kind=schema_kind,
        has_dropdown=key in dropdown_map,
    )
    style = value_styles.get(kind) or value_styles["text"]

    key_cell = ws.cell(row=row, column=1, value=_s(key))
    key_cell.fill = _FILL_KV
    key_cell.font = _FONT_KEY

    label_cell = ws.cell(row=row, column=2, value=_s(label))
    label_cell.fill = _FILL_KV

    val_cell = ws.cell(row=row, column=3, value=_s(value))
    _apply_cell_style(val_cell, style)

    desc_cell = ws.cell(
        row=row,
        column=4,
        value=_s(description_for(key, in_badge_slot=in_badge_slot)),
    )
    desc_cell.fill = _FILL_DESC
    desc_cell.font = _FONT_DESC
    desc_cell.alignment = Alignment(wrap_text=True, vertical="center")

    ws.row_dimensions[row].height = 28
    kv_cells.setdefault(key, []).append(f"C{row}")
    return row + 1


def _write_section(ws: Worksheet, row: int, title: str) -> int:
    for col in range(1, 5):
        cell = ws.cell(row=row, column=col, value=_s(title) if col == 1 else "")
        cell.fill = _FILL_SECTION
        if col == 1:
            cell.font = _FONT_SECTION
            cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 20
    return row + 1


def _write_legend(
    ws: Worksheet,
    row: int,
    *,
    legend_styles: Dict[str, Dict[str, Any]],
) -> int:
    """Строка легенды цветов типов ввода (образцы в C–G)."""
    c0 = ws.cell(row=row, column=1, value="#META:LEGEND")
    c0.font = _FONT_LEGEND
    c1 = ws.cell(row=row, column=2, value="Цвет значения = тип ввода →")
    c1.font = _FONT_LEGEND
    for col_idx, kind in enumerate(INPUT_KIND_ORDER):
        label = INPUT_KIND_LABELS.get(kind, kind)
        cell = ws.cell(row=row, column=3 + col_idx, value=_s(label))
        style = legend_styles.get(kind, {})
        _apply_cell_style(cell, {k: v for k, v in style.items() if v is not None})
    ws.row_dimensions[row].height = 22
    return row + 1


def _write_table(
    ws: Worksheet,
    row: int,
    marker: str,
    columns: Sequence[str],
    rows: List[Dict[str, Any]],
    *,
    value_styles: Dict[str, Dict[str, Any]],
    min_empty: int = 3,
) -> Tuple[int, str, Sequence[str], int, int]:
    """Возвращает (next_row, table_key, columns, data_start, data_end)."""
    table_key = marker.replace("#TABLE:", "").strip().upper()
    if table_key == "REWARD_LINK":
        table_key = "REWARD-LINK"
    hints = TABLE_COLUMN_HINTS.get(table_key, {})
    col_kinds = [input_kind_for_table_col(table_key, c) for c in columns]
    col_styles = [
        value_styles.get(k) or value_styles["text"] for k in col_kinds
    ]

    mcell = ws.cell(row=row, column=1, value=_s(marker))
    mcell.fill = _FILL_TABLE
    mcell.font = Font(bold=True)
    row += 1
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=_s(col_name))
        cell.font = Font(bold=True)
        cell.fill = _FILL_TABLE
    row += 1
    for col_idx, col_name in enumerate(columns, start=1):
        hint = hints.get(col_name, "значение")
        kind_label = INPUT_KIND_LABELS.get(col_kinds[col_idx - 1], "")
        hint_full = f"{hint} · {kind_label}" if kind_label else hint
        text = f"#HINT | {hint_full}" if col_idx == 1 else hint_full
        cell = ws.cell(row=row, column=col_idx, value=_s(text))
        cell.font = _FONT_HINT
        cell.fill = _FILL_HINT
        cell.alignment = Alignment(wrap_text=True)
    row += 1
    data_start = row
    for data_row in rows:
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(
                row=row,
                column=col_idx,
                value=_s(data_row.get(col_name, "")),
            )
            _apply_cell_style(cell, col_styles[col_idx - 1])
        row += 1
    for _ in range(min_empty):
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=row, column=col_idx, value="")
            _apply_cell_style(cell, col_styles[col_idx - 1])
        row += 1
    data_end = row - 1
    return row, table_key, columns, data_start, data_end


def _write_sheet(
    ws: Worksheet,
    payload: Dict[str, Any],
    *,
    value_styles: Dict[str, Dict[str, Any]],
    legend_styles: Dict[str, Dict[str, Any]],
    dropdown_map: Dict[str, List[str]],
    list_ranges: Dict[str, str],
    with_dropdowns: bool,
) -> None:
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 62
    for col_idx in range(5, 8):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    kv_cells: Dict[str, List[str]] = {}

    row = 1
    ws.cell(row=row, column=1, value="#META:FORM_VERSION")
    ws.cell(row=row, column=2, value="6")
    row = 2
    ws.cell(row=row, column=1, value="#META:NOTE")
    note = (
        "Заполняйте цветные ячейки значений (столбец C и таблицы). "
        "Цвет = тип ввода (легенда ниже). Столбец D — описание. "
        "Где зелёный — выпадающий список. Персик — несколько значений через ; . "
        "Розовый — JSON как в SPOD. Лист = номер конкурса."
    )
    note_cell = ws.cell(row=row, column=2, value=_s(note))
    note_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 40
    row = 3
    row = _write_legend(ws, row, legend_styles=legend_styles)
    row = 4
    for col_idx, (title, fill) in enumerate(
        (
            ("Ключ", _FILL_HEADER),
            ("Подпись", _FILL_HEADER),
            ("Значение (заполнять)", _FILL_HEADER_VAL),
            ("Описание / значения", _FILL_HEADER_DESC),
        ),
        start=1,
    ):
        cell = ws.cell(row=row, column=col_idx, value=title)
        cell.font = Font(bold=True)
        cell.fill = fill
    row = 6

    contest_flat: Dict[str, Any] = payload.get("contest_flat") or {}
    contest_arrays: Dict[str, Any] = payload.get("contest_arrays") or {}
    feature: Dict[str, Any] = payload.get("contest_feature") or {}

    def _kv(
        r: int,
        key: str,
        label: str,
        value: Any,
        *,
        in_badge: bool,
        schema_kind: Optional[str] = None,
    ) -> int:
        return _write_kv(
            ws,
            r,
            key,
            label,
            value,
            in_badge_slot=in_badge,
            schema_kind=schema_kind,
            value_styles=value_styles,
            kv_cells=kv_cells,
            dropdown_map=dropdown_map,
        )

    row = _write_section(ws, row, "#SECTION:CONTEST")
    for key, label in schema.CONTEST_FLAT_FIELDS:
        row = _kv(row, key, label, contest_flat.get(key, ""), in_badge=False)
    for key, label in schema.CONTEST_ARRAY_FIELDS:
        row = _kv(
            row,
            key,
            label,
            contest_arrays.get(key, ""),
            in_badge=False,
            schema_kind="list",
        )
    for key, label, kind in schema.CONTEST_FEATURE_FIELDS:
        row = _kv(
            row,
            f"FEATURE.{key}",
            label,
            _feature_value(feature, key, kind),
            in_badge=False,
            schema_kind=kind,
        )
    row += 1

    badges: List[Dict[str, Any]] = list(payload.get("badges") or [])
    contest_type = str(contest_flat.get("CONTEST_TYPE") or "")
    slots = schema.max_badge_slots(contest_type)
    for slot_idx in range(1, slots + 1):
        row = _write_section(ws, row, f"#SECTION:BADGE:{slot_idx}")
        badge = badges[slot_idx - 1] if slot_idx - 1 < len(badges) else {}
        flat = badge.get("flat") or {}
        add_data = badge.get("add_data") or {}
        for key, label in schema.REWARD_FLAT_FIELDS:
            default = "BADGE" if key == "REWARD_TYPE" else ""
            row = _kv(
                row,
                key,
                label,
                flat.get(key, default),
                in_badge=True,
            )
        for key, label, kind in schema.REWARD_ADD_DATA_FIELDS:
            row = _kv(
                row,
                f"ADD.{key}",
                label,
                _feature_value(add_data, key, kind),
                in_badge=True,
                schema_kind=kind,
            )
        row += 1

    table_metas: List[Tuple[str, Sequence[str], int, int]] = []
    for marker, cols, payload_key in (
        ("#TABLE:REWARD-LINK", schema.REWARD_LINK_COLUMNS, "reward_link"),
        ("#TABLE:GROUP", schema.GROUP_COLUMNS, "group"),
        ("#TABLE:INDICATOR", schema.INDICATOR_COLUMNS, "indicator"),
        ("#TABLE:SCHEDULE", schema.SCHEDULE_COLUMNS, "schedule"),
    ):
        row, table_key, cols_out, data_start, data_end = _write_table(
            ws,
            row,
            marker,
            cols,
            list(payload.get(payload_key) or []),
            value_styles=value_styles,
        )
        table_metas.append((table_key, cols_out, data_start, data_end))

    if not with_dropdowns:
        return

    for key, cells in kv_cells.items():
        values = dropdown_map.get(key)
        if not values:
            continue
        formula = _dv_formula(key, values, list_ranges)
        _apply_list_validation(ws, cells, formula)

    for table_key, columns, data_start, data_end in table_metas:
        col_lists = TABLE_DROPDOWNS.get(table_key) or {}
        if not col_lists or data_end < data_start:
            continue
        for col_name, values in col_lists.items():
            if col_name not in columns:
                continue
            col_idx = list(columns).index(col_name) + 1
            letter = get_column_letter(col_idx)
            range_a1 = f"{letter}{data_start}:{letter}{data_end}"
            list_key = f"TBL:{table_key}:{col_name}"
            formula = _dv_formula(list_key, values, list_ranges)
            _apply_list_validation(ws, [range_a1], formula)
