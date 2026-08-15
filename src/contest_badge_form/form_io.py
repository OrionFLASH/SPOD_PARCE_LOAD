# -*- coding: utf-8 -*-
"""Чтение и запись Excel-формы конкурса BADGE (листы 1, 2, 3…)."""

from __future__ import annotations

import logging
import os
import re
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from src.contest_badge_form import schema
from src.contest_badge_form.field_meta import (
    TABLE_COLUMN_HINTS,
    TABLE_DROPDOWNS,
    description_for,
    merge_dropdowns,
)
from src.contest_badge_form.spod_json import (
    form_cell_from_list,
    list_from_form_cell,
    parse_spod_json,
)

_SHEET_NUM_RE = re.compile(r"^\d+$")
_LISTS_SHEET = "LISTS"  # видимый лист справочников (не _LISTS — ломает Excel)
_EXCEL_MAX_CELL = 32000

_FILL_SECTION = PatternFill("solid", fgColor="1F4E79")
_FONT_SECTION = Font(bold=True, color="FFFFFF", size=12)
_FILL_KV = PatternFill("solid", fgColor="D6EAF8")
_FILL_VALUE = PatternFill("solid", fgColor="FFF2CC")
_FILL_DESC = PatternFill("solid", fgColor="F8F9F9")
_FILL_TABLE = PatternFill("solid", fgColor="D5F5E3")
_FILL_HINT = PatternFill("solid", fgColor="E8F8F5")
_FONT_KEY = Font(name="Calibri", size=10, color="566573")
_FONT_DESC = Font(name="Calibri", size=9, color="5D6D7E", italic=True)
_FONT_HINT = Font(name="Calibri", size=8, color="1E8449", italic=True)

# Запрещённые в ячейках Excel управляющие символы
_ILLEGAL_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _excel_safe_text(value: Any) -> str:
    """Текст, безопасный для ячейки Excel (длина и символы)."""
    if value is None:
        return ""
    s = str(value)
    s = _ILLEGAL_RE.sub("", s)
    # Excel не любит NUL и часть control; также срезаем сверхлимит
    if len(s) > _EXCEL_MAX_CELL:
        s = s[: _EXCEL_MAX_CELL - 20] + "…[обрезано]"
    return s


def _clean_dropdown_values(
    values: Sequence[Any], *, pad_for_range: bool = False
) -> List[str]:
    """Убрать пустые; при pad_for_range — минимум 2 строки для диапазона на листе."""
    cleaned: List[str] = []
    for v in values:
        s = _excel_safe_text(v).strip()
        if s == "":
            continue
        if s not in cleaned:
            cleaned.append(s)
    if not cleaned:
        return ["—"]
    if pad_for_range and len(cleaned) == 1:
        cleaned.append(cleaned[0])
    return cleaned


def _dv_formula_for_list(values: List[str], list_ranges: Dict[str, str], key: str) -> str:
    """Формула списка: короткий без запятых — inline; иначе ссылка на LISTS."""
    cleaned = _clean_dropdown_values(values, pad_for_range=False)
    if all("," not in v for v in cleaned) and len(",".join(cleaned)) <= 240:
        return '"' + ",".join(cleaned) + '"'
    if key in list_ranges:
        return list_ranges[key]
    joined = ",".join(v.replace(",", " ") for v in cleaned)
    if len(joined) > 240:
        joined = ",".join(cleaned[:15])
    return f'"{joined}"'


def _write_section_header(ws: Worksheet, row: int, title: str) -> int:
    """Заголовок секции без merge (merge часто ломает Excel на Mac)."""
    for col in range(1, 5):
        cell = ws.cell(row=row, column=col)
        cell.fill = _FILL_SECTION
        if col == 1:
            cell.value = title
            cell.font = _FONT_SECTION
            cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 22
    return row + 1


def _write_kv(
    ws: Worksheet,
    row: int,
    key: str,
    label: str,
    value: Any,
    *,
    in_badge_slot: bool = False,
) -> int:
    """A=ключ, B=подпись, C=значение (жёлтое), D=описание."""
    ws.cell(row=row, column=1, value=key).font = _FONT_KEY
    ws.cell(row=row, column=2, value=label)
    val_cell = ws.cell(
        row=row, column=3, value=_excel_safe_text(value)
    )
    desc_cell = ws.cell(
        row=row,
        column=4,
        value=_excel_safe_text(description_for(key, in_badge_slot=in_badge_slot)),
    )
    ws.cell(row=row, column=1).fill = _FILL_KV
    ws.cell(row=row, column=2).fill = _FILL_KV
    val_cell.fill = _FILL_VALUE
    val_cell.alignment = Alignment(wrap_text=True, vertical="center")
    desc_cell.fill = _FILL_DESC
    desc_cell.font = _FONT_DESC
    desc_cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = 32
    return row + 1


def _write_table(
    ws: Worksheet,
    row: int,
    marker: str,
    columns: Sequence[str],
    rows: List[Dict[str, Any]],
    min_empty: int = 3,
) -> Tuple[int, int, int]:
    """Таблица с #HINT. Возвращает (next_row, data_start, data_end)."""
    table_key = marker.replace("#TABLE:", "").strip().upper()
    if table_key == "REWARD_LINK":
        table_key = "REWARD-LINK"
    hints = TABLE_COLUMN_HINTS.get(table_key, {})

    ws.cell(row=row, column=1, value=marker)
    ws.cell(row=row, column=1).fill = _FILL_TABLE
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=_excel_safe_text(col_name))
        cell.font = Font(bold=True)
        cell.fill = _FILL_TABLE
    row += 1
    ws.cell(row=row, column=1, value="#HINT")
    for col_idx, col_name in enumerate(columns, start=1):
        hint = hints.get(col_name, "значение")
        cell = ws.cell(
            row=row,
            column=col_idx,
            value=_excel_safe_text(
                f"#HINT | {hint}" if col_idx == 1 else hint
            ),
        )
        cell.font = _FONT_HINT
        cell.fill = _FILL_HINT
        cell.alignment = Alignment(wrap_text=True)
    row += 1
    data_start = row
    for data_row in rows:
        for col_idx, col_name in enumerate(columns, start=1):
            val = data_row.get(col_name, "")
            cell = ws.cell(
                row=row,
                column=col_idx,
                value=_excel_safe_text("" if val is None else val),
            )
            cell.fill = _FILL_VALUE
        row += 1
    for _ in range(min_empty):
        for col_idx in range(1, len(columns) + 1):
            # Явная пустая строка + заливка (не «голый» стиль — ломает Excel)
            cell = ws.cell(row=row, column=col_idx, value="")
            cell.fill = _FILL_VALUE
        row += 1
    data_end = row - 1
    return row, data_start, data_end


def _feature_value_for_form(feature: Dict[str, Any], key: str, kind: str) -> str:
    raw = feature.get(key, [] if kind == "list" else "")
    if kind == "list":
        if isinstance(raw, list):
            return form_cell_from_list(raw)
        return form_cell_from_list(list_from_form_cell(raw))
    if raw is None:
        return ""
    if isinstance(raw, (list, dict)):
        return form_cell_from_list(raw) if isinstance(raw, list) else str(raw)
    return str(raw)


def _ensure_lists_sheet(
    wb: Workbook, dropdowns: Dict[str, List[str]]
) -> Dict[str, str]:
    """
    Видимый лист LISTS в конце книги (справочники для длинных списков).
    Пустые значения не пишутся; в диапазоне ≥2 строки.
    """
    for legacy in ("_LISTS", _LISTS_SHEET, "LISTS"):
        if legacy in wb.sheetnames:
            del wb[legacy]
    ws = wb.create_sheet(_LISTS_SHEET)  # в конец
    ws.sheet_state = "visible"
    ranges: Dict[str, str] = {}
    col_idx = 1
    for key, values in sorted(dropdowns.items()):
        cleaned = _clean_dropdown_values(values, pad_for_range=True)
        # На лист — только списки, которые нельзя безопасно дать inline
        if all("," not in v for v in cleaned) and len(",".join(dict.fromkeys(cleaned))) <= 240:
            continue
        # для диапазона — уникальные + pad
        uniq = list(dict.fromkeys(cleaned))
        if len(uniq) == 1:
            uniq.append(uniq[0])
        ws.cell(row=1, column=col_idx, value=_excel_safe_text(key)[:80])
        for i, val in enumerate(uniq, start=2):
            ws.cell(row=i, column=col_idx, value=val)
        last = 1 + len(uniq)
        letter = get_column_letter(col_idx)
        # Имя без кавычек, если нет спецсимволов
        ranges[key] = f"{_LISTS_SHEET}!${letter}$2:${letter}${last}"
        col_idx += 1
    if col_idx == 1:
        # пустой справочник — одна ячейка-заглушка, чтобы лист не был «битым»
        ws.cell(row=1, column=1, value="EMPTY")
        ws.cell(row=2, column=1, value="—")
    return ranges


def _apply_kv_dropdowns(
    ws: Worksheet,
    dropdowns: Dict[str, List[str]],
    list_ranges: Dict[str, str],
) -> None:
    key_rows: Dict[str, List[int]] = {}
    for r in range(1, (ws.max_row or 1) + 1):
        key = str(ws.cell(row=r, column=1).value or "").strip()
        if not key or key.startswith("#"):
            continue
        if key in {"Ключ", "Key"}:
            continue
        key_rows.setdefault(key, []).append(r)

    for field_key, values in dropdowns.items():
        if field_key.startswith("TBL:"):
            continue
        if not values or field_key not in key_rows:
            continue
        cleaned = _clean_dropdown_values(values)
        formula = _dv_formula_for_list(cleaned, list_ranges, field_key)
        dv = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=True,
            showDropDown=False,
            showErrorMessage=False,
            showInputMessage=False,
        )
        ws.add_data_validation(dv)
        for r in key_rows[field_key]:
            dv.add(ws.cell(row=r, column=3))


def _apply_table_dropdowns(
    ws: Worksheet,
    table_marker: str,
    columns: Sequence[str],
    data_start: int,
    data_end: int,
    list_ranges: Dict[str, str],
) -> None:
    table_key = table_marker.replace("#TABLE:", "").strip().upper()
    if table_key == "REWARD_LINK":
        table_key = "REWARD-LINK"
    col_lists = TABLE_DROPDOWNS.get(table_key) or {}
    if not col_lists or data_end < data_start:
        return
    for col_name, values in col_lists.items():
        if col_name not in columns or not values:
            continue
        col_idx = list(columns).index(col_name) + 1
        list_key = f"TBL:{table_key}:{col_name}"
        cleaned = _clean_dropdown_values(values)
        formula = _dv_formula_for_list(cleaned, list_ranges, list_key)
        dv = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=True,
            showDropDown=False,
            showErrorMessage=False,
            showInputMessage=False,
        )
        ws.add_data_validation(dv)
        letter = get_column_letter(col_idx)
        dv.add(f"{letter}{data_start}:{letter}{data_end}")


def write_contest_sheet(
    ws: Worksheet,
    payload: Dict[str, Any],
    dropdowns: Optional[Dict[str, List[str]]] = None,
    list_ranges: Optional[Dict[str, str]] = None,
    *,
    apply_dv: bool = True,
) -> List[Tuple[str, Sequence[str], int, int]]:
    """
    A ключ | B подпись | C значение (жёлтое) | D описание.
    Возвращает спецификации таблиц для отложенной валидации.
    """
    dropdowns = merge_dropdowns(dropdowns)
    list_ranges = list_ranges or {}
    row = 1
    ws.cell(row=row, column=1, value="#META:FORM_VERSION")
    ws.cell(row=row, column=2, value="3")
    row = 2
    ws.cell(row=row, column=1, value="#META:NOTE")
    note = (
        "Заполняйте только жёлтые ячейки (столбец C). "
        "Столбец D — описание и допустимые значения. "
        "Массивы — через точку с запятой (;). "
        "Лист = номер конкурса."
    )
    ws.cell(row=row, column=2, value=_excel_safe_text(note))
    ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[row].height = 48
    row = 3
    headers = ("Ключ", "Подпись", "Значение (заполнять)", "Описание / значения")
    fills = (_FILL_KV, _FILL_KV, _FILL_VALUE, _FILL_DESC)
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=title)
        cell.font = Font(bold=True)
        cell.fill = fills[col_idx - 1]
    row = 5

    contest_flat: Dict[str, Any] = payload.get("contest_flat") or {}
    contest_arrays: Dict[str, Any] = payload.get("contest_arrays") or {}
    feature: Dict[str, Any] = payload.get("contest_feature") or {}

    row = _write_section_header(ws, row, "#SECTION:CONTEST")
    for key, label in schema.CONTEST_FLAT_FIELDS:
        row = _write_kv(ws, row, key, label, contest_flat.get(key, ""))
    for key, label in schema.CONTEST_ARRAY_FIELDS:
        row = _write_kv(ws, row, key, label, contest_arrays.get(key, ""))
    for key, label, kind in schema.CONTEST_FEATURE_FIELDS:
        form_key = f"FEATURE.{key}"
        row = _write_kv(
            ws,
            row,
            form_key,
            label,
            _feature_value_for_form(feature, key, kind),
        )
    row += 1

    badges: List[Dict[str, Any]] = list(payload.get("badges") or [])
    contest_type = str(contest_flat.get("CONTEST_TYPE") or "")
    slots = schema.max_badge_slots(contest_type)
    for slot_idx in range(1, slots + 1):
        row = _write_section_header(ws, row, f"#SECTION:BADGE:{slot_idx}")
        badge = badges[slot_idx - 1] if slot_idx - 1 < len(badges) else {}
        flat = badge.get("flat") or {}
        add_data = badge.get("add_data") or {}
        for key, label in schema.REWARD_FLAT_FIELDS:
            default = "BADGE" if key == "REWARD_TYPE" else ""
            row = _write_kv(
                ws,
                row,
                key,
                label,
                flat.get(key, default),
                in_badge_slot=True,
            )
        for key, label, kind in schema.REWARD_ADD_DATA_FIELDS:
            form_key = f"ADD.{key}"
            row = _write_kv(
                ws,
                row,
                form_key,
                label,
                _feature_value_for_form(add_data, key, kind),
                in_badge_slot=True,
            )
        row += 1

    table_specs = [
        ("#TABLE:REWARD-LINK", schema.REWARD_LINK_COLUMNS, "reward_link"),
        ("#TABLE:GROUP", schema.GROUP_COLUMNS, "group"),
        ("#TABLE:INDICATOR", schema.INDICATOR_COLUMNS, "indicator"),
        ("#TABLE:SCHEDULE", schema.SCHEDULE_COLUMNS, "schedule"),
    ]
    table_meta: List[Tuple[str, Sequence[str], int, int]] = []
    for marker, cols, payload_key in table_specs:
        row, data_start, data_end = _write_table(
            ws,
            row,
            marker,
            cols,
            list(payload.get(payload_key) or []),
        )
        table_meta.append((marker, cols, data_start, data_end))
        if apply_dv:
            _apply_table_dropdowns(
                ws, marker, cols, data_start, data_end, list_ranges
            )

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 96
    for col_idx in range(5, 17):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16

    # freeze_panes и data validation намеренно не ставим — на Excel Mac часто «файл повреждён»
    return table_meta


def build_workbook(
    payloads: List[Dict[str, Any]],
    dropdowns: Optional[Dict[str, List[str]]] = None,
    *,
    with_dropdowns: bool = False,
) -> Workbook:
    """
    Собрать книгу: листы 1…N.
    Data validation по умолчанию ВЫКЛ (Excel Mac часто не открывает файлы с десятками DV).
    Допустимые значения остаются в столбце D (описания).
    """
    # dropdowns зарезервированы; списки — в field_meta / столбец D
    _ = dropdowns
    _ = with_dropdowns

    wb = Workbook()
    default = wb.active
    first = True
    use_payloads = payloads if payloads else [{}]
    for idx, payload in enumerate(use_payloads, start=1):
        if first:
            ws = default
            ws.title = str(idx)
            first = False
        else:
            ws = wb.create_sheet(str(idx))
        write_contest_sheet(
            ws,
            payload,
            dropdowns={},
            list_ranges={},
            apply_dv=False,
        )
    return wb


def save_workbook(
    wb: Workbook, path: str, *, keep_data_validations: bool = False
) -> None:
    """Сохранить книгу и починить OOXML пустых ячеек (иначе Excel пишет «ошибка содержимого»)."""
    parent = os.path.dirname(os.path.abspath(path))
    if parent:
        os.makedirs(parent, exist_ok=True)
    wb.save(path)
    _postprocess_xlsx(path, keep_data_validations=keep_data_validations)
    logging.info("[contest_badge_form] Форма записана: %s", path)


def _postprocess_xlsx(path: str, *, keep_data_validations: bool = False) -> None:
    """
    openpyxl 3.x для пустой строки часто пишет:
      <c r="C6" s="9" t="inlineStr" />
    без <is><t/></is> — Excel считает книгу повреждённой.
    Также <c ... t="n" /> без <v> — недопустимо.
    """
    import io
    import zipfile

    buf = io.BytesIO()
    with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(
        buf, "w", compression=zipfile.ZIP_DEFLATED
    ) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.startswith("xl/worksheets/sheet") and item.filename.endswith(
                ".xml"
            ):
                text = data.decode("utf-8")
                # пустой inlineStr → корректный пустой rich string
                text = re.sub(
                    r'<c r="([^"]+)"([^>]*) t="inlineStr"\s*/>',
                    r'<c r="\1"\2 t="inlineStr"><is><t></t></is></c>',
                    text,
                )
                # styled empty number cell → просто стиль без типа
                text = re.sub(
                    r'<c r="([^"]+)" s="(\d+)" t="n"\s*/>',
                    r'<c r="\1" s="\2"/>',
                    text,
                )
                if not keep_data_validations:
                    # DV иногда ломает Excel Mac — убираем только для «простых» книг
                    text = re.sub(
                        r"<dataValidations\b[^>]*>.*?</dataValidations>",
                        "",
                        text,
                        flags=re.DOTALL,
                    )
                data = text.encode("utf-8")
            elif item.filename == "xl/workbook.xml":
                text = data.decode("utf-8")
                text = text.replace("<workbookProtection />", "")
                text = text.replace("<workbookProtection/>", "")
                data = text.encode("utf-8")
            zout.writestr(item.filename, data)
    with open(path, "wb") as f:
        f.write(buf.getvalue())


def empty_contest_payload(
    *,
    contest_type: str = "ТУРНИРНЫЙ",
) -> Dict[str, Any]:
    """Пустой payload одного листа формы (с FIELD_DEFAULTS из field_meta)."""
    from src.contest_badge_form.field_meta import default_for

    contest_flat: Dict[str, str] = {}
    for key, _ in schema.CONTEST_FLAT_FIELDS:
        contest_flat[key] = default_for(key)
    contest_flat["CONTEST_TYPE"] = contest_type
    contest_arrays: Dict[str, str] = {}
    for key, _ in schema.CONTEST_ARRAY_FIELDS:
        contest_arrays[key] = default_for(key)

    feature: Dict[str, Any] = schema.empty_feature_template()
    for leaf, _label, kind in schema.CONTEST_FEATURE_FIELDS:
        form_key = f"FEATURE.{leaf}"
        raw = default_for(form_key)
        if raw == "":
            continue
        if kind == "list":
            feature[leaf] = [p.strip() for p in raw.split(";") if p.strip()]
        else:
            feature[leaf] = raw

    badges: List[Dict[str, Any]] = []
    for _ in range(schema.max_badge_slots(contest_type)):
        flat = {key: "" for key, _ in schema.REWARD_FLAT_FIELDS}
        for key, _ in schema.REWARD_FLAT_FIELDS:
            flat[key] = default_for(key, in_badge_slot=True)
        flat["REWARD_TYPE"] = "BADGE"
        add_data = schema.empty_add_data_template()
        for leaf, _label, kind in schema.REWARD_ADD_DATA_FIELDS:
            form_key = f"ADD.{leaf}"
            raw = default_for(form_key)
            if raw == "":
                continue
            if kind == "list":
                add_data[leaf] = [
                    p.strip() for p in raw.split(";") if p.strip()
                ]
            else:
                add_data[leaf] = raw
        badges.append({"flat": flat, "add_data": add_data})
    return {
        "contest_flat": contest_flat,
        "contest_arrays": contest_arrays,
        "contest_feature": feature,
        "badges": badges,
        "reward_link": [],
        "group": [],
        "indicator": [],
        "schedule": [],
    }


def create_blank_form(
    path: str,
    *,
    sheet_count: int = 1,
    contest_type: str = "ТУРНИРНЫЙ",
    dropdowns: Optional[Dict[str, List[str]]] = None,
    catalog_path: Optional[str] = None,
) -> str:
    """
    Создать пустую Excel-форму для заполнения (stdlib OOXML + dropdowns).
    Если задан catalog_path — подписи/описания/дефолты/списки из catalog.json.
    """
    from src.contest_badge_form.catalog_loader import (
        clear_param_catalog,
        load_param_catalog,
    )

    loaded = False
    if catalog_path:
        load_param_catalog(catalog_path)
        loaded = True
    try:
        n = max(1, int(sheet_count))
        payloads = [
            empty_contest_payload(contest_type=contest_type) for _ in range(n)
        ]
        from src.contest_badge_form.xlsx_write import write_form_xlsx

        return write_form_xlsx(
            path, payloads, dropdowns=dropdowns, with_dropdowns=True
        )
    finally:
        if loaded:
            clear_param_catalog()


def _is_marker(value: Any) -> bool:
    s = str(value or "").strip()
    return (
        s.startswith("#SECTION:")
        or s.startswith("#TABLE:")
        or s.startswith("#META:")
        or s.startswith("#HINT")
    )


def _read_kv_block(
    ws: Worksheet, start_row: int
) -> Tuple[Dict[str, str], int]:
    """Читать KEY / VALUE(C) до следующего маркера."""
    data: Dict[str, str] = {}
    row = start_row
    max_row = ws.max_row or start_row
    while row <= max_row:
        a = ws.cell(row=row, column=1).value
        if _is_marker(a) and row != start_row:
            break
        if a is not None and str(a).strip() and not str(a).startswith("#"):
            key = str(a).strip()
            if key in {"Ключ", "Key"}:
                row += 1
                continue
            val = ws.cell(row=row, column=3).value
            data[key] = "" if val is None else str(val)
        row += 1
        if row - start_row > 500:
            break
    return data, row


def _read_table_block(
    ws: Worksheet, start_row: int, columns_hint: Sequence[str]
) -> Tuple[List[Dict[str, str]], int]:
    """#TABLE → заголовки → #HINT → данные."""
    header_row = start_row + 1
    headers: List[str] = []
    for col_idx in range(1, max(len(columns_hint), 30) + 1):
        h = ws.cell(row=header_row, column=col_idx).value
        if h is None or str(h).strip() == "":
            if col_idx == 1:
                break
            break
        headers.append(str(h).strip())
    if not headers:
        headers = list(columns_hint)
    rows: List[Dict[str, str]] = []
    row = header_row + 1
    max_row = ws.max_row or row
    while row <= max_row:
        a = str(ws.cell(row=row, column=1).value or "").strip()
        if a.startswith("#HINT"):
            row += 1
            continue
        break
    while row <= max_row:
        a = ws.cell(row=row, column=1).value
        if _is_marker(a):
            break
        values = [
            ws.cell(row=row, column=c).value for c in range(1, len(headers) + 1)
        ]
        if all(v is None or str(v).strip() == "" for v in values):
            peek = row + 1
            if peek <= max_row and _is_marker(ws.cell(row=peek, column=1).value):
                row = peek
                break
            row += 1
            if row <= max_row:
                nxt = [
                    ws.cell(row=row, column=c).value
                    for c in range(1, len(headers) + 1)
                ]
                if all(v is None or str(v).strip() == "" for v in nxt):
                    while row <= max_row and not _is_marker(
                        ws.cell(row=row, column=1).value
                    ):
                        row += 1
                    break
            continue
        item = {
            headers[i]: ("" if values[i] is None else str(values[i]))
            for i in range(len(headers))
        }
        if str(item.get(headers[0], "")).startswith("#"):
            row += 1
            continue
        rows.append(item)
        row += 1
    return rows, row


def read_contest_sheet(ws: Worksheet) -> Dict[str, Any]:
    """Разобрать один лист формы в payload."""
    payload: Dict[str, Any] = {
        "contest_flat": {},
        "contest_arrays": {},
        "contest_feature": schema.empty_feature_template(),
        "badges": [],
        "reward_link": [],
        "group": [],
        "indicator": [],
        "schedule": [],
    }
    row = 1
    max_row = ws.max_row or 1
    current_badge_slot: Optional[int] = None
    badge_kv: Dict[str, str] = {}

    def _flush_badge() -> None:
        nonlocal badge_kv, current_badge_slot
        if current_badge_slot is None:
            return
        flat = {k: badge_kv.get(k, "") for k, _ in schema.REWARD_FLAT_FIELDS}
        add_data: Dict[str, Any] = {}
        for key, _label, kind in schema.REWARD_ADD_DATA_FIELDS:
            raw = badge_kv.get(f"ADD.{key}", "")
            if kind == "list":
                add_data[key] = list_from_form_cell(raw)
            else:
                add_data[key] = raw
        # Пустой слот (нет кода) — не добавляем
        if str(flat.get("REWARD_CODE") or "").strip():
            if not flat.get("REWARD_TYPE"):
                flat["REWARD_TYPE"] = "BADGE"
            payload["badges"].append({"flat": flat, "add_data": add_data})
        badge_kv = {}
        current_badge_slot = None

    while row <= max_row:
        a = str(ws.cell(row=row, column=1).value or "").strip()
        if a.startswith("#SECTION:CONTEST"):
            _flush_badge()
            kv, row = _read_kv_block(ws, row + 1)
            flat_keys = {k for k, _ in schema.CONTEST_FLAT_FIELDS}
            array_keys = {k for k, _ in schema.CONTEST_ARRAY_FIELDS}
            feature: Dict[str, Any] = schema.empty_feature_template()
            contest_flat: Dict[str, str] = {}
            contest_arrays: Dict[str, str] = {}
            for key, val in kv.items():
                if key.startswith("FEATURE."):
                    leaf = key[len("FEATURE.") :]
                    kind = "list"
                    for fk, _fl, fknd in schema.CONTEST_FEATURE_FIELDS:
                        if fk == leaf:
                            kind = fknd
                            break
                    feature[leaf] = (
                        list_from_form_cell(val) if kind == "list" else val
                    )
                elif key in array_keys:
                    contest_arrays[key] = val
                elif key in flat_keys:
                    contest_flat[key] = val
                else:
                    contest_flat[key] = val
            payload["contest_flat"] = contest_flat
            payload["contest_arrays"] = contest_arrays
            payload["contest_feature"] = feature
            continue
        if a.startswith("#SECTION:BADGE:"):
            _flush_badge()
            try:
                current_badge_slot = int(a.split(":")[-1])
            except ValueError:
                current_badge_slot = 1
            badge_kv, row = _read_kv_block(ws, row + 1)
            continue
        if a.startswith("#TABLE:"):
            _flush_badge()
            cols = schema.table_columns_for(a)
            table_rows, row = _read_table_block(ws, row, cols)
            marker = a.replace("#TABLE:", "").strip().upper()
            if marker in {"REWARD-LINK", "REWARD_LINK"}:
                payload["reward_link"] = table_rows
            elif marker == "GROUP":
                payload["group"] = table_rows
            elif marker == "INDICATOR":
                payload["indicator"] = table_rows
            elif marker in {"SCHEDULE", "TOURNAMENT-SCHEDULE"}:
                payload["schedule"] = table_rows
            continue
        row += 1

    _flush_badge()
    return payload


def read_form_workbook(path: str) -> List[Dict[str, Any]]:
    """Прочитать все числовые листы формы."""
    wb = load_workbook(path, data_only=True)
    payloads: List[Dict[str, Any]] = []
    sheet_names = [n for n in wb.sheetnames if _SHEET_NUM_RE.match(str(n))]
    sheet_names.sort(key=lambda x: int(x))
    if not sheet_names:
        logging.warning(
            "[contest_badge_form] В книге нет листов 1,2,3…: %s", path
        )
    for name in sheet_names:
        payloads.append(read_contest_sheet(wb[name]))
        logging.info(
            "[contest_badge_form] Лист %s: CONTEST_CODE=%s, badges=%s",
            name,
            (payloads[-1].get("contest_flat") or {}).get("CONTEST_CODE"),
            len(payloads[-1].get("badges") or []),
        )
    return payloads


def df_rows_to_dicts(df_rows: Any, columns: Sequence[str]) -> List[Dict[str, str]]:
    """DataFrame/records → список dict по колонкам."""
    out: List[Dict[str, str]] = []
    if df_rows is None:
        return out
    # pandas DataFrame
    if hasattr(df_rows, "iterrows"):
        for _, series in df_rows.iterrows():
            item = {
                c: ("" if c not in series or series[c] is None else str(series[c]))
                for c in columns
            }
            out.append(item)
        return out
    for rec in df_rows:
        item = {c: str(rec.get(c, "") or "") for c in columns}
        out.append(item)
    return out


def payload_from_csv_bundle(
    contest_row: Dict[str, Any],
    badges_rows: List[Dict[str, Any]],
    add_data_list: List[Dict[str, Any]],
    reward_link: List[Dict[str, str]],
    group: List[Dict[str, str]],
    indicator: List[Dict[str, str]],
    schedule: List[Dict[str, str]],
) -> Dict[str, Any]:
    """Собрать payload формы из строк CSV."""
    contest_flat: Dict[str, str] = {}
    for key, _ in schema.CONTEST_FLAT_FIELDS:
        contest_flat[key] = str(contest_row.get(key, "") or "")
    contest_arrays: Dict[str, str] = {}
    for key, _ in schema.CONTEST_ARRAY_FIELDS:
        raw = contest_row.get(key, "")
        parsed = parse_spod_json(raw) if raw else None
        if isinstance(parsed, list):
            contest_arrays[key] = form_cell_from_list(parsed)
        else:
            contest_arrays[key] = "" if raw is None else str(raw)

    feature_raw = contest_row.get("CONTEST_FEATURE", "")
    feature_obj = parse_spod_json(feature_raw) if feature_raw else None
    if not isinstance(feature_obj, dict):
        feature_obj = schema.empty_feature_template()
    else:
        # дополнить отсутствующие ключи
        base = schema.empty_feature_template()
        base.update(feature_obj)
        feature_obj = base

    badges: List[Dict[str, Any]] = []
    for idx, brow in enumerate(badges_rows):
        flat = {k: str(brow.get(k, "") or "") for k, _ in schema.REWARD_FLAT_FIELDS}
        add = add_data_list[idx] if idx < len(add_data_list) else {}
        if not isinstance(add, dict):
            add = {}
        base_add = schema.empty_add_data_template()
        base_add.update(add)
        badges.append({"flat": flat, "add_data": base_add})

    return {
        "contest_flat": contest_flat,
        "contest_arrays": contest_arrays,
        "contest_feature": feature_obj,
        "badges": badges,
        "reward_link": reward_link,
        "group": group,
        "indicator": indicator,
        "schedule": schedule,
    }
