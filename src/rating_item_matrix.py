# -*- coding: utf-8 -*-
"""
Матрица на листе RATING: колонки ITEM из REWARD, заказы по ORDER, четыре состояния ячейки (число / Y / N).

Фильтр ORDER по статусу; лимиты item_order_groups; красная шапка при itemAmount.
Выполняется после merge_fields_advanced (развёрнутые колонки ADD_DATA на REWARD).
"""

from __future__ import annotations

import logging
import unicodedata
from collections import defaultdict
from typing import Any, Dict, List, Optional, Set, Union

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from src.reward_item_catalog import (
    build_item_catalog_from_reward_df,
    item_accessible_for_manager,
    rules_for_matrix_column,
)


def _as_float(x: Any) -> Optional[float]:
    """Преобразование значения ячейки в float; None если не число или пусто."""
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return None
    if isinstance(x, str):
        s = x.strip().replace(",", ".")
        if not s or s == "-":
            return None
        try:
            return float(s)
        except ValueError:
            return None
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def _norm_str(x: Any) -> str:
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return ""
    return str(x).strip()


# Запасные имена колонок для типичных выгрузок (русские заголовки и англ. поля gamification)
_DEFAULT_ORDER_EMP: List[str] = [
    "Табельный номер",
    "PERSON_NUMBER",
    "PERSON_NUMBER_ADD",
    "personNumber",
    "EMPLOYEE_NUMBER",
]
_DEFAULT_ORDER_PROD: List[str] = [
    "Код товара",
    "REWARD_CODE",
    "ITEM_CODE",
    "PRODUCT_CODE",
    "rewardCode",
    "REWARD",
]
_DEFAULT_RATING_EMP: List[str] = _DEFAULT_ORDER_EMP
_DEFAULT_COUNTRY_RANK: List[str] = [
    "Место в рейтинге по стране",
    "RANK_COUNTRY",
    "COUNTRY_RANK",
    "PLACE_COUNTRY",
    "countryRatingPlace",
    "ratingPlaceCountry",
]
_DEFAULT_TB_RANK: List[str] = [
    "Место в рейтинге ТБ",
    "RANK_TB",
    "TB_RANK",
    "PLACE_TB",
    "tbRatingPlace",
    "ratingPlaceTB",
]
_DEFAULT_GOSB_RANK: List[str] = [
    "Место в рейтинге ГОСБ",
    "RANK_GOSB",
    "GOSB_RANK",
    "PLACE_GOSB",
    "gosbRatingPlace",
    "ratingPlaceGOSB",
]
# Лист LIST-REWARDS: табельный и код награды (русские заголовки и запасные англ. имена)
_DEFAULT_LIST_RW_EMP: List[str] = [
    "Табельный номер сотрудника",
    "PERSON_NUMBER",
    "PERSON_NUMBER_ADD",
    "personNumber",
    "Табельный номер",
    "EMPLOYEE_NUMBER",
]
_DEFAULT_LIST_RW_CODE: List[str] = [
    "Код награды",
    "REWARD_CODE",
    "rewardCode",
    "Код",
]
_DEFAULT_CRYSTALS: List[str] = [
    "Количество кристаллов",
    "CRYSTAL_COUNT",
    "crystalsEarnedTotal",
    "crystalEarnedTotal",
]
_DEFAULT_ORDER_STATUS: List[str] = [
    "Статус заказа",
    "ORDER_STATUS",
    "orderStatus",
    "Статус",
]

# Ключи заливки ячеек матрицы (4 состояния + шапка при исчерпании itemAmount)
FILL_ORDERED_AVAILABLE = "ordered_available"
FILL_ORDERED_UNAVAILABLE = "ordered_unavailable"
FILL_AVAILABLE_NOT_ORDERED = "available_not_ordered"
FILL_UNAVAILABLE_NOT_ORDERED = "unavailable_not_ordered"


def _norm_header(s: str) -> str:
    """Нормализация имени столбца для сопоставления (BOM, NFKC, регистр)."""
    t = unicodedata.normalize("NFKC", str(s)).strip()
    if t.startswith("\ufeff"):
        t = t.lstrip("\ufeff").strip()
    return t.casefold()


def _build_header_index(df: pd.DataFrame) -> Dict[str, str]:
    """casefold-ключ -> исходное имя столбца в DataFrame."""
    out: Dict[str, str] = {}
    for c in df.columns:
        out[_norm_header(str(c))] = c
    return out


def _resolve_column(df: pd.DataFrame, names: Union[str, List[str], None], defaults: List[str]) -> Optional[str]:
    """
    Первое найденное имя столбца: сначала варианты из конфига (строка или список), затем defaults.
    Сравнение по нормализованным строкам и точному вхождению в columns.
    """
    candidates: List[str] = []
    if names is not None:
        if isinstance(names, str) and names.strip():
            candidates.append(names.strip())
        elif isinstance(names, list):
            candidates.extend(str(x).strip() for x in names if str(x).strip())
    candidates.extend(d for d in defaults if d not in candidates)

    idx = _build_header_index(df)
    for want in candidates:
        wn = _norm_header(want)
        if wn in idx:
            resolved = idx[wn]
            if want != resolved:
                logging.info(f"[rating_item_matrix] Столбец «{want}» сопоставлен с «{resolved}»")
            return resolved
        if want in df.columns:
            return want
    return None


def _find_column_by_fragments(df: pd.DataFrame, fragments: List[str]) -> Optional[str]:
    """Первый столбец, в имени которого встречаются все подстроки (без учёта регистра)."""
    fr = [f.casefold() for f in fragments]
    for c in df.columns:
        s = str(c).casefold()
        if all(f in s for f in fr):
            return c
    return None


def _strip_hex(color: str) -> str:
    """ARGB для openpyxl без символа #."""
    return str(color or "").strip().lstrip("#").upper()[:8]


def _resolve_fill_colors(cfg: Dict[str, Any]) -> Dict[str, str]:
    """Четыре цвета матрицы и шапки; устаревшие fill_accessibility_* — запасные."""
    ok_legacy = _strip_hex(cfg.get("fill_accessibility_ok") or "C6EFCE")
    fail_legacy = _strip_hex(cfg.get("fill_accessibility_fail") or "FFC7CE")
    return {
        FILL_ORDERED_AVAILABLE: _strip_hex(cfg.get("fill_ordered_available") or ok_legacy),
        FILL_ORDERED_UNAVAILABLE: _strip_hex(cfg.get("fill_ordered_unavailable") or "FFB6C1"),
        FILL_AVAILABLE_NOT_ORDERED: _strip_hex(cfg.get("fill_available_not_ordered") or "92D050"),
        FILL_UNAVAILABLE_NOT_ORDERED: _strip_hex(
            cfg.get("fill_unavailable_not_ordered") or fail_legacy
        ),
        "header_stock_out": _strip_hex(cfg.get("fill_header_stock_out") or "FF0000"),
    }


def _filter_order_dataframe(
    order_df: pd.DataFrame,
    cfg: Dict[str, Any],
) -> pd.DataFrame:
    """
    Исключает заказы со статусом из order_status_exclude.
    Если колонка статуса не найдена — все строки (WARNING), как в плане.
    """
    exclude_raw = cfg.get("order_status_exclude")
    if exclude_raw is None:
        exclude_list = ["Отклонён", "Отменён"]
    else:
        exclude_list = [str(x) for x in exclude_raw]
    exclude_set = {_norm_str(x) for x in exclude_list if _norm_str(x)}
    if not exclude_set:
        return order_df

    status_col = _resolve_column(
        order_df,
        cfg.get("order_status_col"),
        _DEFAULT_ORDER_STATUS,
    )
    if status_col is None:
        logging.warning(
            "[rating_item_matrix] Колонка статуса заказа не найдена — "
            "в расчёт попадают все строки ORDER"
        )
        return order_df

    before = len(order_df)
    status_series = order_df[status_col].map(_norm_str)
    filtered = order_df.loc[~status_series.isin(exclude_set)].copy()
    removed = before - len(filtered)
    logging.debug(
        f"[rating_item_matrix] ORDER: отфильтровано по статусу «{status_col}»: "
        f"{removed} строк (исключены: {sorted(exclude_set)})"
    )
    return filtered


def _order_counts_by_employee(
    order_df: pd.DataFrame,
    emp_col: str,
    prod_col: str,
) -> Dict[str, Dict[str, int]]:
    """Табельный -> {код товара -> число строк заказов}."""
    counts: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for _, row in order_df.iterrows():
        emp = _norm_str(row.get(emp_col))
        code = _norm_str(row.get(prod_col))
        if emp and code:
            counts[emp][code] += 1
    return {emp: dict(codes) for emp, codes in counts.items()}


def _parse_item_order_groups(cfg: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Группы кодов с лимитом суммарных заказов на строку RATING."""
    raw = cfg.get("item_order_groups") or []
    groups: List[Dict[str, Any]] = []
    for item in raw:
        if not isinstance(item, dict):
            continue
        codes = [_norm_str(c) for c in (item.get("codes") or []) if _norm_str(c)]
        try:
            max_orders = int(item.get("max_orders") or item.get("max_orders_in_group") or 0)
        except (TypeError, ValueError):
            max_orders = 0
        if codes and max_orders > 0:
            groups.append(
                {
                    "id": str(item.get("id") or ""),
                    "max_orders": max_orders,
                    "codes": codes,
                }
            )
    return groups


def _blocked_codes_for_row(
    count_by_code: Dict[str, int],
    groups: List[Dict[str, Any]],
) -> Set[str]:
    """
    Коды группы, по которым сумма заказов менеджера >= max_orders —
    все коды группы считаются недоступными для строки.
    """
    blocked: Set[str] = set()
    for grp in groups:
        codes: List[str] = grp.get("codes") or []
        try:
            limit = int(grp.get("max_orders") or 0)
        except (TypeError, ValueError):
            limit = 0
        if limit <= 0 or not codes:
            continue
        total = sum(int(count_by_code.get(c, 0)) for c in codes)
        if total >= limit:
            blocked.update(codes)
    return blocked


def _matrix_fill_key(accessible: bool, ordered: bool) -> str:
    """Ключ заливки по доступности и факту заказа."""
    if ordered:
        return FILL_ORDERED_AVAILABLE if accessible else FILL_ORDERED_UNAVAILABLE
    return FILL_AVAILABLE_NOT_ORDERED if accessible else FILL_UNAVAILABLE_NOT_ORDERED


def _item_amount_limit(rules: Dict[str, Any]) -> Optional[int]:
    """Лимит itemAmount из каталога REWARD; None — не задан."""
    raw = rules.get("itemAmount")
    if raw is None:
        return None
    try:
        val = int(float(raw))
        return val if val > 0 else None
    except (TypeError, ValueError):
        return None


def _make_unique_col_name(base: str, existing: set) -> str:
    """Уникальное имя колонки в Excel (дубликаты REWARD_CODE)."""
    name = base.strip() or "ITEM"
    if name not in existing:
        existing.add(name)
        return name
    k = 2
    while f"{name}__{k}" in existing:
        k += 1
    out = f"{name}__{k}"
    existing.add(out)
    return out


def _collect_item_reward_specs(
    reward_df: pd.DataFrame,
    rating_df: pd.DataFrame,
    cfg: Dict[str, Any],
) -> List[Dict[str, Any]]:
    """
    Строки REWARD с REWARD_TYPE == ITEM: код, сезон, пороги рейтинга из развёрнутых колонок.
    """
    rtc = cfg.get("reward_type_col") or "REWARD_TYPE"
    rcc = cfg.get("reward_code_col") or "REWARD_CODE"
    exact_season = cfg.get("col_season_code") or "ADD_DATA => getCondition => employeeRating => seasonCode"
    exact_bank = cfg.get("col_min_rating_bank") or "ADD_DATA => getCondition => employeeRating => minRatingBANK"
    exact_tb = cfg.get("col_min_rating_tb") or "ADD_DATA => getCondition => employeeRating => minRatingTB"
    exact_gosb = cfg.get("col_min_rating_gosb") or "ADD_DATA => getCondition => employeeRating => minRatingGOSB"

    def _metric_col(exact: str, frags: List[str]) -> Optional[str]:
        if exact in reward_df.columns:
            return exact
        found = _find_column_by_fragments(reward_df, frags)
        if found:
            logging.info(f"[rating_item_matrix] Порог рейтинга: вместо «{exact}» используется «{found}»")
        return found

    c_season = _metric_col(exact_season, ["employeeRating", "seasonCode"])
    c_bank = _metric_col(exact_bank, ["employeeRating", "minRatingBANK"])
    c_tb = _metric_col(exact_tb, ["employeeRating", "minRatingTB"])
    c_gosb = _metric_col(exact_gosb, ["employeeRating", "minRatingGOSB"])

    rtc_res = _resolve_column(reward_df, rtc, ["REWARD_TYPE", "reward_type", "RewardType"])
    rcc_res = _resolve_column(reward_df, rcc, ["REWARD_CODE", "reward_code", "RewardCode"])
    if rtc_res is None or rcc_res is None:
        logging.warning(
            f"[rating_item_matrix] На листе REWARD не найдены колонки типа/кода награды "
            f"(ожидались «{rtc}», «{rcc}»). Заголовки (первые 40): {list(reward_df.columns)[:40]}"
        )
        return []

    mask = reward_df[rtc_res].astype(str).str.strip().str.upper() == "ITEM"
    sub = reward_df.loc[mask].copy()
    if sub.empty:
        logging.info(
            f"[rating_item_matrix] Нет строк REWARD с типом ITEM (колонка «{rtc_res}»). "
            f"Уникальные значения типа (до 15): {reward_df[rtc_res].astype(str).str.strip().unique()[:15].tolist()}"
        )
        return []

    # Имена новых колонок не должны пересекаться с уже существующими на RATING
    existing_names: set = set(str(c) for c in rating_df.columns)
    rows: List[Dict[str, Any]] = []
    for _, row in sub.iterrows():
        code = _norm_str(row.get(rcc_res))
        if not code:
            continue
        season = _norm_str(row.get(c_season)) if c_season and c_season in sub.columns else ""
        col_name = code
        if season:
            col_name = f"{code} ({season})"
        col_name = _make_unique_col_name(col_name, existing_names)
        rows.append(
            {
                "col_name": col_name,
                "match_code": code,
                "min_bank": _as_float(row.get(c_bank)) if c_bank and c_bank in sub.columns else None,
                "min_tb": _as_float(row.get(c_tb)) if c_tb and c_tb in sub.columns else None,
                "min_gosb": _as_float(row.get(c_gosb)) if c_gosb and c_gosb in sub.columns else None,
                "sort_season": season or "",
            }
        )

    rows.sort(key=lambda r: (_norm_str(r["match_code"]).lower(), r["sort_season"].lower()))
    return rows


def apply_rating_item_matrix_enrichment(
    sheets_data: Dict[str, Any],
    cfg: Dict[str, Any],
) -> Optional[Dict[str, Any]]:
    """
    Добавляет на лист RATING колонки со счётчиками заказов по кодам ITEM (аналог СЧЁТЕСЛИМН).

    Возвращает метаданные для последующей подсветки или None, если шаг пропущен.
    """
    if not cfg or not bool(cfg.get("enabled")):
        return None

    sr = cfg.get("sheet_rating") or "RATING"
    so = cfg.get("sheet_order") or "ORDER"
    rw = cfg.get("sheet_reward") or "REWARD"

    for key, name in (("rating", sr), ("order", so), ("reward", rw)):
        if name not in sheets_data or sheets_data[name] is None:
            logging.warning(f"[rating_item_matrix] Нет листа «{name}» в данных — шаг пропущен")
            return None

    rating_t = sheets_data[sr]
    order_t = sheets_data[so]
    reward_t = sheets_data[rw]
    rating_df = rating_t[0]
    order_df = order_t[0]
    if not isinstance(rating_df, pd.DataFrame) or not isinstance(order_df, pd.DataFrame):
        return None

    emp_o = _resolve_column(order_df, cfg.get("order_employee_col"), _DEFAULT_ORDER_EMP)
    prod_o = _resolve_column(order_df, cfg.get("order_product_col"), _DEFAULT_ORDER_PROD)
    emp_r = _resolve_column(rating_df, cfg.get("rating_employee_col"), _DEFAULT_RATING_EMP)
    country_c = _resolve_column(rating_df, cfg.get("country_rank_col"), _DEFAULT_COUNTRY_RANK)
    tb_c = _resolve_column(rating_df, cfg.get("tb_rank_col"), _DEFAULT_TB_RANK)
    gosb_c = _resolve_column(rating_df, cfg.get("gosb_rank_col"), _DEFAULT_GOSB_RANK)

    if emp_o is None or prod_o is None or emp_r is None:
        logging.warning(
            f"[rating_item_matrix] Не удалось сопоставить обязательные столбцы ORDER/RATING. "
            f"ORDER: сотрудник={emp_o!r}, товар/код={prod_o!r}; RATING: сотрудник={emp_r!r}. "
            f"Колонки ORDER (до 30): {list(order_df.columns)[:30]}; RATING (до 30): {list(rating_df.columns)[:30]}"
        )
        return None

    if not any((country_c, tb_c, gosb_c)):
        logging.warning(
            f"[rating_item_matrix] Нет ни одного столбца места в рейтинге — "
            f"country={country_c!r}, tb={tb_c!r}, gosb={gosb_c!r}. "
            "Критерии minRating* по отсутствующим колонкам будут считаться невыполненными (красная ячейка), если порог > 0."
        )
    elif not (country_c and tb_c and gosb_c):
        logging.info(
            f"[rating_item_matrix] Часть столбцов рейтинга не найдена (country={country_c!r}, tb={tb_c!r}, gosb={gosb_c!r}) — "
            "доступность считается только по найденным местам в рейтинге."
        )

    specs = _collect_item_reward_specs(reward_t[0], rating_df, cfg)
    if not specs:
        logging.warning(
            "[rating_item_matrix] Список ITEM-наград пуст — колонки на RATING не добавлялись "
            "(проверьте REWARD_TYPE=ITEM и разворот ADD_DATA на листе REWARD)."
        )
        return None

    order_df = _filter_order_dataframe(order_df, cfg)
    counts_by_emp = _order_counts_by_employee(order_df, emp_o, prod_o)
    order_groups = _parse_item_order_groups(cfg)

    rating_df = rating_df.copy()
    thresholds: Dict[str, Dict[str, Optional[float]]] = {}
    added: List[str] = []
    col_values: Dict[str, List[Any]] = {sp["col_name"]: [] for sp in specs}

    for sp in specs:
        cname = sp["col_name"]
        added.append(cname)
        thresholds[cname] = {
            "bank": sp["min_bank"],
            "tb": sp["min_tb"],
            "gosb": sp["min_gosb"],
        }

    reward_df0 = reward_t[0]
    rtc_res2 = _resolve_column(
        reward_df0, cfg.get("reward_type_col"), ["REWARD_TYPE", "reward_type", "RewardType"]
    )
    rcc_res2 = _resolve_column(
        reward_df0, cfg.get("reward_code_col"), ["REWARD_CODE", "reward_code", "RewardCode"]
    )
    adc = str(cfg.get("reward_add_data_col") or "REWARD_ADD_DATA")
    if adc not in reward_df0.columns:
        adc = "REWARD_ADD_DATA"
    if rtc_res2 is None or rcc_res2 is None:
        catalog = {}
        logging.warning("[rating_item_matrix] Каталог ITEM из REWARD не построен: нет колонок типа/кода награды")
    else:
        catalog = build_item_catalog_from_reward_df(
            reward_df0,
            rtc_res2,
            rcc_res2,
            full_name_col="FULL_NAME",
            add_data_col=adc,
        )

    order_by_emp: Dict[str, Set[str]] = {}
    for emp_key, code_counts in counts_by_emp.items():
        order_by_emp[emp_key] = {c for c, n in code_counts.items() if n > 0}

    slr = cfg.get("sheet_list_rewards") or "LIST-REWARDS"
    rewards_by_emp: Dict[str, Set[str]] = defaultdict(set)
    if slr in sheets_data and sheets_data[slr] is not None:
        lr_df = sheets_data[slr][0]
        if isinstance(lr_df, pd.DataFrame):
            le = _resolve_column(lr_df, cfg.get("list_rewards_employee_col"), _DEFAULT_LIST_RW_EMP)
            lc = _resolve_column(lr_df, cfg.get("list_rewards_code_col"), _DEFAULT_LIST_RW_CODE)
            if le and lc:
                for _, lrow in lr_df.iterrows():
                    e2 = _norm_str(lrow.get(le))
                    c2 = _norm_str(lrow.get(lc))
                    if e2 and c2:
                        rewards_by_emp[e2].add(c2)
            else:
                logging.warning(
                    f"[rating_item_matrix] Лист «{slr}»: не найдены колонки табельного/кода награды — "
                    f"критерий rewardCode для всех товаров считается невыполненным, если список кодов не пуст."
                )
    else:
        logging.info(f"[rating_item_matrix] Лист «{slr}» отсутствует — множество наград сотрудника пустое.")

    cry_c = _resolve_column(rating_df, cfg.get("crystals_col"), _DEFAULT_CRYSTALS)
    role_c = _resolve_column(rating_df, cfg.get("rating_role_col"), ["Наименование Роли"])
    period_c = _resolve_column(rating_df, cfg.get("rating_period_col"), ["Период"])

    matrix_cells: List[Dict[str, Any]] = []
    header_stock_out_cols: Set[str] = set()

    for pos, (_, row) in enumerate(rating_df.iterrows()):
        emp_key = _norm_str(row.get(emp_r))
        rc = _as_float(row.get(country_c)) if country_c else None
        rt = _as_float(row.get(tb_c)) if tb_c else None
        rg = _as_float(row.get(gosb_c)) if gosb_c else None
        cry = _as_float(row.get(cry_c)) if cry_c else None
        count_by_code = counts_by_emp.get(emp_key, {})
        blocked_codes = _blocked_codes_for_row(count_by_code, order_groups)
        order_codes = order_by_emp.get(emp_key, set())
        rw_codes = rewards_by_emp.get(emp_key, set())
        excel_row = pos + 2

        for sp in specs:
            code = sp["match_code"]
            cname = sp["col_name"]
            count = int(count_by_code.get(code, 0))
            ordered = count > 0

            rules = rules_for_matrix_column(
                code,
                catalog,
                min_bank=sp.get("min_bank"),
                min_tb=sp.get("min_tb"),
                min_gosb=sp.get("min_gosb"),
            )
            base_accessible = item_accessible_for_manager(
                rules,
                rank_country=rc,
                rank_tb=rt,
                rank_gosb=rg,
                crystals=cry,
                order_product_codes=order_codes,
                list_reward_codes=rw_codes,
                manager_tab=emp_key or None,
            )
            accessible = base_accessible and code not in blocked_codes

            if ordered:
                cell_val: Any = count
            else:
                cell_val = "Y" if accessible else "N"
            col_values[cname].append(cell_val)

            fill_key = _matrix_fill_key(accessible, ordered)
            matrix_cells.append(
                {
                    "row_excel": excel_row,
                    "col_name": cname,
                    "accessible": accessible,
                    "ordered": ordered,
                    "count": count,
                    "fill_key": fill_key,
                }
            )

            limit_amt = _item_amount_limit(rules)
            if limit_amt is not None and count >= limit_amt:
                header_stock_out_cols.add(cname)

    for cname in added:
        rating_df[cname] = col_values.get(cname, [])

    sheets_data[sr] = (rating_df, rating_t[1])
    fills = _resolve_fill_colors(cfg)
    logging.info(
        f"[rating_item_matrix] Лист «{sr}»: колонок ITEM-матрицы: {len(added)}, "
        f"ячеек для подсветки: {len(matrix_cells)}, красных заголовков: {len(header_stock_out_cols)}"
    )
    if role_c is None or period_c is None:
        logging.debug(
            f"[rating_item_matrix] Колонки роль/период на RATING: role={role_c!r}, period={period_c!r} "
            "(для itemAmount используется только счётчик заказов менеджера по коду)"
        )

    return {
        "sheet_rating": sr,
        "added_columns": added,
        "thresholds": thresholds,
        "country_rank_col": country_c,
        "tb_rank_col": tb_c,
        "gosb_rank_col": gosb_c,
        "matrix_cells": matrix_cells,
        "fills": fills,
        "header_stock_out_columns": sorted(header_stock_out_cols),
        "fill_accessibility_ok": fills[FILL_ORDERED_AVAILABLE],
        "fill_accessibility_fail": fills[FILL_UNAVAILABLE_NOT_ORDERED],
        "skip_colors": False,
    }


def _header_col_map(ws: Any) -> Dict[str, int]:
    """Имя заголовка (строка 1) -> индекс столбца openpyxl (1-based)."""
    m: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is not None:
            m[str(v).strip()] = c
    return m


def apply_rating_item_matrix_colors(
    xlsx_path: str,
    meta: Dict[str, Any],
    cfg: Dict[str, Any],
) -> None:
    """Подсветка ячеек матрицы (4 состояния) и заголовков при исчерпании itemAmount."""
    if not cfg or not bool(cfg.get("enabled")):
        return
    if meta.get("skip_colors"):
        logging.info("[rating_item_matrix] Подсветка отключена флагом skip_colors")
        return
    cells = meta.get("matrix_cells") or meta.get("accessibility_cells") or []
    if not cells:
        logging.warning("[rating_item_matrix] Нет предвычисленных ячеек матрицы — подсветка пропущена")
        return
    try:
        wb = load_workbook(xlsx_path)
    except OSError as e:
        logging.warning(f"[rating_item_matrix] Не удалось открыть файл для подсветки: {e}")
        return

    sn = meta.get("sheet_rating") or "RATING"
    if sn not in wb.sheetnames:
        wb.close()
        return
    ws = wb[sn]
    hmap = _header_col_map(ws)

    def col_for(name: str) -> Optional[int]:
        return hmap.get(str(name).strip())

    fills_cfg = meta.get("fills") or _resolve_fill_colors(cfg or {})
    fill_by_key: Dict[str, PatternFill] = {}
    for key in (
        FILL_ORDERED_AVAILABLE,
        FILL_ORDERED_UNAVAILABLE,
        FILL_AVAILABLE_NOT_ORDERED,
        FILL_UNAVAILABLE_NOT_ORDERED,
    ):
        hx = _strip_hex(fills_cfg.get(key) or "FFFFFF")
        fill_by_key[key] = PatternFill(fill_type="solid", start_color=hx, end_color=hx)
    header_hex = _strip_hex(fills_cfg.get("header_stock_out") or "FF0000")
    fill_header = PatternFill(fill_type="solid", start_color=header_hex, end_color=header_hex)

    n_header = 0
    for cname in meta.get("header_stock_out_columns") or []:
        ci = col_for(str(cname))
        if ci is not None:
            ws.cell(row=1, column=ci).fill = fill_header
            n_header += 1

    n_applied = 0
    for item in cells:
        r = int(item.get("row_excel") or 0)
        cname = item.get("col_name")
        if r < 2 or not cname:
            continue
        ci = col_for(str(cname))
        if ci is None:
            continue
        fill_key = item.get("fill_key")
        if not fill_key and "ok" in item:
            # Совместимость со старым meta (только ok)
            fill_key = (
                FILL_ORDERED_AVAILABLE if item.get("ok") else FILL_UNAVAILABLE_NOT_ORDERED
            )
        pf = fill_by_key.get(str(fill_key)) or fill_by_key[FILL_UNAVAILABLE_NOT_ORDERED]
        ws.cell(row=r, column=ci).fill = pf
        n_applied += 1

    wb.save(xlsx_path)
    wb.close()
    logging.info(
        f"[rating_item_matrix] Подсветка ITEM: {n_applied} ячеек, {n_header} заголовков в {xlsx_path}"
    )
