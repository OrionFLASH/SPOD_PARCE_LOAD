# -*- coding: utf-8 -*-
"""
Полный каталог настраиваемых параметров PROM/SPOD → Excel.

Обходит все строки указанных CSV (не только первые), разворачивает JSON-колонки
в дерево ключей, собирает типы, зависимости, примеры, дубли и правила
консистентности из config/CONFIG_CHECKS.json и пояснения из глоссариев.

Запуск из корня проекта:
  python src/Tools/build_spod_params_excel.py
  python src/Tools/build_spod_params_excel.py --input-dir IN/PROM/SPOD --out Docs/params_catalog/SPOD_PARAMS_CATALOG_LEAF_v3.xlsx
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Set, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from src.csv_headers import normalize_csv_column_header  # noqa: E402

DEFAULT_INPUT = ROOT / "IN" / "SPOD_UPLOAD"
DEFAULT_OUT = ROOT / "Docs" / "params_catalog" / "SPOD_PARAMS_CATALOG_LEAF_v3.xlsx"
CHECKS_PATH = ROOT / "config" / "CONFIG_CHECKS.json"
FORMATS_PATH = ROOT / "config" / "CONFIG_FORMATS.json"
GLOSSARY_DIR = ROOT / "src" / "Tools" / "catalog_glossary"
DELIM = ";"
ENC = "utf-8"

# Префиксы развёрнутого JSON на родном листе (не «чужой лист=>колонка»).
_JSON_EXPAND_PREFIXES: Set[str] = {
    "ADD_DATA",
    "CONTEST_FEATURE",
    "REWARD_ADD_DATA",
    "FEATURE",
}

# Соответствие файла выгрузки → имя таблицы (листа) в проекте.
FILE_TO_TABLE: Dict[str, str] = {
    "CONTEST": "CONTEST-DATA",
    "REWARD-LINK": "REWARD-LINK",
    "REWARD": "REWARD",
    "GROUP": "GROUP",
    "INDICATOR": "INDICATOR",
    "SCHEDULE": "TOURNAMENT-SCHEDULE",
    "REPORT": "REPORT",
    "ORG_UNIT_V20": "ORG_UNIT_V20",
    "EMPLOYEE": "EMPLOYEE",
    "USER_ROLE_SB": "USER_ROLE SB",
    "USER_ROLE": "USER_ROLE",
}

# Колонки-дискриминаторы для зависимостей схемы JSON.
DEPENDENCY_DISCRIMINATORS: Dict[str, List[str]] = {
    "REWARD": ["REWARD_TYPE"],
    "CONTEST-DATA": ["CONTEST_TYPE", "BUSINESS_STATUS"],
    "GROUP": ["GROUP_CODE"],
    "INDICATOR": ["INDICATOR_MARK_TYPE", "INDICATOR_CALC_TYPE"],
    "TOURNAMENT-SCHEDULE": ["TOURNAMENT_STATUS", "PERIOD_TYPE"],
}

# Краткие назначения плоских колонок (по имени таблицы).
COLUMN_DESCRIPTIONS: Dict[str, Dict[str, str]] = {
    "CONTEST-DATA": {
        "CONTEST_CODE": "Уникальный код конкурса; ключ связей с GROUP, INDICATOR, REWARD-LINK, SCHEDULE, REPORT.",
        "FULL_NAME": (
            "Отображаемое название конкурса/турнира "
            "(на странице Турниры/Детальная карточка)."
        ),
        "CREATE_DT": "Дата начала YYYY-MM-DD. Часто начало года.",
        "CLOSE_DT": "Дата окончания; 4000-01-01 обычно означает «без срока».",
        "BUSINESS_STATUS": "Бизнес-статус конкурса: АКТИВНЫЙ | АРХИВНЫЙ.",
        "CONTEST_TYPE": (
            "ТУРНИРНЫЙ (до 3 BADGE) | ИНДИВИДУАЛЬНЫЙ | "
            "ИНДИВИДУАЛЬНЫЙ НАКОПИТЕЛЬНЫЙ (1 BADGE); влияет на схему CONTEST_FEATURE."
        ),
        "CONTEST_DESCRIPTION": (
            "Текст описания для конкурса/турнира "
            "(на странице Детальная карточка турнира)."
        ),
        "CONTEST_FEATURE": "JSON с параметрами UI, рассылок, фильтров ТБ/ГОСБ, награждения.",
        "SHOW_INDICATOR": (
            "Единица/подпись индикатора: шт. | Факт | % | … "
            "на списке показателей подпись к единицам данных."
        ),
        "PRODUCT_GROUP": (
            "Группа продукта из классификатора "
            "(можно будет в дальнейшем дополнить)."
        ),
        "PRODUCT": "Продукт / тематика конкурса (свободный текст).",
        "CONTEST_SUBJECT": "Предмет конкурса. Обычно: EMPLOYEE. Кто участник.",
        "FACTOR_MARK_TYPE": (
            "CRITERION | RATING_MAX | RATING_MIN — способ выбора победителей "
            "(достиг показателя / больше других / меньше других)."
        ),
        "CONTEST_INDICATOR_METHOD": "INTEGRAL | RELATION. Метод расчета показателя.",
        "CONTEST_FACTOR_METHOD": (
            "FACT | FACT0-FACT1 | FACT0-RUN_RATE1_DOWN | RUN_RATE — "
            "для автоматических турниров способ расчета на данных источников."
        ),
        "PLAN_METHOD_CODE": (
            "DEPENDS_PREVIOUS_PERIOD | PRESET_VALUE — метод расчета планового "
            "показателя (из прошлого периода / фиксированное значение)."
        ),
        "PLAN_MOD_METOD": "Модификатор плана. Обычно: MULTIPLIER.",
        "PLAN_MOD_VALUE": "Значение планового показателя (0, 1, 1000, …).",
        "FACTOR_MATCH": "Сравнение фактора: = | > | >=.",
        "CONTEST_PERIOD": "Периоды через ; или пусто → []. Обычно пусто.",
        "TARGET_TYPE": "Среда конкурса: ПРОМ | ТЕСТ.",
        "SOURCE_UPD_FREQUENCY": (
            "Частота обновления источника: 1 | 7 | 10 (дни). (не используется)"
        ),
        "CALC_TYPE": (
            "Тип расчёта: 0 | 1. (не используется) "
            "0 — промышленный / 1 — ручной."
        ),
        "BUSINESS_BLOCK": "Бизнес-блок(и) конкурса; часто JSON-массив кодов.",
        "FACT_POST_PROCESSING": (
            "Правило постобработки показателя конкурса; часто пусто. "
            "PERCENTILE — перцентиль от фактического показателя."
        ),
    },
    "REWARD": {
        "REWARD_CODE": "Уникальный код награды; связь с REWARD-LINK и отчётами.",
        "REWARD_TYPE": "Тип награды: ITEM / BADGE / LABEL / CRYSTAL — задаёт схему REWARD_ADD_DATA.",
        "FULL_NAME": "Краткое название бейджа / отображаемое название награды.",
        "REWARD_DESCRIPTION": "Полное описание награды.",
        "REWARD_CONDITION": "Класс/код условия начисления (часто пусто или код).",
        "REWARD_COST": "Стоимость в кристаллах (часто 0…14).",
        "REWARD_ADD_DATA": "JSON с UI-признаками, сезонами, условиями выдачи; структура зависит от REWARD_TYPE.",
    },
    "REWARD-LINK": {
        "CONTEST_CODE": "Код конкурса, к которому привязана награда.",
        "GROUP_CODE": "Код группы расчёта на конкурсе (BANK, TB, …).",
        "REWARD_CODE": "Код награды из справочника REWARD.",
    },
    "GROUP": {
        "CONTEST_CODE": "Код конкурса.",
        "GROUP_CODE": "Код группы расчёта (BANK, TB, GOSB, …).",
        "GROUP_VALUE": "Значение группы: `*`, код или JSON-массив кодов.",
        "GET_CALC_METHOD": "Метод получения расчёта.",
        "GET_CALC_CRITERION": "Критерий GET-расчёта.",
        "ADD_CALC_CRITERION": "Дополнительный критерий расчёта.",
        "ADD_CALC_CRITERION_2": "Второй дополнительный критерий.",
        "BASE_CALC_CODE": "Базовый код метода расчёта.",
    },
    "INDICATOR": {
        "CONTEST_CODE": "Код конкурса.",
        "INDICATOR_CALC_TYPE": "Тип расчёта индикатора.",
        "INDICATOR_ADD_CALC_TYPE": "Доп. тип расчёта (часть составного ключа).",
        "FULL_NAME": "Имя/метка индикатора.",
        "INDICATOR_CODE": "Код индикатора (WAIT, RATING, …).",
        "INDICATOR_AGG_FUNCTION": "Агрегирующая функция.",
        "INDICATOR_WEIGHT": "Вес индикатора в формуле.",
        "INDICATOR_OBJECT": "Объект применения индикатора.",
        "INDICATOR_MARK_TYPE": "Тип отметки (RATING и др.).",
        "INDICATOR_MATCH": "Условие совпадения (MIN, MAX, …).",
        "INDICATOR_VALUE": "Порог / константа индикатора.",
        "CONTEST_CRITERION": "Критерий конкурса для индикатора.",
        "INDICATOR_FILTER": "Фильтр отбора; может быть JSON.",
        "CONTESTANT_SELECTION": "Правило выбора участников.",
        "CALC_TYPE": "Тип расчёта (числовой код).",
        "N": "Порядковый/множительный параметр N.",
    },
    "TOURNAMENT-SCHEDULE": {
        "TOURNAMENT_CODE": "Уникальный код турнира/слота расписания.",
        "PERIOD_TYPE": "Тип/метка периода (текст).",
        "START_DT": "Дата старта турнира.",
        "END_DT": "Дата окончания турнира.",
        "RESULT_DT": "Дата подведения итогов турнира.",
        "PLAN_PERIOD_START_DT": "Плановое начало периода.",
        "PLAN_PERIOD_END_DT": "Плановое окончание периода.",
        "CRITERION_MARK_TYPE": "Тип отметки критерия.",
        "CRITERION_MARK_VALUE": "Значение отметки критерия.",
        "FILTER_PERIOD_ARR": "JSON-массив/объект фильтров периода.",
        "TOURNAMENT_STATUS": "Статус турнира (АКТИВНЫЙ, ЗАВЕРШЕН, УДАЛЕН, …).",
        "CONTEST_CODE": "Код конкурса.",
        "TARGET_TYPE": "Тип цели; часто JSON-объект/массив.",
        "CALC_TYPE": "Тип расчёта.",
        "TRN_INDICATOR_FILTER": "Фильтр индикатора на уровне турнира.",
    },
    "REPORT": {
        "MANAGER_PERSON_NUMBER": "Табельный номер участника/менеджера (20 знаков).",
        "CONTEST_CODE": "Код конкурса.",
        "TOURNAMENT_CODE": "Код турнира из расписания.",
        "CONTEST_DATE": "Дата среза показателя.",
        "PLAN_VALUE": "Плановое значение (число, 5 знаков после точки).",
        "FACT_VALUE": "Фактическое значение (число, 5 знаков после точки).",
        "priority_type": "Тип приоритета строки отчёта.",
    },
    "ORG_UNIT_V20": {
        "TB_CODE": "Код территориального банка.",
        "TB_FULL_NAME": "Полное название ТБ.",
        "TB_SHORT_NAME": "Краткое название ТБ.",
        "GOSB_CODE": "Код ГОСБ (0 — аппарат ТБ).",
        "GOSB_NAME": "Полное название ГОСБ.",
        "GOSB_SHORT_NAME": "Краткое название ГОСБ.",
        "CLUSTER_CODE": "Код кластера.",
        "GROUPING_CODE": "Код группировки в иерархии.",
        "GOSB_CNT": "Счётчик ГОСБ.",
        "GROUPING_CNT": "Счётчик группировки.",
        "ORG_UNIT_CODE": "Уникальный код оргподразделения (ключ).",
    },
    "EMPLOYEE": {
        "PERSON_NUMBER": "Табельный номер (20 цифр с ведущими нулями).",
        "PERSON_NUMBER_ADD": "Доп./нормализованный табельный номер (тоже 20 цифр).",
        "SURNAME": "Фамилия.",
        "FIRST_NAME": "Имя.",
        "MIDDLE_NAME": "Отчество.",
        "MANAGER_FULL_NAME": "ФИО руководителя строкой.",
        "POSITION_NAME": "Наименование должности.",
        "TB_CODE": "Код территориального банка.",
        "GOSB_CODE": "Код ГОСБ.",
        "BUSINESS_BLOCK": "Код бизнес-блока сотрудника.",
        "PRIORITY_TYPE": "Тип приоритета.",
        "KPK_CODE": "Код КПК (если применимо).",
        "KPK_NAME": "Наименование КПК.",
        "ROLE_CODE": "Код роли в промо.",
        "UCH_CODE": "Код участка/учёта.",
        "GENDER": "Пол (код).",
        "ORG_UNIT_CODE": "Код оргподразделения (связь с ORG_UNIT_V20).",
    },
    "USER_ROLE": {
        "RULE_NUM": "Номер правила роли (уникальный ключ).",
        "ROLE_CODE": "Код роли.",
        "ROLE_NAME": "Наименование роли.",
        "PERSON_NUMBER_ARR": "JSON-массив табельных номеров (явный список сотрудников).",
        "STAGE_ETALONE_CODE_ARR": "JSON-массив кодов эталонов стадий.",
        "POST_ETALONE_CODE_ARR": "JSON-массив кодов эталонов должностей.",
        "DIV_CODE_ARR": "JSON-массив кодов подразделений включения.",
        "EXCLUDE_DIV_CODE_ARR": "JSON-массив кодов подразделений исключения.",
        "BUSINESS_BLOCK": "Бизнес-блок правила.",
        "UCH_CODE": "Код участка.",
        "ORG_UNIT_CODE": "Код оргподразделения.",
        "TB_CODE": "Код ТБ (фильтр).",
        "GOSB_CODE": "Код ГОСБ (фильтр).",
    },
    "USER_ROLE SB": {
        "RULE_NUM": "Номер правила роли (уникальный ключ) — вариант SB.",
        "ROLE_CODE": "Код роли.",
        "ROLE_NAME": "Наименование роли.",
        "PERSON_NUMBER_ARR": "JSON-массив табельных номеров.",
        "STAGE_ETALONE_CODE_ARR": "JSON-массив кодов эталонов стадий.",
        "POST_ETALONE_CODE_ARR": "JSON-массив кодов эталонов должностей.",
        "DIV_CODE_ARR": "JSON-массив кодов подразделений включения.",
        "EXCLUDE_DIV_CODE_ARR": "JSON-массив кодов подразделений исключения.",
        "BUSINESS_BLOCK": "Бизнес-блок правила.",
        "UCH_CODE": "Код участка.",
        "ORG_UNIT_CODE": "Код оргподразделения.",
        "TB_CODE": "Код ТБ (фильтр).",
        "GOSB_CODE": "Код ГОСБ (фильтр).",
    },
}

EXCEL_HEADERS: List[str] = [
    "Название таблицы",
    "Вид параметра",
    "Колонка CSV",
    "Тип ячейки (JSON / -)",
    "Полный путь ключа",
    "Имя параметра",
    "Тип данных",
    "Зависимости",
    "Описание параметра",
    "Идентификатор (EN)",
    "Признак дублей (предположение)",
    "Пример значения 1",
    "Пример значения 2",
    "Пример значения 3",
    "Условия консистентности",
    "Строк в таблице",
    "Заполнено / ключ есть",
    "Пусто / ключа нет",
    "% заполнения",
    "Уникальных значений",
    "Разнообразие (уник./заполн.)",
    "Статистика встречаемости",
    "Источник файла",
    "Excel: тип",
    "Excel: ограничения",
    "Excel: выравнивание",
    "Excel: цвет заголовка",
    "Excel: ширина",
]


def normalize_json_cell(raw: str) -> str:
    """Нормализация SPOD-JSON: тройные кавычки и лишняя кавычка после ] или }."""
    s = raw.replace('"""', '"').strip()
    # типичный артефакт CSV: ...}]"  /  ...}"
    if len(s) >= 2 and s[-1] == '"' and s[-2] in "]}":
        s = s[:-1]
    return s


def try_parse_json(raw: str) -> Tuple[Optional[Any], bool]:
    """Возвращает (объект, удалось_ли). Пустая строка → (None, False)."""
    s = (raw or "").strip()
    if not s:
        return None, False
    nv = normalize_json_cell(s)
    if not (nv.startswith("{") or nv.startswith("[")):
        return None, False
    try:
        return json.loads(nv), True
    except (json.JSONDecodeError, TypeError, ValueError):
        pass
    # доп. эвристики (как в json_utils.safe_json_loads)
    try:
        fixed = nv
        fixed = re.sub(r'"{2,}([^"\s]+)"{2,}\s*:', r'"\1":', fixed)
        fixed = re.sub(r':\s*"{2,}([^"\s]+)"{2,}', r':"\1"', fixed)
        fixed = re.sub(r',\s*([}\]])', r'\1', fixed)
        if len(fixed) >= 2 and fixed[-1] == '"' and fixed[-2] in "]}":
            fixed = fixed[:-1]
        return json.loads(fixed), True
    except (json.JSONDecodeError, TypeError, ValueError):
        return None, False


def type_label_scalar(v: Any) -> str:
    if v is None:
        return "null"
    if isinstance(v, bool):
        return "boolean"
    if isinstance(v, int) and not isinstance(v, bool):
        return "целое число"
    if isinstance(v, float):
        return "число"
    if isinstance(v, str):
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", v):
            return "дата (строка YYYY-MM-DD)"
        if v in ("Y", "N"):
            return "строка (флаг Y/N)"
        if re.fullmatch(r"-?\d+", v):
            return "строка (число)"
        if re.fullmatch(r"-?\d+\.\d+", v):
            return "строка (десятичное число)"
        return "строка"
    if isinstance(v, list):
        return "массив"
    if isinstance(v, dict):
        return "объект"
    return type(v).__name__


def infer_array_element_type(samples: Iterable[Any]) -> str:
    labels: Counter[str] = Counter()
    for v in samples:
        labels[type_label_scalar(v)] += 1
    if not labels:
        return "массив (пустой / без элементов в выборке)"
    top = labels.most_common(1)[0][0]
    if "дата" in top:
        return "массив с датами"
    if "целое" in top or top == "число":
        return "массив с числами"
    if "флаг" in top:
        return "массив со строками (флаги Y/N)"
    if top.startswith("строка"):
        return "массив со строками"
    if top == "объект":
        return "массив объектов"
    if top == "массив":
        return "массив массивов"
    return f"массив ({top})"


def infer_flat_column_type(values: List[str]) -> str:
    non_empty = [v for v in values if v is not None and str(v).strip() != ""]
    if not non_empty:
        return "пусто"
    # JSON?
    json_ok = 0
    for v in non_empty[: min(200, len(non_empty))]:
        _, ok = try_parse_json(v)
        if ok:
            json_ok += 1
    if json_ok >= max(1, int(0.5 * min(200, len(non_empty)))):
        # уточним корень
        sample_obj, _ = try_parse_json(non_empty[0])
        if isinstance(sample_obj, list):
            return "JSON-массив (в ячейке)"
        return "JSON-объект (в ячейке)"
    date_n = sum(1 for v in non_empty if re.fullmatch(r"\d{4}-\d{2}-\d{2}", v.strip()))
    if date_n >= max(1, int(0.8 * len(non_empty))):
        return "дата (строка YYYY-MM-DD)"
    int_n = sum(1 for v in non_empty if re.fullmatch(r"-?\d+", v.strip()))
    if int_n >= max(1, int(0.9 * len(non_empty))):
        return "целое число (строкой в CSV)"
    float_n = sum(1 for v in non_empty if re.fullmatch(r"-?\d+\.\d+", v.strip()))
    if float_n >= max(1, int(0.9 * len(non_empty))):
        return "число с дробной частью (строкой в CSV)"
    yn = sum(1 for v in non_empty if v.strip() in ("Y", "N"))
    if yn >= max(1, int(0.9 * len(non_empty))):
        return "строка (флаг Y/N)"
    return "строка"


def merge_path(base: str, key: str) -> str:
    if not base:
        return key
    return f"{base}.{key}"


def leaf_name(path: str) -> str:
    if not path:
        return ""
    p = path.replace("[]", "")
    return p.split(".")[-1] if "." in p else p


def make_param_id(table: str, column: str, json_path: str, is_json_col: bool) -> str:
    """Короткий уникальный EN-идентификатор параметра."""
    t = re.sub(r"[^A-Za-z0-9]+", "_", table).strip("_").upper()
    c = re.sub(r"[^A-Za-z0-9]+", "_", column).strip("_")
    if not is_json_col or json_path in ("-", ""):
        return f"{t}__COL__{c}"
    # путь без имени колонки-префикса если уже есть
    jp = json_path
    if jp.startswith(column + "."):
        jp = jp[len(column) + 1 :]
    elif jp == column:
        jp = "root"
    jp = jp.replace("[]", "_ITEM").replace(".", "__")
    jp = re.sub(r"[^A-Za-z0-9_]+", "_", jp).strip("_")
    return f"{t}__JSON__{c}__{jp or 'root'}"


def parse_glossary_meanings(md_text: str) -> Dict[str, str]:
    """Извлекает из markdown-глоссария блоки ##### `key` → текст смысла."""
    result: Dict[str, str] = {}
    parts = re.split(r"\n##### `([^`]+)`\n", md_text)
    for i in range(1, len(parts), 2):
        key = parts[i].strip()
        body = parts[i + 1] if i + 1 < len(parts) else ""
        body = re.split(r"\n#### ", body, maxsplit=1)[0]
        lines: List[str] = []
        for line in body.splitlines():
            s = line.strip()
            if not s.startswith("- "):
                continue
            content = s[2:].strip()
            # приоритет смысловых строк
            if content.startswith("**Смысл") or "гипотеза" in content.lower():
                lines.insert(0, content)
            elif content.startswith("**Тип") or content.startswith("**REWARD") or content.startswith("**CONTEST") or content.startswith("**Домен") or content.startswith("**Значен"):
                lines.append(content)
            elif len(lines) < 4:
                lines.append(content)
        text = " ".join(lines).strip()
        if text:
            result[key] = text

    # таблица раздела getCondition (и подобные): | `path` | тип | описание |
    in_detail_table = False
    for line in md_text.splitlines():
        if "Блок `getCondition`" in line or line.strip().startswith("#### 5."):
            in_detail_table = True
        if in_detail_table and line.startswith("#### ") and "getCondition" not in line and not line.strip().startswith("#### 5"):
            in_detail_table = False
        if not in_detail_table:
            continue
        m = re.match(r"\|\s*`([^`]+)`\s*\|\s*([^|]+)\|\s*([^|]+)\|", line)
        if not m:
            continue
        key = m.group(1).strip()
        typ = m.group(2).strip()
        desc = m.group(3).strip()
        if key in ("Подполе", "---") or typ.startswith("-"):
            continue
        if key and desc and key not in result:
            result[key] = f"Тип: {typ}. {desc}"
        # также leaf
        leaf = key.split(".")[-1]
        if leaf and leaf not in result and desc:
            result[leaf] = f"Тип: {typ}. {desc}"
    return result


@dataclass
class PathAccumulator:
    """Статистика по одному параметру (колонка или JSON-путь)."""

    table: str
    column: str
    json_path: str  # "-" для плоских; для JSON — полный путь от корня
    is_json_column: bool
    types: Counter = field(default_factory=Counter)
    examples: List[str] = field(default_factory=list)
    example_set: Set[str] = field(default_factory=set)
    present_rows: int = 0
    # dependency_key -> Counter of values where this path appeared
    dep_presence: Dict[str, Counter] = field(default_factory=lambda: defaultdict(Counter))
    array_elem_types: Counter = field(default_factory=Counter)
    scalar_for_array: List[Any] = field(default_factory=list)
    value_counter: Counter = field(default_factory=Counter)
    empty_string_count: int = 0
    null_count: int = 0
    empty_array_count: int = 0
    touched_rows: Set[int] = field(default_factory=set)

    def _value_key(self, v: Any) -> str:
        if v is None:
            return "null"
        if isinstance(v, (dict, list)):
            try:
                s = json.dumps(v, ensure_ascii=False, separators=(",", ":"), sort_keys=True)
            except TypeError:
                s = str(v)
        else:
            s = str(v)
        if len(s) > 200:
            s = s[:197] + "…"
        return s

    def add_example(self, v: Any) -> None:
        if v is None:
            self.null_count += 1
            s = "null"
        elif isinstance(v, str) and v.strip() == "":
            self.empty_string_count += 1
            s = '""'
        elif isinstance(v, list) and len(v) == 0:
            self.empty_array_count += 1
            s = "[]"
        elif isinstance(v, (dict, list)):
            try:
                s = json.dumps(v, ensure_ascii=False, separators=(",", ":"))
            except TypeError:
                s = str(v)
        else:
            s = str(v)
        vk = self._value_key(v)
        if len(self.value_counter) < 5000 or vk in self.value_counter:
            self.value_counter[vk] += 1
        if len(s) > 300:
            s = s[:297] + "…"
        if s not in self.example_set and len(self.examples) < 3:
            self.example_set.add(s)
            self.examples.append(s)
        elif len(self.examples) < 3:
            self.examples.append(s)

    def note_type(self, label: str) -> None:
        self.types[label] += 1


def leaf_key_name(full_path: str) -> str:
    """Последний сегмент пути: getCondition.rewards[].rewardCode → rewardCode."""
    p = (full_path or "").strip()
    if not p or p == "[]":
        return ""
    # убрать хвостовой [] у массива скаляров: feature[] → feature
    if p.endswith("[]") and p.count("[]") == 1 and "." not in p[:-2]:
        return p[:-2]
    core = p[:-2] if p.endswith("[]") and not p.endswith("[].") else p
    # путь вида a.b[].c или [].c
    seg = core.split(".")[-1]
    seg = seg.replace("[]", "")
    return seg or ""


def display_full_path(path_from_root: str) -> str:
    """Полный путь ключа1.ключ2.…ключ (с [] для контекста массива объектов)."""
    p = (path_from_root or "").strip()
    return p if p else "-"


def _norm_fmt_header(name: Optional[str]) -> str:
    """Нормализация имени колонки для сопоставления с CONFIG_FORMATS."""
    return normalize_csv_column_header(name)


def is_native_format_column_entry(col_entry: str, table: str) -> bool:
    """
    True, если запись columns в FORMATS относится к родному листу/развороту JSON,
    а не к чужой merge-колонке (REPORT=>…, LIST-TOURNAMENT=>…).
    """
    s = (col_entry or "").strip()
    if not s:
        return False
    if "=>" not in s:
        return True
    # «A => B» с пробелами — разворот JSON; «A=>B» без пробелов вокруг => — часто merge
    left = s.split("=>", 1)[0].strip()
    if left in _JSON_EXPAND_PREFIXES:
        return True
    # чужой лист: REPORT, LIST-TOURNAMENT, REWARD, CONTEST-DATA, …
    if left.upper() == table.upper():
        return True
    # если слева похоже на имя листа (есть дефис/подчёркивание типичных листов) — не native
    known_foreign = {
        "REPORT",
        "LIST-TOURNAMENT",
        "LIST-REWARDS",
        "SUMMARY",
        "REWARD",
        "CONTEST-DATA",
        "GROUP",
        "INDICATOR",
        "TOURNAMENT-SCHEDULE",
        "REWARD-LINK",
        "ORG_UNIT_V20",
        "EMPLOYEE",
        "USER_ROLE",
        "USER_ROLE SB",
        "YEAR_STATA",
    }
    if left in known_foreign and left.upper() != table.upper():
        return False
    # иначе считаем разворотом JSON на листе (напр. ADD_DATA уже покрыт)
    return True


def json_path_to_excel_segments(json_path: str) -> str:
    """Путь каталога `a.b[].c` → сегменты Excel `a => b[] => c`."""
    p = (json_path or "").strip()
    if not p or p == "-":
        return ""
    parts = p.split(".")
    return " => ".join(parts)


def format_header_candidates(
    table: str,
    csv_column: str,
    json_path: str,
    is_json_key: bool,
) -> List[str]:
    """Кандидаты имён заголовка Excel для матчинга с CONFIG_FORMATS на родном листе."""
    out: List[str] = []
    col = (csv_column or "").strip()
    if not is_json_key:
        if col:
            out.append(col)
        return out
    path_seg = json_path_to_excel_segments(json_path if json_path != "-" else "")
    if not path_seg:
        return out
    prefixes = [col] if col else []
    if col == "REWARD_ADD_DATA":
        prefixes.append("ADD_DATA")
    elif col == "CONTEST_FEATURE":
        prefixes.append("CONTEST_FEATURE")
    seen: Set[str] = set()
    for pref in prefixes:
        cand = f"{pref} => {path_seg}"
        n = _norm_fmt_header(cand)
        if n and n not in seen:
            seen.add(n)
            out.append(cand)
    return out


def format_constraints_text(rule: Dict[str, Any]) -> str:
    """Человекочитаемые ограничения из правила column_formats."""
    dtype = str(rule.get("data_type") or "general").lower()
    if dtype == "number":
        places = int(rule.get("decimal_places", 0))
        dec = rule.get("decimal_separator", ",")
        thou = "да" if rule.get("thousands_separator", False) else "нет"
        return f"знаки={places}; дес={dec}; тысячи={thou}"
    if dtype == "date":
        return f"шаблон={rule.get('date_format') or 'YYYY-MM-DD'}"
    if dtype == "text":
        return "text"
    return "-"


def format_alignment_text(rule: Dict[str, Any]) -> str:
    h = str(rule.get("horizontal", "left")).lower()
    v = str(rule.get("vertical", "center")).lower()
    wrap = "да" if rule.get("wrap_text", False) else "нет"
    return f"гориз={h}; верт={v}; перенос={wrap}"


def format_width_from_rule(rule: Dict[str, Any]) -> str:
    """
    Явная ширина колонки из полей правила (не листовой дефолт).
    Поддерживает width / width_mode / min_width / max_width / col_width и т.п.
    """
    parts: List[str] = []
    mode = rule.get("width_mode", rule.get("width", rule.get("col_width_mode")))
    if mode is not None and str(mode).strip() != "":
        parts.append(f"mode={mode}")
    mn = rule.get("min_width", rule.get("col_min_width"))
    mx = rule.get("max_width", rule.get("col_max_width"))
    if mn is not None:
        parts.append(f"min={mn}")
    if mx is not None:
        parts.append(f"max={mx}")
    return "; ".join(parts) if parts else "-"


def load_formats_config(path: Path = FORMATS_PATH) -> Dict[str, Any]:
    """Читает CONFIG_FORMATS.json."""
    if not path.is_file():
        return {"color_scheme": [], "column_formats": []}
    return json.loads(path.read_text(encoding="utf-8"))


def find_column_format_rule(
    formats_cfg: Dict[str, Any],
    table: str,
    candidates: List[str],
) -> Optional[Dict[str, Any]]:
    """Первое правило column_formats на родном листе, совпавшее с кандидатом."""
    cand_norm = {_norm_fmt_header(c) for c in candidates if c}
    if not cand_norm:
        return None
    for rule in formats_cfg.get("column_formats") or []:
        if not isinstance(rule, dict):
            continue
        if str(rule.get("sheet") or "") != table:
            continue
        cols = rule.get("columns") or []
        for entry in cols:
            if not is_native_format_column_entry(str(entry), table):
                continue
            if _norm_fmt_header(str(entry)) in cand_norm:
                return rule
        # column_prefixes
        for prefix in rule.get("column_prefixes") or []:
            if not is_native_format_column_entry(str(prefix), table):
                continue
            pnorm = _norm_fmt_header(str(prefix))
            if pnorm and any(c.startswith(pnorm) for c in cand_norm):
                return rule
    return None


def find_color_scheme_rule(
    formats_cfg: Dict[str, Any],
    table: str,
    candidates: List[str],
) -> Optional[Dict[str, Any]]:
    """Наиболее специфичное правило color_scheme для заголовка на родном листе."""
    cand_norm = {_norm_fmt_header(c) for c in candidates if c}
    schemes = [
        g
        for g in (formats_cfg.get("color_scheme") or [])
        if isinstance(g, dict) and table in (g.get("sheets") or [])
    ]
    # сначала с явным списком columns
    for g in schemes:
        cols = g.get("columns") or []
        if not cols:
            continue
        for entry in cols:
            if not is_native_format_column_entry(str(entry), table):
                continue
            if _norm_fmt_header(str(entry)) in cand_norm:
                return g
    # затем группа с пустым columns (фон «Исходные данные»)
    for g in schemes:
        if not (g.get("columns") or []):
            return g
    return None


def format_color_text(color_rule: Optional[Dict[str, Any]]) -> str:
    if not color_rule:
        return "-"
    bg = color_rule.get("header_bg") or "-"
    fg = color_rule.get("header_fg") or "-"
    group = color_rule.get("group") or "-"
    return f"bg={bg}; fg={fg}; группа={group}"


def excel_format_fields_for_row(
    formats_cfg: Dict[str, Any],
    row: Dict[str, Any],
) -> Dict[str, str]:
    """Заполняет пять полей Excel:* для строки каталога."""
    table = str(row.get("table") or "")
    csv_col = str(row.get("csv_column") or "")
    jp = str(row.get("json_path") or "-")
    is_json = str(row.get("param_kind") or "") == "JSON-КЛЮЧ"
    candidates = format_header_candidates(table, csv_col, jp, is_json)
    # для цвета у КОЛОНКИ JSON-ячейки также пробуем имя колонки
    if not is_json and candidates:
        pass
    fmt_rule = find_column_format_rule(formats_cfg, table, candidates)
    color_rule = find_color_scheme_rule(formats_cfg, table, candidates)

    if fmt_rule:
        dtype = str(fmt_rule.get("data_type") or "general").lower()
        excel_type = dtype if dtype else "general"
        excel_limits = format_constraints_text(fmt_rule)
        excel_align = format_alignment_text(fmt_rule)
        excel_width = format_width_from_rule(fmt_rule)
    else:
        excel_type = "-"
        excel_limits = "-"
        excel_align = "-"
        excel_width = "-"

    return {
        "excel_type": excel_type,
        "excel_limits": excel_limits,
        "excel_align": excel_align,
        "excel_color": format_color_text(color_rule),
        "excel_width": excel_width,
    }


def apply_excel_format_fields(
    rows: List[Dict[str, Any]],
    formats_cfg: Optional[Dict[str, Any]] = None,
) -> None:
    """Проставляет Excel:* поля во все строки каталога (in-place)."""
    cfg = formats_cfg if formats_cfg is not None else load_formats_config()
    for r in rows:
        r.update(excel_format_fields_for_row(cfg, r))


def is_scalar_or_null(v: Any) -> bool:
    return not isinstance(v, (dict, list))


def is_array_of_scalars(v: Any) -> bool:
    if not isinstance(v, list):
        return False
    if not v:
        return True
    return all(is_scalar_or_null(x) for x in v)


def ensure_leaf_acc(
    acc: Dict[str, PathAccumulator],
    path: str,
    column: str,
    table: str,
) -> PathAccumulator:
    if path not in acc:
        acc[path] = PathAccumulator(
            table=table,
            column=column,
            json_path=display_full_path(path),
            is_json_column=True,
        )
    return acc[path]


def register_leaf(
    acc: Dict[str, PathAccumulator],
    path: str,
    value: Any,
    column: str,
    table: str,
    dep_ctx: Dict[str, str],
    row_idx: int,
) -> None:
    """Регистрирует конечный параметр (скаляр или массив скаляров)."""
    node = ensure_leaf_acc(acc, path, column, table)
    if row_idx not in node.touched_rows:
        node.touched_rows.add(row_idx)
        node.present_rows += 1
        for dk, dv in dep_ctx.items():
            if dv:
                node.dep_presence[dk][dv] += 1
    if isinstance(value, list):
        node.note_type("массив")
        node.add_example(value if len(value) <= 5 else list(value[:5]) + ["…"])
        for item in value:
            lab = type_label_scalar(item)
            node.array_elem_types[lab] += 1
            if len(node.scalar_for_array) < 50:
                node.scalar_for_array.append(item)
            vk = node._value_key(item)
            if len(node.value_counter) < 5000 or vk in node.value_counter:
                node.value_counter[vk] += 1
    else:
        node.note_type(type_label_scalar(value))
        node.add_example(value)


def walk_json_leaves(
    obj: Any,
    path: str,
    column: str,
    table: str,
    acc: Dict[str, PathAccumulator],
    dep_ctx: Dict[str, str],
    row_idx: int,
) -> None:
    """
    Обход только конечных параметров (как в плоских каталогах полей / BigQuery field path).

    - объект-контейнер не регистрируется, только спуск внутрь;
    - массив объектов не регистрируется, регистрируются поля элементов (путь с []);
    - массив скаляров / скаляр — конечный параметр;
    - корневой массив скаляров не даёт отдельных ключей (это сама колонка).
    """
    if isinstance(obj, dict):
        for k, v in obj.items():
            child = f"{path}.{k}" if path else k
            if isinstance(v, dict):
                walk_json_leaves(v, child, column, table, acc, dep_ctx, row_idx)
            elif isinstance(v, list):
                if is_array_of_scalars(v):
                    register_leaf(acc, child, v, column, table, dep_ctx, row_idx)
                elif all(isinstance(x, dict) for x in v):
                    for item in v:
                        walk_json_leaves(
                            item, f"{child}[]", column, table, acc, dep_ctx, row_idx
                        )
                else:
                    register_leaf(acc, child, v, column, table, dep_ctx, row_idx)
            else:
                register_leaf(acc, child, v, column, table, dep_ctx, row_idx)
        return

    if isinstance(obj, list):
        if not path:
            if is_array_of_scalars(obj):
                return
            if all(isinstance(x, dict) for x in obj):
                for item in obj:
                    walk_json_leaves(item, "[]", column, table, acc, dep_ctx, row_idx)
            return
        if is_array_of_scalars(obj):
            register_leaf(acc, path, obj, column, table, dep_ctx, row_idx)
        return

    if path:
        register_leaf(acc, path, obj, column, table, dep_ctx, row_idx)


def resolve_table_name(filename: str) -> Optional[str]:
    """Определяет имя таблицы по префиксу имени файла выгрузки."""
    stem = Path(filename).name
    head = stem.split("__")[0].split(" ")[0].strip()
    head_u = head.upper().replace(" ", "_")
    # порядок: более длинные/специфичные префиксы первыми
    ordered = [
        ("REWARD-LINK", "REWARD-LINK"),
        ("REWARD_LINK", "REWARD-LINK"),
        ("USER_ROLE_SB", "USER_ROLE SB"),
        ("USER-ROLE-SB", "USER_ROLE SB"),
        ("ORG_UNIT_V20", "ORG_UNIT_V20"),
        ("TOURNAMENT-SCHEDULE", "TOURNAMENT-SCHEDULE"),
        ("SCHEDULE", "TOURNAMENT-SCHEDULE"),
        ("CONTEST", "CONTEST-DATA"),
        ("REWARD", "REWARD"),
        ("GROUP", "GROUP"),
        ("INDICATOR", "INDICATOR"),
        ("REPORT", "REPORT"),
        ("EMPLOYEE", "EMPLOYEE"),
        ("USER_ROLE", "USER_ROLE"),
    ]
    for prefix, table in ordered:
        if head_u == prefix.upper() or head_u.startswith(prefix.upper() + "_"):
            return table
    return FILE_TO_TABLE.get(head_u)


def detect_json_columns(rows: List[Dict[str, str]], columns: List[str]) -> Set[str]:
    """Колонка считается JSON, если ≥30% непустых ячеек парсятся как JSON (или ≥10 удачных)."""
    json_cols: Set[str] = set()
    for col in columns:
        ok = 0
        nonempty = 0
        for row in rows:
            v = (row.get(col) or "").strip()
            if not v:
                continue
            nonempty += 1
            _, parsed = try_parse_json(v)
            if parsed:
                ok += 1
        if nonempty == 0:
            continue
        if ok >= 10 or (ok / nonempty) >= 0.3:
            json_cols.add(col)
    return json_cols


def format_dependencies(
    node: PathAccumulator,
    disc_totals: Dict[str, Counter],
    is_json: bool,
) -> str:
    """Каждая зависимость — с новой строки внутри ячейки."""
    lines: List[str] = []
    if not is_json:
        return "-"
    for disc, counter in sorted(node.dep_presence.items()):
        totals = disc_totals.get(disc, Counter())
        present_vals = sorted(counter.keys())
        all_vals = sorted(totals.keys())
        missing = [v for v in all_vals if v not in counter]
        lines.append(f"Дискриминатор: {disc}")
        for v in present_vals:
            tot = totals.get(v, 0)
            cnt = counter[v]
            if tot and cnt / tot >= 0.95:
                lines.append(f"  есть при {disc}={v}: {cnt}/{tot} (почти всегда)")
            elif tot:
                lines.append(f"  есть при {disc}={v}: {cnt}/{tot} строк")
            else:
                lines.append(f"  есть при {disc}={v}: {cnt}")
        for v in missing:
            tot = totals.get(v, 0)
            lines.append(f"  нет при {disc}={v} (таких строк дискриминатора: {tot})")
    return "\n".join(lines) if lines else "-"


def diversity_ratio(unique_n: int, filled_n: int) -> str:
    if filled_n <= 0:
        return "-"
    return f"{unique_n / filled_n:.2f}"


def format_occurrence_stats(
    *,
    total: int,
    filled: int,
    empty: int,
    unique_n: int,
    value_counter: Optional[Counter] = None,
    empty_string: int = 0,
    null_n: int = 0,
    empty_array: int = 0,
    json_ok: Optional[int] = None,
    key_absent_in_json: Optional[int] = None,
    is_json_key: bool = False,
) -> str:
    """Текстовая сводка встречаемости — каждая метрика с новой строки."""
    lines: List[str] = [
        f"строк таблицы: {total}",
        f"заполнено / ключ встречается: {filled}",
        f"пусто / ключа нет: {empty}",
    ]
    pct = (100.0 * filled / total) if total else 0.0
    lines.append(f"% заполнения: {pct:.1f}%")
    lines.append(f"уникальных значений: {unique_n}")
    if filled:
        lines.append(f"разнообразие (уник./заполн.): {unique_n / filled:.2f}")
    if is_json_key and json_ok is not None:
        lines.append(f"JSON колонки распарсен: {json_ok}/{total}")
    if key_absent_in_json is not None:
        lines.append(f"в распарсенном JSON ключ отсутствует: {key_absent_in_json}")
    if empty_string:
        lines.append(f"пустая строка \"\": {empty_string}")
    if null_n:
        lines.append(f"null: {null_n}")
    if empty_array:
        lines.append(f"пустой массив []: {empty_array}")
    if value_counter:
        top = value_counter.most_common(5)
        if top:
            lines.append("топ значений:")
            for val, cnt in top:
                show = val if len(val) <= 80 else val[:77] + "…"
                lines.append(f"  {cnt}× {show}")
    return "\n".join(lines)


def dominant_type(node: PathAccumulator) -> str:
    if not node.types:
        return "неизвестно"
    # если массив — уточнить элементы
    if "массив" in node.types and node.array_elem_types:
        return infer_array_element_type(
            [k for k, n in node.array_elem_types.items() for _ in range(min(n, 5))]
        )
    if node.scalar_for_array:
        return infer_array_element_type(node.scalar_for_array)
    # смесь типов
    items = node.types.most_common()
    if len(items) == 1:
        return items[0][0]
    top = ", ".join(f"{k}×{v}" for k, v in items[:4])
    return f"смесь: {top}"


def load_consistency_index(path: Path) -> Dict[Tuple[str, str], List[str]]:
    """Индекс (table, column_or_json_key) → тексты условий из CONFIG_CHECKS."""
    idx: Dict[Tuple[str, str], List[str]] = defaultdict(list)
    if not path.is_file():
        return idx
    data = json.loads(path.read_text(encoding="utf-8"))
    rules = data.get("consistency_checks", {}).get("rules", [])

    def add(table: str, field_name: str, text: str) -> None:
        if not table or not field_name:
            return
        idx[(table, field_name)].append(text)

    for r in rules:
        rid = r.get("id", "")
        name = r.get("name") or rid
        rtype = r.get("type", "")
        enabled = r.get("enabled", True)
        en = "вкл" if enabled else "выкл"
        base = f"[{rid}] ({rtype}, {en}) {name}"

        sheet = r.get("sheet") or r.get("sheet_src") or ""
        json_col = r.get("json_column") or ""

        if rtype == "referential":
            src_s = r.get("sheet_src", "")
            src_c = r.get("column_src", "")
            ref_s = r.get("sheet_ref", "")
            ref_c = r.get("column_ref", "")
            add(src_s, src_c, f"{base}: значения {src_s}.{src_c} ⊆ {ref_s}.{ref_c}")
            add(ref_s, ref_c, f"{base}: справочник для {src_s}.{src_c}")
        elif rtype == "referential_composite":
            src_s = r.get("sheet_src", "")
            ref_s = r.get("sheet_ref", "")
            src_fields = list(r.get("columns_src") or r.get("src_fields") or [])
            ref_fields = list(r.get("columns_ref") or r.get("ref_fields") or [])
            for c in src_fields:
                add(
                    src_s,
                    c,
                    f"{base}: составная ссылка ({', '.join(src_fields)}) → "
                    f"{ref_s}({', '.join(map(str, ref_fields))})",
                )
            for c in ref_fields:
                add(ref_s, c, f"{base}: справочник составного ключа для {src_s}")
        elif rtype == "unique":
            fields = r.get("fields") or r.get("unique_fields") or r.get("columns") or []
            if isinstance(fields, dict):
                fields = list(fields.keys())
            scope = r.get("unique_scope_conditions") or []
            scope_txt = ""
            if scope:
                scope_txt = " при условии " + "; ".join(
                    f"{c.get('column')}{c.get('op', '=')}{c.get('value')}"
                    for c in scope
                    if isinstance(c, dict)
                )
            for c in fields:
                add(sheet, c, f"{base}: уникальность по ({', '.join(fields)}){scope_txt}")
        elif rtype == "field_length":
            fields = r.get("fields") or {}
            if isinstance(fields, dict):
                for col, spec in fields.items():
                    lim = spec.get("limit")
                    op = spec.get("operator", "<=")
                    add(sheet, col, f"{base}: длина {col} {op} {lim}")
        elif rtype == "field_format":
            col = r.get("field") or r.get("column")
            fmt = r.get("format") or {}
            if isinstance(fmt, dict):
                fmt_s = (
                    f"type={fmt.get('type')}, date_format={fmt.get('date_format')}, "
                    f"pattern={fmt.get('pattern')}, allow_empty={fmt.get('allow_empty')}"
                )
            else:
                fmt_s = str(fmt)
            if col:
                add(sheet, col, f"{base}: формат {fmt_s}")
        elif rtype == "field_in_values":
            col = r.get("field") or r.get("column")
            vals = r.get("allowed_values") or []
            jkey = r.get("json_key") or r.get("json_path") or ""
            allow_empty = r.get("allow_empty", True)
            if json_col and jkey:
                txt = f"{base}: {json_col}.{jkey} ∈ {vals}; allow_empty={allow_empty}"
                add(sheet, json_col, txt)
                add(sheet, leaf_name(jkey), txt)
            elif col:
                add(sheet, col, f"{base}: {col} ∈ {vals}; allow_empty={allow_empty}")
        elif rtype == "json_spod_format":
            req = "обязательно" if r.get("json_required") else "если заполнено"
            add(sheet, json_col, f"{base}: ячейка {json_col} — валидный SPOD-JSON ({req})")
            for nk in r.get("numeric_value_keys") or []:
                add(sheet, nk, f"{base}: ключ {nk} в {json_col} — числовое значение SPOD-JSON")
        elif rtype == "json_field_equals_column":
            jkey = r.get("json_key") or r.get("json_field") or "parentRewardCode"
            cmp_col = r.get("column_compare") or r.get("compare_column") or "REWARD_CODE"
            op = "≠" if r.get("must_not_equal") else "="
            bits = [f"{json_col}.{jkey} {op} {cmp_col}"]
            if r.get("filter_column"):
                bits.append(f"{r['filter_column']}={r.get('filter_value')}")
            if r.get("json_filter_key"):
                bits.append(f"{json_col}.{r['json_filter_key']}={r.get('json_filter_value')}")
            txt = f"{base}: " + "; ".join(bits)
            add(sheet, json_col, txt)
            add(sheet, leaf_name(str(jkey)), txt)
            add(sheet, cmp_col, txt + " (колонка сравнения)")
            if r.get("json_filter_key"):
                add(sheet, str(r["json_filter_key"]), txt + " (фильтр JSON)")
        elif rtype == "json_field_in_column":
            jkey = r.get("json_key") or r.get("json_field") or "parentRewardCode"
            ref_col = r.get("column_in_sheet") or r.get("column") or "REWARD_CODE"
            txt = f"{base}: уникальные {json_col}.{jkey} ⊆ {sheet}.{ref_col}"
            add(sheet, json_col, txt)
            add(sheet, leaf_name(str(jkey)), txt)
            add(sheet, ref_col, txt + " (справочник)")
        elif rtype == "json_priority_unique_per_contest_link":
            jkey = r.get("json_key") or "priority"
            link = r.get("link_sheet") or "REWARD-LINK"
            txt = (
                f"{base}: через {link} для каждого CONTEST_CODE у связанных наград "
                f"поле {jkey} в {json_col} либо отсутствует у всех, либо задано у всех и уникально"
            )
            add(sheet, json_col, txt)
            add(sheet, leaf_name(str(jkey)), txt)
            add(link, r.get("link_contest_column") or "CONTEST_CODE", txt + " (связь)")
            add(link, r.get("link_reward_column") or "REWARD_CODE", txt + " (связь)")
        else:
            col = r.get("field") or r.get("column") or json_col
            if sheet and col:
                add(sheet, col, base)

    return idx


def consistency_for_param(
    idx: Dict[Tuple[str, str], List[str]],
    table: str,
    column: str,
    json_path: str,
    is_json: bool,
) -> str:
    lines: List[str] = []
    seen: Set[str] = set()

    def push(items: List[str]) -> None:
        for t in items:
            if t not in seen:
                seen.add(t)
                lines.append(t)

    push(idx.get((table, column), []))
    if is_json and json_path not in ("-", ""):
        push(idx.get((table, json_path), []))
        leaf = leaf_name(json_path)
        if leaf:
            push(idx.get((table, leaf), []))
        # путь без префикса колонки
        if json_path.startswith(column + "."):
            rel = json_path[len(column) + 1 :]
            push(idx.get((table, rel), []))
            push(idx.get((table, leaf_name(rel)), []))
    return "\n".join(lines) if lines else "-"


def description_for_param(
    table: str,
    column: str,
    json_path: str,
    is_json: bool,
    glossary: Dict[str, str],
) -> str:
    if not is_json or json_path in ("-", ""):
        return COLUMN_DESCRIPTIONS.get(table, {}).get(column, "Колонка выгрузки SPOD (описание по данным/коду).")
    leaf = leaf_name(json_path)
    # прямое совпадение
    if leaf in glossary:
        return glossary[leaf]
    # относительный путь
    rel = json_path
    if rel.startswith(column + "."):
        rel = rel[len(column) + 1 :]
    if rel in glossary:
        return glossary[rel]
    # getCondition.* и т.п.
    for gkey, gval in glossary.items():
        if rel.endswith(gkey) or leaf == gkey.split(".")[-1]:
            if gkey == leaf or gkey.endswith("." + leaf) or leaf == gkey:
                return gval
    # эвристика по имени
    hints = {
        "businessBlock": "Коды бизнес-блока, к которым относится сущность; влияет на видимость и отбор аудитории.",
        "feature": (
            "Тексты особенностей (турнир — в детальной карточке; награда — в Награде)."
        ),
        "hidden": "Флаг скрытия сущности в интерфейсе (Y/N).",
        "nftFlg": "Признак NFT-награды (Y/N).",
        "outstanding": "Признак «выдающейся» награды для ленты/акцентов (Y/N).",
        "rewardRule": "Текстовое правило/условие получения награды для пользователя.",
        "seasonItem": "Коды сезонов, к которым привязана награда/предмет.",
        "getCondition": "Объект условий выдачи ITEM (рейтинг, нужные/запрещённые награды).",
        "priority": "Приоритет отображения/ранжирования бейджа в рамках конкурса.",
        "parentRewardCode": "Код родительской награды (иерархия BADGE/LABEL).",
        "masterBadge": "Мастер-бейдж: Y — для награды / N — для турнира.",
        "vid": "Контур/вид промо (ПРОМ / ТЕСТ).",
        "typeRewarding": "Вручаем одну из 3 наград или все (one | all).",
        "momentRewarding": (
            "Момент награждения: AFTER — после закрытия турнира / "
            "DURIN — во время турнира."
        ),
        "masking": "Маскирование данных в отображении (Y/N).",
        "minNumber": (
            "Мин. число участников чтобы считать победителей "
            "(исключаем соревнование сам с собой): 1 | 2 | 3."
        ),
        "accuracy": "Число знаков после запятой для отображения (0 | 1 | 2 | 3 | 5).",
        "capacity": (
            "Масштаб отображения: пусто | MILLIONS | THOUSANDS "
            "(уменьшение показателя до млн/тыс.)."
        ),
        "tournamentStartMailing": "Флаг рассылки при старте турнира.",
        "tournamentEndMailing": "Флаг рассылки при окончании турнира.",
        "tournamentLikeMailing": "Флаг «лайк»-рассылки турнира.",
        "tournamentRewardingMailing": "Флаг рассылки о награждении.",
        "persomanNumberVisible": (
            "Если указаны табельные — только эти сотрудники увидят турнир."
        ),
        "persomanNumberHidden": (
            "Если указаны табельные — эти сотрудники НЕ увидят турнир."
        ),
        "tbVisible": "Коды ТБ с видимостью.",
        "tbHidden": "Коды ТБ со скрытием.",
        "gosbVisible": "Коды ГОСБ с видимостью.",
        "gosbHidden": "Коды ГОСБ со скрытием.",
        "helpCodeList": "Коды для вывода окна с доп. описанием конкурса.",
        "preferences": "Преференции за получение награды, если предусмотрены.",
        "newsType": "Тип новости: AIPROMPT (ИИ) | TEMPLATE (по шаблону).",
        "winCriterion": "Текст критерия победы для ИИ создания новости.",
        "itemAmount": "Доступное количество предметов на складе/лимите.",
        "itemLimitCount": "Лимит заказов предмета.",
        "itemLimitPeriod": "Период лимита (например once).",
        "bookingRequired": "Требуется бронирование для получения ITEM.",
        "deliveryRequired": "Требуется доставка для ITEM.",
        "commingSoon": "Признак «скоро» (coming soon; опечатка в имени).",
        "tagColor": "Цветовой токен метки LABEL.",
        "tagEndDT": "Дата окончания действия метки LABEL.",
        "period_code": "Код периода внутри CONTEST_PERIOD.",
        "criterion_mark_value": "Значение отметки критерия в JSON периода.",
    }
    if leaf in hints:
        return "Предположение: " + hints[leaf]
    return (
        f"JSON-ключ `{leaf}` в колонке `{column}` таблицы `{table}`. "
        "Смысл восстановлен по имени и встречаемости в данных; уточняется по ТЗ СПОД."
    )


def analyze_file(
    path: Path,
    table: str,
    consistency_idx: Dict[Tuple[str, str], List[str]],
    glossary: Dict[str, str],
) -> List[Dict[str, Any]]:
    with path.open(encoding=ENC, newline="") as fh:
        reader = csv.DictReader(fh, delimiter=DELIM)
        columns = list(reader.fieldnames or [])
        rows: List[Dict[str, str]] = []
        for raw in reader:
            rows.append({k: (raw.get(k) if raw.get(k) is not None else "") for k in columns})

    total = len(rows)
    json_cols = detect_json_columns(rows, columns)
    disc_names = DEPENDENCY_DISCRIMINATORS.get(table, [])
    disc_totals: Dict[str, Counter] = {d: Counter() for d in disc_names}
    for row in rows:
        for d in disc_names:
            val = (row.get(d) or "").strip()
            if val:
                disc_totals[d][val] += 1

    # плоские колонки
    flat_values: Dict[str, List[str]] = {c: [] for c in columns}
    for row in rows:
        for c in columns:
            flat_values[c].append(row.get(c) or "")

    # JSON accumulators per column
    json_acc_by_col: Dict[str, Dict[str, PathAccumulator]] = {c: {} for c in json_cols}
    json_ok_by_col: Dict[str, int] = {c: 0 for c in json_cols}

    for row_idx, row in enumerate(rows):
        dep_ctx = {d: (row.get(d) or "").strip() for d in disc_names}
        for col in json_cols:
            raw = row.get(col) or ""
            obj, ok = try_parse_json(raw)
            if not ok:
                continue
            json_ok_by_col[col] += 1
            walk_json_leaves(obj, "", col, table, json_acc_by_col[col], dep_ctx, row_idx)

    # сбор строк результата: только КОЛОНКИ и конечные JSON-ключи
    out_rows: List[Dict[str, Any]] = []

    for col in columns:
        is_json = col in json_cols
        vals = flat_values[col]
        nonempty = [v for v in vals if str(v).strip() != ""]
        empty_n = total - len(nonempty)
        uniq = len(set(nonempty))
        flat_counter: Counter = Counter()
        for v in nonempty:
            key = v if len(v) <= 200 else v[:197] + "…"
            if len(flat_counter) < 5000 or key in flat_counter:
                flat_counter[key] += 1

        examples: List[str] = []
        seen_ex: Set[str] = set()
        for v in nonempty:
            if is_json:
                obj, ok = try_parse_json(v)
                if ok:
                    try:
                        s = json.dumps(obj, ensure_ascii=False, separators=(",", ":"))
                    except TypeError:
                        s = str(obj)
                else:
                    s = v
            else:
                s = v
            if len(s) > 300:
                s = s[:297] + "…"
            if s not in seen_ex:
                seen_ex.add(s)
                examples.append(s)
            if len(examples) >= 3:
                break
        while len(examples) < 3:
            examples.append("-")

        dtype = infer_flat_column_type(vals)
        if is_json:
            root_kinds: Counter = Counter()
            for v in nonempty[:500]:
                obj, ok = try_parse_json(v)
                if ok:
                    root_kinds[type_label_scalar(obj)] += 1
            if root_kinds:
                top = root_kinds.most_common(1)[0][0]
                if top == "массив":
                    dtype = "JSON-массив (в ячейке)"
                elif top == "объект":
                    dtype = "JSON-объект (в ячейке)"
                else:
                    dtype = f"JSON ({top})"
            else:
                dtype = "JSON (не удалось массово распарсить)"
        param_id = make_param_id(table, col, "-", is_json)
        desc = description_for_param(table, col, "-", False, glossary)
        if is_json:
            desc = COLUMN_DESCRIPTIONS.get(table, {}).get(
                col,
                "Колонка содержит JSON; конечные ключи — отдельными строками каталога.",
            )

        cons = consistency_for_param(consistency_idx, table, col, "-", is_json)

        if is_json and disc_names:
            deps_col = "Колонка JSON.\nНабор вложенных ключей зависит от:\n" + "\n".join(
                f"  {d}" for d in disc_names
            )
        elif is_json:
            deps_col = "Колонка JSON.\nНабор вложенных ключей зависит от контекста строки."
        else:
            deps_col = "-"

        occ = format_occurrence_stats(
            total=total,
            filled=len(nonempty),
            empty=empty_n,
            unique_n=uniq,
            value_counter=flat_counter,
            json_ok=json_ok_by_col.get(col) if is_json else None,
            is_json_key=False,
        )

        out_rows.append(
            {
                "table": table,
                "param_kind": "КОЛОНКА",
                "csv_column": col,
                "col_type": "JSON" if is_json else "-",
                "json_path": "-",
                "name": col,
                "data_type": dtype,
                "deps": deps_col,
                "description": desc,
                "param_id": param_id,
                "dup_hint": "",
                "ex1": examples[0],
                "ex2": examples[1],
                "ex3": examples[2],
                "consistency": cons,
                "rows_total": str(total),
                "filled": str(len(nonempty)),
                "empty": str(empty_n),
                "fill_pct": f"{(100.0 * len(nonempty) / total) if total else 0:.1f}%",
                "uniq": str(uniq),
                "diversity": diversity_ratio(uniq, len(nonempty)),
                "occurrence": occ,
                "source": path.name,
                "_sort": (table, col, "", 0),
            }
        )

        if not is_json:
            continue

        json_ok = json_ok_by_col.get(col, 0)
        acc_map = json_acc_by_col[col]
        paths_set = set(acc_map.keys())
        skip_paths = {
            p
            for p in paths_set
            if any(ch == f"{p}[]" or ch.startswith(f"{p}[].") for ch in paths_set)
        }
        for rel in sorted(acc_map.keys(), key=lambda p: (p.count("."), p.count("[]"), p)):
            if rel in skip_paths:
                continue
            node = acc_map[rel]
            full_path = display_full_path(rel)
            nm = leaf_key_name(rel)
            if not nm or nm == "[]":
                continue
            dtype_j = dominant_type(node)
            # массивы скаляров — тип «массив с …»
            if "массив" in node.types:
                dtype_j = dominant_type(node)

            deps = format_dependencies(node, disc_totals, True)
            desc_j = description_for_param(table, col, rel, True, glossary)
            pid = make_param_id(table, col, rel, True)
            cons_j = consistency_for_param(consistency_idx, table, col, rel, True)

            exs = list(node.examples)
            while len(exs) < 3:
                exs.append("-")

            filled_k = len(node.touched_rows) if node.touched_rows else node.present_rows
            key_absent = max(0, json_ok - filled_k) if filled_k <= json_ok else 0
            empty_est = (total - json_ok) + key_absent
            uniq_k = len(node.value_counter)
            occ_k = format_occurrence_stats(
                total=total,
                filled=filled_k,
                empty=empty_est,
                unique_n=uniq_k,
                value_counter=node.value_counter,
                empty_string=node.empty_string_count,
                null_n=node.null_count,
                empty_array=node.empty_array_count,
                json_ok=json_ok,
                key_absent_in_json=key_absent,
                is_json_key=True,
            )

            out_rows.append(
                {
                    "table": table,
                    "param_kind": "JSON-КЛЮЧ",
                    "csv_column": col,
                    "col_type": "JSON",
                    "json_path": full_path,
                    "name": nm,
                    "data_type": dtype_j,
                    "deps": deps,
                    "description": desc_j,
                    "param_id": pid,
                    "dup_hint": "",
                    "ex1": exs[0],
                    "ex2": exs[1],
                    "ex3": exs[2],
                    "consistency": cons_j,
                    "rows_total": str(total),
                    "filled": str(filled_k),
                    "empty": str(empty_est),
                    "fill_pct": f"{(100.0 * filled_k / total) if total else 0:.1f}%",
                    "uniq": str(uniq_k),
                    "diversity": diversity_ratio(uniq_k, filled_k),
                    "occurrence": occ_k,
                    "source": path.name,
                    "_sort": (table, col, rel, 1),
                }
            )

    return out_rows



def mark_duplicates(all_rows: List[Dict[str, Any]]) -> None:
    """Проставляет признак дублей по имени / полному пути между таблицами."""
    by_name: Dict[str, List[str]] = defaultdict(list)
    by_path: Dict[str, List[str]] = defaultdict(list)
    for r in all_rows:
        label = f"{r['table']}.{r.get('csv_column', r['name'])}"
        if r.get("json_path") and r["json_path"] != "-":
            label += f"::{r['json_path']}"
        by_name[r["name"]].append(label)
        if r.get("json_path") and r["json_path"] != "-":
            by_path[r["json_path"]].append(label)

    by_leaf_tables: Dict[str, Set[str]] = defaultdict(set)
    for r in all_rows:
        by_leaf_tables[r["name"]].add(r["table"])

    for r in all_rows:
        tables = sorted(by_leaf_tables.get(r["name"], set()))
        same = by_name.get(r["name"], [])
        others = [x for x in same if not x.startswith(r["table"] + ".")]
        bits: List[str] = []
        if len(tables) > 1:
            bits.append(
                "имя встречается в таблицах: " + ", ".join(tables)
            )
        if others:
            bits.append("примеры: " + "; ".join(others[:6]))
        jp = r.get("json_path") or "-"
        if jp != "-" and len(by_path.get(jp, [])) > 1:
            bits.append(f"тот же путь `{jp}` в нескольких местах")
        r["dup_hint"] = (
            "Да (предположение): " + "; ".join(bits)
            if bits
            else "Нет (уникальное имя в каталоге)"
        )


def write_excel(rows: List[Dict[str, Any]], out_path: Path, meta: Dict[str, str]) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # --- Лист PARAMETERS ---
    ws = wb.active
    ws.title = "PARAMETERS"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    # данные: по высоте по центру; заголовок — по высоте и по ширине
    data_align = Alignment(wrap_text=True, vertical="center", horizontal="left")
    header_align = Alignment(wrap_text=True, vertical="center", horizontal="center")

    for col_idx, h in enumerate(EXCEL_HEADERS, start=1):
        cell = ws.cell(1, col_idx, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    key_order = [
        "table", "param_kind", "csv_column", "col_type", "json_path", "name",
        "data_type", "deps", "description", "param_id", "dup_hint",
        "ex1", "ex2", "ex3", "consistency",
        "rows_total", "filled", "empty", "fill_pct",
        "uniq", "diversity", "occurrence", "source",
        "excel_type", "excel_limits", "excel_align", "excel_color", "excel_width",
    ]
    for r_idx, r in enumerate(rows, start=2):
        for c_idx, k in enumerate(key_order, start=1):
            val = r.get(k, "")
            cell = ws.cell(r_idx, c_idx, val)
            cell.alignment = data_align
            cell.border = thin
            if k == "col_type" and val == "JSON":
                cell.fill = PatternFill("solid", fgColor="FFF2CC")
            if k == "param_kind" and val == "JSON-КЛЮЧ":
                cell.fill = PatternFill("solid", fgColor="E2EFDA")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(EXCEL_HEADERS))}{len(rows) + 1}"
    widths = [
        16, 12, 22, 12, 40, 24, 28, 42, 48, 36, 36,
        28, 28, 28, 48,
        12, 14, 14, 12, 12, 14, 42, 36,
        12, 28, 28, 42, 18,
    ]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 36

    # --- Лист META ---
    ws2 = wb.create_sheet("META")
    ws2["A1"] = "Параметр"
    ws2["B1"] = "Значение"
    ws2["A1"].font = header_font
    ws2["B1"].font = header_font
    ws2["A1"].fill = header_fill
    ws2["B1"].fill = header_fill
    ws2["A1"].alignment = header_align
    ws2["B1"].alignment = header_align
    meta_rows = [
        ("generated_at", meta.get("generated_at", "")),
        ("input_dir", meta.get("input_dir", "")),
        ("files", meta.get("files", "")),
        ("total_parameters", str(len(rows))),
        (
            "note",
            "Полный обход всех строк CSV; JSON после замены \"\"\" → \". "
            "Описания — из глоссариев/эвристик. "
            "Колонки Excel:* — из config/CONFIG_FORMATS.json (родной лист; ширина только per-column).",
        ),
    ]
    for i, (a, b) in enumerate(meta_rows, start=2):
        c1 = ws2.cell(i, 1, a)
        c2 = ws2.cell(i, 2, b)
        c1.alignment = data_align
        c2.alignment = data_align
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 100

    # --- Лист TABLES ---
    ws3 = wb.create_sheet("TABLES")
    for col_idx, h in enumerate(
        ["Таблица", "Файл", "Строк данных", "Колонок", "Параметров в каталоге"], start=1
    ):
        cell = ws3.cell(1, col_idx, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    by_table: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for r in rows:
        by_table[r["table"]].append(r)
    t_idx = 2
    for table in sorted(by_table.keys()):
        trs = by_table[table]
        src = trs[0]["source"]
        total_rows = trs[0].get("rows_total", "")
        n_cols = sum(1 for r in trs if "__COL__" in r["param_id"])
        for c_idx, val in enumerate(
            [table, src, total_rows, n_cols, len(trs)], start=1
        ):
            cell = ws3.cell(t_idx, c_idx, val)
            cell.alignment = data_align
        t_idx += 1
    for i, w in enumerate([20, 48, 14, 12, 18], start=1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    wb.save(out_path)


def load_all_glossaries() -> Dict[str, str]:
    merged: Dict[str, str] = {}
    if not GLOSSARY_DIR.is_dir():
        return merged
    for p in sorted(GLOSSARY_DIR.glob("*.md")):
        if p.name.upper().startswith("README"):
            continue
        merged.update(parse_glossary_meanings(p.read_text(encoding="utf-8")))
    return merged


def list_input_files(input_dir: Path) -> List[Tuple[Path, str]]:
    found: List[Tuple[Path, str]] = []
    for p in sorted(input_dir.glob("*.csv")):
        table = resolve_table_name(p.name)
        if not table:
            print(f"SKIP (неизвестная таблица): {p.name}")
            continue
        found.append((p, table))
    # Одна таблица — один файл (новейший по mtime); USER_ROLE и USER_ROLE SB — разные таблицы
    by_table: Dict[str, Path] = {}
    for p, table in found:
        prev = by_table.get(table)
        if prev is None or p.stat().st_mtime >= prev.stat().st_mtime:
            by_table[table] = p
    return [(path, table) for table, path in sorted(by_table.items(), key=lambda x: x[0])]


def main() -> int:
    parser = argparse.ArgumentParser(description="Каталог параметров SPOD → Excel")
    parser.add_argument("--input-dir", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT)
    parser.add_argument("--checks", type=Path, default=CHECKS_PATH)
    parser.add_argument("--formats", type=Path, default=FORMATS_PATH)
    args = parser.parse_args()

    if not args.input_dir.is_dir():
        print(f"Нет каталога входных CSV: {args.input_dir}")
        return 1

    glossary = load_all_glossaries()
    consistency_idx = load_consistency_index(args.checks)
    formats_cfg = load_formats_config(args.formats)
    files = list_input_files(args.input_dir)
    if not files:
        print("CSV не найдены")
        return 1

    all_rows: List[Dict[str, Any]] = []
    for path, table in files:
        print(f"Анализ {path.name} → {table} …")
        part = analyze_file(path, table, consistency_idx, glossary)
        print(f"  параметров: {len(part)}")
        all_rows.extend(part)

    mark_duplicates(all_rows)
    apply_excel_format_fields(all_rows, formats_cfg)
    all_rows.sort(key=lambda r: r.get("_sort", (r["table"], r["name"], r["json_path"], 0)))

    meta = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "input_dir": str(args.input_dir),
        "files": "; ".join(f"{t}:{p.name}" for p, t in files),
    }
    write_excel(all_rows, args.out, meta)
    print(f"Готово: {args.out} ({len(all_rows)} параметров)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
